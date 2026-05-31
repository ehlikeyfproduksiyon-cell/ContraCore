# -*- coding: utf-8 -*-
"""
E-YMM Karşıt İnceleme Otomasyonu — Production GUI
ContraCore embedded widget; KarsitWorker sinyalleri tam bağlı.
"""
import json
import os
import shutil
import sys
from datetime import datetime

from PySide6.QtCore import Qt, QTimer, QPoint, QSize, QRectF, Signal, QEvent, QObject
from PySide6.QtGui  import (QColor, QFont, QIcon, QLinearGradient, QBrush,
                             QPainter, QPainterPath, QPen, QPixmap, QFontMetrics)
from PySide6.QtWidgets import (
    QApplication, QComboBox, QDialog, QFileDialog, QFrame, QGraphicsDropShadowEffect,
    QGraphicsOpacityEffect,
    QGridLayout, QHBoxLayout, QLabel, QLineEdit, QMainWindow, QMessageBox,
    QPlainTextEdit, QProgressBar, QPushButton, QScrollArea, QVBoxLayout, QWidget,
)

from .karsit_worker import KarsitWorker
from .karsit_import import excel_oku, kaydet_db
from .karsit_db     import init_db

# ── Config ────────────────────────────────────────────────────────────────────
_CFG_PATH = os.path.join(
    os.environ.get('APPDATA', os.path.expanduser('~')),
    'ContraCore', 'karsit_ymm_config.json'
)

# Master Excel şablon kaynağı
_MASTER_XLSM_SRC = os.path.normpath(os.path.join(
    os.path.dirname(__file__), '..', '..', 'modules', 'e-ymm',
    'karsit_otomasyon', 'karsit_master.xlsm'
))

# Proje genelinde kullanılan ikon dizini
_ICON_DIR     = os.path.normpath(os.path.join(os.path.dirname(__file__), '..', '..', 'Icon'))
_PROFILE_PNG  = os.path.join(_ICON_DIR, 'profile.png')


def _cfg_load() -> dict:
    try:
        with open(_CFG_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _cfg_save(data: dict):
    try:
        os.makedirs(os.path.dirname(_CFG_PATH), exist_ok=True)
        with open(_CFG_PATH, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
    except Exception:
        pass

# ── Son Kullanılan Yollar ─────────────────────────────────────────────────────
_RECENT_FILE = os.path.join(
    os.environ.get('APPDATA', os.path.expanduser('~')),
    'ContraCore', 'karsit_ymm_recent.json'
)

def _recent_load() -> dict:
    try:
        with open(_RECENT_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def _recent_add(tag: str, path: str):
    d = _recent_load()
    lst = d.get(tag, [])
    if path in lst:
        lst.remove(path)
    lst.insert(0, path)
    d[tag] = lst[:5]
    try:
        os.makedirs(os.path.dirname(_RECENT_FILE), exist_ok=True)
        with open(_RECENT_FILE, 'w', encoding='utf-8') as f:
            json.dump(d, f, ensure_ascii=False)
    except Exception:
        pass

# ── Scroll bloklayıcı — QComboBox'larda scroll wheel etkisiz ─────────────────
class _NoScrollFilter(QObject):
    def eventFilter(self, obj, event):
        if event.type() == QEvent.Wheel:
            # Combo seçeneğini değiştirme — ama scroll area'yı kaydır
            parent = obj.parent()
            while parent:
                from PySide6.QtWidgets import QScrollArea, QAbstractScrollArea
                if isinstance(parent, (QScrollArea, QAbstractScrollArea)):
                    parent.wheelEvent(event)
                    break
                parent = parent.parent()
            return True  # combo'nun kendi scroll'unu yut
        return False

_NO_SCROLL = _NoScrollFilter()

def _cmb_no_scroll(c: 'QComboBox'):
    """Verilen combo'ya scroll bloğu uygular."""
    c.setFocusPolicy(Qt.StrongFocus)
    c.installEventFilter(_NO_SCROLL)
    return c

# ── Renk Paleti ───────────────────────────────────────────────────────────────
BG      = '#F2F4F7'
CARD    = '#FFFFFF'
NAVY    = '#0B1F3A'
NAVY2   = '#162D4E'
GREEN   = '#22C55E'
GREEN2  = '#16A34A'
RED     = '#EF4444'
GOLD    = '#C9A46A'
BORDER  = '#E5E7EB'
TEXT    = '#111827'
TEXT2   = '#6B7280'
TEXT3   = '#9CA3AF'
BLUE    = '#3B82F6'
BLUE2   = '#2563EB'
BLUE_BG = '#EFF6FF'
LOG_BG  = '#0F172A'
ORANGE  = '#F97316'

# ── Dark Mode Paleti ──────────────────────────────────────────────────────────
_DARK_BG      = '#0F172A'
_DARK_CARD    = '#1E293B'
_DARK_BORDER  = '#334155'
_DARK_TEXT    = '#F1F5F9'
_DARK_TEXT2   = '#94A3B8'
_DARK_TEXT3   = '#64748B'
_DARK_BLUE_BG = '#1E3A5F'
_DARK_INPUT   = '#0F1E33'

_dark_mode: bool = False   # Global tema durumu

def _dark_app_qss() -> str:
    return f'''
    QScrollArea {{ background: {_DARK_BG}; border: none; }}
    QScrollBar:vertical {{ background: {_DARK_BG}; width: 6px; border: none; }}
    QScrollBar::handle:vertical {{ background: #334155; border-radius: 3px; min-height: 30px; }}
    QScrollBar::handle:vertical:hover {{ background: #475569; }}
    QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height: 0; }}
    QToolTip {{
        background: {_DARK_CARD}; color: {_DARK_TEXT};
        border: 1px solid {_DARK_BORDER}; border-radius: 6px; padding: 4px 8px;
    }}
    '''


def _cc_pix(sz: int = 24, dark: bool = False) -> QPixmap:
    font = QFont('Coolvetica', sz, QFont.Bold)
    fm   = QFontMetrics(font)
    cw   = fm.horizontalAdvance('Contra')
    ow   = fm.horizontalAdvance('CORE')
    h    = fm.height() + 4
    pix  = QPixmap(cw + ow + 4, h)
    pix.fill(Qt.transparent)
    p    = QPainter(pix)
    p.setRenderHint(QPainter.Antialiasing)
    pa1  = QPainterPath(); pa1.addText(2, fm.ascent() + 2, font, 'Contra')
    g1   = QLinearGradient(0, 2, 0, h - 2)
    if dark:
        g1.setColorAt(0, QColor('#FFFFFF')); g1.setColorAt(1, QColor('#E2E8F0'))
    else:
        g1.setColorAt(0, QColor('#0a1e43')); g1.setColorAt(1, QColor('#081631'))
    p.setBrush(QBrush(g1)); p.setPen(Qt.NoPen); p.drawPath(pa1)
    pa2  = QPainterPath(); pa2.addText(2 + cw, fm.ascent() + 2, font, 'CORE')
    g2   = QLinearGradient(0, 2, 0, h - 2)
    g2.setColorAt(0, QColor('#c8a45b')); g2.setColorAt(1, QColor('#96732d'))
    p.setBrush(QBrush(g2)); p.drawPath(pa2)
    p.end()
    return pix


def _icon_pix(filename: str, size: int = 16) -> QPixmap:
    """Icon klasöründen PNG yükler, belirtilen boyuta ölçekler."""
    path = os.path.join(_ICON_DIR, filename)
    pix  = QPixmap(path)
    if pix.isNull():
        return pix
    return pix.scaled(size, size, Qt.KeepAspectRatio, Qt.SmoothTransformation)


def _icon_lbl(filename: str, size: int = 16) -> QLabel:
    """Küçük ikon QLabel döner; PNG yoksa boş QLabel."""
    lbl = QLabel()
    lbl.setStyleSheet('background:transparent;border:none;')
    pix = _icon_pix(filename, size)
    if not pix.isNull():
        lbl.setPixmap(pix)
        lbl.setFixedSize(size, size)
    return lbl


def _section_hdr(text: str, icon_file: str = '') -> QWidget:
    """Opsiyonel ikon + başlık metnini yan yana gösteren section header widget."""
    w   = QWidget()
    w.setStyleSheet('background:transparent;')
    lay = QHBoxLayout(w)
    lay.setContentsMargins(0, 4, 0, 4)
    lay.setSpacing(8)
    if icon_file:
        lay.addWidget(_icon_lbl(icon_file, 18))
    lbl = QLabel(text)
    lbl.setStyleSheet(
        f'font-size:15px;font-weight:800;color:{NAVY};'
        f'background:transparent;border:none;letter-spacing:1px;'
    )
    lay.addWidget(lbl)
    lay.addStretch()
    # Alt çizgi
    sep = QFrame()
    sep.setFrameShape(QFrame.HLine)
    sep.setStyleSheet(f'background:{NAVY};border:none;max-height:2px;opacity:0.2;')
    container = QWidget()
    container.setStyleSheet('background:transparent;')
    cl = QVBoxLayout(container)
    cl.setContentsMargins(0, 0, 0, 0)
    cl.setSpacing(4)
    cl.addWidget(w)
    cl.addWidget(sep)
    return container


# ── AccordionSection — katlanabilir bölüm ────────────────────────────────────
class AccordionSection(QFrame):
    def __init__(self, title: str, icon_file: str = '',
                 initial_open: bool = True, accent: str = NAVY, parent=None):
        super().__init__(parent)
        self._open = initial_open
        self._accent = accent
        self.setStyleSheet('QFrame{background:transparent;border:none;}')

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(0)

        # Tıklanabilir başlık
        self._hdr = QFrame()
        self._hdr.setCursor(Qt.PointingHandCursor)
        self._hdr.setFixedHeight(36)
        self._hdr.setStyleSheet(f'''
            QFrame {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {accent}18, stop:1 transparent);
                border-radius: 8px;
                border-left: 3px solid {accent};
            }}
            QFrame:hover {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {accent}28, stop:1 transparent);
            }}
        ''')
        hl = QHBoxLayout(self._hdr)
        hl.setContentsMargins(10, 0, 10, 0)
        hl.setSpacing(8)

        self._arrow = QLabel('▼' if initial_open else '▶')
        self._arrow.setStyleSheet(
            f'font-size:10px;color:{accent};background:transparent;border:none;'
        )
        hl.addWidget(self._arrow)

        if icon_file:
            hl.addWidget(_icon_lbl(icon_file, 14))

        title_lbl = QLabel(title)
        title_lbl.setStyleSheet(
            f'font-size:12px;font-weight:800;color:{accent};'
            f'background:transparent;border:none;letter-spacing:0.5px;'
        )
        hl.addWidget(title_lbl)
        hl.addStretch()
        root.addWidget(self._hdr)

        # İçerik alanı
        self._body = QFrame()
        self._body.setStyleSheet('QFrame{background:transparent;border:none;}')
        self._body_lay = QVBoxLayout(self._body)
        self._body_lay.setContentsMargins(0, 8, 0, 4)
        self._body_lay.setSpacing(8)
        self._body.setVisible(initial_open)
        root.addWidget(self._body)

        self._hdr.mousePressEvent = lambda e: self.toggle()

    def toggle(self):
        self._open = not self._open
        self._body.setVisible(self._open)
        self._arrow.setText('▼' if self._open else '▶')
        if self._open:
            from core import theme as _t
            _t.apply_to_widget(self._body)
            for inp in self._body.findChildren(InputRow):
                inp._apply_normal_style()

    def layout(self) -> QVBoxLayout:
        return self._body_lay

    def add_widget(self, w):
        self._body_lay.addWidget(w)

    def add_layout(self, lay):
        self._body_lay.addLayout(lay)

    def set_open(self, v: bool):
        if v != self._open:
            self.toggle()

    def showEvent(self, event):
        super().showEvent(event)
        from core import theme as _t
        _t.apply_to_widget(self)
        for inp in self.findChildren(InputRow):
            inp._apply_normal_style()


# ── RecentPopup — son kullanılan yollar menüsü ────────────────────────────────
class RecentPopup(QFrame):
    picked = Signal(str)

    def __init__(self, tag: str, label: str = 'Son Kullanılan', parent=None):
        super().__init__(parent, Qt.Popup | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self._tag   = tag
        self._label = label
        self.setFixedWidth(320)
        self._lay = QVBoxLayout(self)
        self._lay.setContentsMargins(0, 0, 0, 0)

    def show_at(self, widget):
        self._refresh()
        self.adjustSize()
        p = widget.mapToGlobal(QPoint(0, widget.height() + 4))
        screen = QApplication.screenAt(p) or QApplication.primaryScreen()
        sg = screen.availableGeometry()
        x = max(sg.left(), min(p.x(), sg.right()  - self.width()))
        y = max(sg.top(),  min(p.y(), sg.bottom() - self.height()))
        self.move(x, y)
        self.show()

    def _refresh(self):
        while self._lay.count():
            item = self._lay.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        card = QFrame()
        card.setStyleSheet(
            f'QFrame{{background:{CARD};border-radius:12px;border:1px solid {BORDER};}}'
        )
        cl = QVBoxLayout(card)
        cl.setContentsMargins(12, 12, 12, 12)
        cl.setSpacing(4)
        ttl = QLabel(self._label)
        ttl.setStyleSheet(
            f'font-size:11px;font-weight:700;color:{TEXT};background:transparent;border:none;'
        )
        cl.addWidget(ttl)
        paths = _recent_load().get(self._tag, [])
        if not paths:
            lbl = QLabel('Henüz seçim yapılmadı.')
            lbl.setStyleSheet(
                f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
            )
            cl.addWidget(lbl)
        else:
            for p in paths:
                b = QPushButton(f'  {os.path.basename(p)}')
                b.setToolTip(p)
                b.setCursor(Qt.PointingHandCursor)
                b.setStyleSheet(f'''
                    QPushButton {{
                        background: #F9FAFB; border: 1px solid {BORDER};
                        border-radius: 8px; color: {TEXT2}; font-size: 11px;
                        padding: 6px 10px; text-align: left;
                    }}
                    QPushButton:hover {{
                        background: {BLUE_BG}; border-color: {NAVY}; color: {TEXT};
                    }}
                ''')
                b.clicked.connect(lambda _, x=p: self._pick(x))
                cl.addWidget(b)
        self._lay.addWidget(card)

    def _pick(self, path: str):
        self.picked.emit(path)
        self.close()


# ── ProfilePopup ──────────────────────────────────────────────────────────────
class ProfilePopup(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent, Qt.Popup | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedWidth(290)
        self._just_closed = False

        outer = QVBoxLayout(self)
        outer.setContentsMargins(6, 6, 6, 6)

        card = QFrame()
        card.setStyleSheet('''
            QFrame {
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #162D4E, stop:1 #0B1F3A);
                border-radius: 16px;
                border: 1px solid #C9A46A;
            }
        ''')
        glow = QGraphicsDropShadowEffect()
        glow.setBlurRadius(28)
        glow.setColor(QColor('#C9A46A44'))
        glow.setOffset(0, 4)
        card.setGraphicsEffect(glow)

        cl = QVBoxLayout(card)
        cl.setContentsMargins(0, 0, 0, 16)
        cl.setSpacing(0)

        header = QFrame()
        header.setFixedHeight(64)
        header.setStyleSheet('''
            QFrame {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1E3660, stop:1 #0B1F3A);
                border-radius: 16px;
                border-bottom-left-radius: 0px;
                border-bottom-right-radius: 0px;
                border: none;
            }
        ''')
        hl = QHBoxLayout(header)
        hl.setContentsMargins(16, 10, 16, 10)
        hl.setSpacing(12)

        av_lbl = QLabel()
        av_lbl.setFixedSize(42, 42)
        av_lbl.setAlignment(Qt.AlignCenter)
        av_lbl.setStyleSheet('background:transparent;border:none;')
        _av_pix = QPixmap(_PROFILE_PNG)
        if not _av_pix.isNull():
            av_lbl.setPixmap(_av_pix.scaled(36, 36, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            av_lbl.setText('👤'); av_lbl.setStyleSheet('background:transparent;border:none;font-size:22px;')
        hl.addWidget(av_lbl)

        name_col = QVBoxLayout()
        name_col.setSpacing(1)
        name_lbl = QLabel('Serkan ŞAHİN')
        name_lbl.setStyleSheet(
            'font-size:13px;font-weight:700;color:#FFFFFF;background:transparent;border:none;'
        )
        role_lbl = QLabel('Yazılım Geliştirici')
        role_lbl.setStyleSheet(f'font-size:10px;color:{GOLD};background:transparent;border:none;')
        name_col.addWidget(name_lbl)
        name_col.addWidget(role_lbl)
        hl.addLayout(name_col)
        hl.addStretch()
        cl.addWidget(header)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFixedHeight(1)
        sep.setStyleSheet('background:#C9A46A44;border:none;max-height:1px;')
        cl.addWidget(sep)
        cl.addSpacing(12)

        for ico, label, val in [
            ('📱', 'WhatsApp', '0531 087 93 39'),
            ('✉️', 'Mail',     'serkan.opsiyonymm@gmail.com'),
        ]:
            row_frame = QFrame()
            row_frame.setStyleSheet('''
                QFrame {
                    background: rgba(255,255,255,0.04);
                    border-radius: 8px;
                    border: 1px solid rgba(201,164,106,0.15);
                }
                QFrame:hover {
                    background: rgba(201,164,106,0.10);
                    border: 1px solid rgba(201,164,106,0.35);
                }
            ''')
            row_lay = QHBoxLayout(row_frame)
            row_lay.setContentsMargins(12, 8, 12, 8)
            row_lay.setSpacing(10)

            ico_lbl = QLabel(ico)
            ico_lbl.setFixedWidth(20)
            ico_lbl.setStyleSheet('background:transparent;border:none;font-size:14px;')

            info_col = QVBoxLayout()
            info_col.setSpacing(0)
            lbl_top = QLabel(label)
            lbl_top.setStyleSheet(
                f'font-size:9px;font-weight:600;color:{GOLD};background:transparent;border:none;'
            )
            lbl_val = QLabel(val)
            lbl_val.setStyleSheet('font-size:11px;color:#E2E8F0;background:transparent;border:none;')
            lbl_val.setTextInteractionFlags(Qt.TextSelectableByMouse)
            info_col.addWidget(lbl_top)
            info_col.addWidget(lbl_val)

            row_lay.addWidget(ico_lbl)
            row_lay.addLayout(info_col)
            row_lay.addStretch()
            cl.addWidget(row_frame)
            cl.addSpacing(6)

        outer.addWidget(card)

    def hideEvent(self, event):
        super().hideEvent(event)
        self._just_closed = True
        QTimer.singleShot(200, lambda: setattr(self, '_just_closed', False))

    def show_at(self, widget):
        p = widget.mapToGlobal(QPoint(0, widget.height() + 4))
        x = p.x() - self.width() + widget.width()
        screen = QApplication.screenAt(p) or QApplication.primaryScreen()
        sg = screen.availableGeometry()
        x = max(sg.left(), min(x, sg.right() - self.width()))
        y = max(sg.top(), min(p.y(), sg.bottom() - self.height()))
        self.move(x, y)
        self.show()


# ── Yaygın Validatorlar ───────────────────────────────────────────────────────
import re as _re

def _val_tarih(v: str):
    """GG.AA.YYYY formatı."""
    if not v:
        return None
    if not _re.fullmatch(r'\d{2}\.\d{2}\.\d{4}', v):
        return 'GG.AA.YYYY formatında girin'
    g, a, y = int(v[:2]), int(v[3:5]), int(v[6:])
    if not (1 <= g <= 31 and 1 <= a <= 12 and 2000 <= y <= 2099):
        return 'Geçerli bir tarih girin'
    return None

def _val_donem(v: str):
    """AA.YYYY formatı."""
    if not v:
        return None
    if not _re.fullmatch(r'\d{2}\.\d{4}', v):
        return 'AA.YYYY formatında girin'
    a, y = int(v[:2]), int(v[3:])
    if not (1 <= a <= 12 and 2000 <= y <= 2099):
        return 'Geçerli bir dönem girin'
    return None

def _val_kul_kodu(v: str):
    """Kullanıcı kodu: 6-12 rakam."""
    if not v:
        return None
    if not _re.fullmatch(r'\d{6,12}', v):
        return '6-12 haneli numara girin'
    return None

def _val_telefon(v: str):
    """10 rakam (mask ile veya ham)."""
    digits = ''.join(c for c in v if c.isdigit())
    if not v:
        return None
    if len(digits) != 10:
        return '10 haneli numara girin'
    return None

def _val_tutanak(v: str):
    """Tutanak sayısı — serbest metin."""
    return None

def _val_sozlesme_no(v: str):
    """E-##### veya benzeri."""
    if not v:
        return None
    if not _re.fullmatch(r'[A-Za-z0-9\-]{3,20}', v):
        return 'Geçerli sözleşme no girin (ör: E-65526)'
    return None


# ── Tarih / Dönem otomatik format yardımcıları ───────────────────────────────
def _date_format(digits: str) -> str:
    """GG.AA.YYYY — noktaları otomatik ekler."""
    n = min(len(digits), 8)
    if n == 0:
        return ''
    r = digits[:min(2, n)]
    if n > 2:
        r += '.' + digits[2:min(4, n)]
    if n > 4:
        r += '.' + digits[4:8]
    return r


def _period_format(digits: str) -> str:
    """AA.YYYY — noktayı otomatik ekler."""
    n = min(len(digits), 6)
    if n == 0:
        return ''
    r = digits[:min(2, n)]
    if n > 2:
        r += '.' + digits[2:6]
    return r


class _DateEdit(QLineEdit):
    """Tarih maskı: GG.AA.YYYY — noktaları otomatik ekler."""
    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Backspace:
            digits = ''.join(c for c in self.text() if c.isdigit())
            new_txt = _date_format(digits[:-1]) if digits else ''
            self.setText(new_txt)
            self.setCursorPosition(len(new_txt))
            return
        if e.text() and e.text().isdigit():
            digits = ''.join(c for c in self.text() if c.isdigit())
            if len(digits) < 8:
                new_txt = _date_format(digits + e.text())
                self.setText(new_txt)
                self.setCursorPosition(len(new_txt))
            return
        super().keyPressEvent(e)


class _PeriodEdit(QLineEdit):
    """Dönem maskı: AA.YYYY — noktayı otomatik ekler."""
    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Backspace:
            digits = ''.join(c for c in self.text() if c.isdigit())
            new_txt = _period_format(digits[:-1]) if digits else ''
            self.setText(new_txt)
            self.setCursorPosition(len(new_txt))
            return
        if e.text() and e.text().isdigit():
            digits = ''.join(c for c in self.text() if c.isdigit())
            if len(digits) < 6:
                new_txt = _period_format(digits + e.text())
                self.setText(new_txt)
                self.setCursorPosition(len(new_txt))
            return
        super().keyPressEvent(e)


# ── Telefon formatı yardımcıları ─────────────────────────────────────────────
def _phone_format(digits: str) -> str:
    """Rakamları (5xx) XXX XX-XX formatına çevirir."""
    n = min(len(digits), 10)
    if n == 0:
        return ''
    r = f'({digits[:min(3, n)]}'
    if n >= 3:
        r += ')'
    if n > 3:
        r += f' {digits[3:min(6, n)]}'
    if n > 6:
        r += f' {digits[6:min(8, n)]}'
    if n > 8:
        r += f'-{digits[8:10]}'
    return r


class _PhoneEdit(QLineEdit):
    """Telefon maskı: (5xx) XXX XX-XX. Backspace her zaman bir rakam siler."""
    def keyPressEvent(self, e):
        if e.key() == Qt.Key_Backspace:
            digits = ''.join(c for c in self.text() if c.isdigit())
            new_txt = _phone_format(digits[:-1]) if digits else ''
            self.setText(new_txt)
            self.setCursorPosition(len(new_txt))
            return
        if e.text() and e.text().isdigit():
            digits = ''.join(c for c in self.text() if c.isdigit())
            if len(digits) < 10:
                new_txt = _phone_format(digits + e.text())
                self.setText(new_txt)
                self.setCursorPosition(len(new_txt))
            return
        super().keyPressEvent(e)


# ── InputRow ──────────────────────────────────────────────────────────────────
class InputRow(QFrame):
    def __init__(self, label: str, placeholder: str = '',
                 echo_password: bool = False,
                 has_btn: bool = False, btn_tooltip: str = '',
                 icon_file: str = '',
                 phone_mask: bool = False,
                 date_mask: bool = False,
                 period_mask: bool = False,
                 validator=None,       # callable(str) -> str|None  (None=ok, str=hata)
                 parent=None):
        super().__init__(parent)
        self._phone_mask  = phone_mask
        self._date_mask   = date_mask
        self._period_mask = period_mask
        self._validator   = validator
        self._is_valid    = True
        self.setStyleSheet('QFrame{background:transparent;border:none;}')

        root = QVBoxLayout(self)
        root.setContentsMargins(0, 0, 0, 0)
        root.setSpacing(3)

        lbl_row = QHBoxLayout()
        lbl_row.setContentsMargins(0, 0, 0, 0)
        lbl_row.setSpacing(5)
        if icon_file:
            lbl_row.addWidget(_icon_lbl(icon_file, 14))
        lbl = QLabel(label)
        lbl.setStyleSheet(
            f'font-size:11px;font-weight:700;color:{TEXT2};'
            f'background:transparent;border:none;letter-spacing:0.4px;'
        )
        lbl_row.addWidget(lbl)
        lbl_row.addStretch()
        root.addLayout(lbl_row)

        row = QHBoxLayout()
        row.setContentsMargins(0, 0, 0, 0)
        row.setSpacing(6)

        if phone_mask:
            self.edit = _PhoneEdit()
        elif date_mask:
            self.edit = _DateEdit()
        elif period_mask:
            self.edit = _PeriodEdit()
        else:
            self.edit = QLineEdit()
        self.edit.setPlaceholderText(placeholder)
        if echo_password:
            self.edit.setEchoMode(QLineEdit.Password)
        self.edit.setFixedHeight(40)
        self._apply_normal_style()
        if validator:
            self.edit.textChanged.connect(self._run_validator)
        row.addWidget(self.edit, 1)

        if has_btn:
            self.btn = QPushButton('📁')
            self.btn.setFixedSize(40, 40)
            self.btn.setToolTip(btn_tooltip)
            self.btn.setCursor(Qt.PointingHandCursor)
            self.btn.setStyleSheet(f'''
                QPushButton {{
                    background: #F3F4F6;
                    border: 1.5px solid {BORDER};
                    border-radius: 10px;
                    font-size: 16px;
                }}
                QPushButton:hover {{ background: {BLUE_BG}; border-color: {BLUE}; }}
                QPushButton:pressed {{ background: #DBEAFE; }}
            ''')
            row.addWidget(self.btn)

        root.addLayout(row)

        # Hata mesajı label (başta gizli)
        self._err_lbl = QLabel()
        self._err_lbl.setStyleSheet(
            f'font-size:10px;color:{RED};background:transparent;'
            f'border:none;padding-left:4px;'
        )
        self._err_lbl.setVisible(False)
        root.addWidget(self._err_lbl)

    def _apply_normal_style(self):
        from core import theme as _t
        if _t.is_dark():
            bg, bg_f, txt, txt3, brd = '#0F1E33', '#162030', _t.DARK_TEXT, _t.DARK_TEXT3, _t.DARK_BORDER
        else:
            bg, bg_f, txt, txt3, brd = '#FAFAFA', '#FFFFFF', TEXT, TEXT3, BORDER
        self.edit.setStyleSheet(f'''
            QLineEdit {{
                background: {bg}; border: 1.5px solid {brd};
                border-radius: 10px; padding: 0 12px;
                font-size: 13px; color: {txt};
            }}
            QLineEdit:focus {{ border-color: {BLUE}; background: {bg_f}; }}
            QLineEdit::placeholder {{ color: {txt3}; }}
        ''')

    def _apply_error_style(self):
        from core import theme as _t
        if _t.is_dark():
            bg, txt, txt3 = '#2A1010', _t.DARK_TEXT, _t.DARK_TEXT3
        else:
            bg, txt, txt3 = '#FFF5F5', TEXT, TEXT3
        self.edit.setStyleSheet(f'''
            QLineEdit {{
                background: {bg}; border: 1.5px solid {RED};
                border-radius: 10px; padding: 0 12px;
                font-size: 13px; color: {txt};
            }}
            QLineEdit:focus {{ border-color: #DC2626; background: {bg}; }}
            QLineEdit::placeholder {{ color: {txt3}; }}
        ''')

    def _apply_ok_style(self):
        from core import theme as _t
        if _t.is_dark():
            bg, bg_f, txt, txt3 = '#0D2010', '#0F2512', _t.DARK_TEXT, _t.DARK_TEXT3
        else:
            bg, bg_f, txt, txt3 = '#F0FDF4', '#F0FDF4', TEXT, TEXT3
        self.edit.setStyleSheet(f'''
            QLineEdit {{
                background: {bg}; border: 1.5px solid {GREEN};
                border-radius: 10px; padding: 0 12px;
                font-size: 13px; color: {txt};
            }}
            QLineEdit:focus {{ border-color: {GREEN2}; background: {bg_f}; }}
            QLineEdit::placeholder {{ color: {txt3}; }}
        ''')

    def _run_validator(self, text: str):
        if not self._validator:
            return
        # Şifre alanlarını doğrulamıyoruz (privacy)
        if self.edit.echoMode() == QLineEdit.Password:
            return
        v = text.strip()
        if not v:
            # Boşken hata gösterme
            self._err_lbl.setVisible(False)
            self._apply_normal_style()
            self._is_valid = True
            return
        err = self._validator(v)
        if err:
            self._err_lbl.setText(f'⚠  {err}')
            self._err_lbl.setVisible(True)
            self._apply_error_style()
            self._is_valid = False
        else:
            self._err_lbl.setVisible(False)
            self._apply_ok_style()
            self._is_valid = True

    def value(self) -> str:
        if self._phone_mask:
            return ''.join(c for c in self.edit.text() if c.isdigit())
        return self.edit.text().strip()

    def set_value(self, v: str):
        if not v:
            self.edit.setText('')
            return
        if self._phone_mask:
            digits = ''.join(c for c in v if c.isdigit())
            self.edit.setText(_phone_format(digits))
        elif self._date_mask:
            digits = ''.join(c for c in v if c.isdigit())
            self.edit.setText(_date_format(digits))
        elif self._period_mask:
            digits = ''.join(c for c in v if c.isdigit())
            self.edit.setText(_period_format(digits))
        else:
            self.edit.setText(v)

    def set_error(self, has_error: bool):
        if has_error:
            from core import theme as _t
            if _t.is_dark():
                bg, bg_f, txt, txt3 = '#1E1200', '#201400', _t.DARK_TEXT, _t.DARK_TEXT3
            else:
                bg, bg_f, txt, txt3 = '#FFF7ED', '#FFFBF5', TEXT, TEXT3
            self.edit.setStyleSheet(f'''
                QLineEdit {{
                    background: {bg}; border: 1.5px solid {ORANGE};
                    border-radius: 10px; padding: 0 12px;
                    font-size: 13px; color: {txt};
                }}
                QLineEdit:focus {{ border-color: {ORANGE}; background: {bg_f}; }}
                QLineEdit::placeholder {{ color: {txt3}; }}
            ''')
        else:
            self._apply_normal_style()


# ── _drop_btns_style — ortak buton stilleri ───────────────────────────────────
def _make_drop_btn_sel() -> QPushButton:
    b = QPushButton('Seç')
    b.setFixedSize(52, 32)
    b.setCursor(Qt.PointingHandCursor)
    b.setStyleSheet(f'''
        QPushButton {{ background:{NAVY}; color:#FFF;
            border-radius:8px; border:none;
            font-size:11px; font-weight:600; }}
        QPushButton:hover {{ background:{NAVY2}; }}
    ''')
    return b

def _make_drop_btn_icon(icon_file: str, hover_bg: str, hover_border: str) -> QPushButton:
    b = QPushButton()
    b.setFixedSize(28, 32)
    b.setCursor(Qt.PointingHandCursor)
    b.setStyleSheet(f'''
        QPushButton {{ background:#F3F4F6;
            border-radius:8px; border:1px solid {BORDER}; }}
        QPushButton:hover {{
            background:{hover_bg}; border-color:{hover_border}; }}
    ''')
    pix = _icon_pix(icon_file, 14)
    if not pix.isNull():
        b.setIcon(QIcon(pix))
    return b


# ── DropFolderZone — sürükle-bırak klasör seçici ─────────────────────────────
class DropFolderZone(QFrame):
    changed = Signal(str)

    def __init__(self, parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QSizePolicy
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setAcceptDrops(True)
        self._path = ''
        self._set_idle_style()
        self.setFixedHeight(72)

        self._recent = RecentPopup('word_klasor', 'Son Kullanılan Klasörler')
        self._recent.picked.connect(self.set_value)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 0, 10, 0)
        lay.setSpacing(8)

        ico = QLabel()
        ico.setFixedSize(32, 32)
        ico.setAlignment(Qt.AlignCenter)
        ico.setStyleSheet('background:#EFF6FF;border-radius:7px;border:none;')
        _pix_kl = _icon_pix('klasor.png', 20)
        ico.setPixmap(_pix_kl) if not _pix_kl.isNull() else ico.setText('📂')
        lay.addWidget(ico)

        text_col = QVBoxLayout()
        text_col.setSpacing(2)
        self._lbl_title = QLabel('WORD KLASÖRÜ')
        self._lbl_title.setStyleSheet(
            f'font-size:11px;font-weight:700;color:{TEXT2};'
            f'background:transparent;border:none;letter-spacing:0.4px;'
        )
        self._lbl_path = QLabel('Klasör seçin veya buraya sürükleyin...')
        self._lbl_path.setStyleSheet(
            f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
        )
        text_col.addWidget(self._lbl_title)
        text_col.addWidget(self._lbl_path)
        lay.addLayout(text_col, 1)

        btn_lay = QHBoxLayout()
        btn_lay.setSpacing(6)
        self._btn_sel = _make_drop_btn_sel()
        self._btn_rec = _make_drop_btn_icon('clock.png', BLUE_BG, NAVY)
        self._btn_rec.setToolTip('Son kullanılan klasörler')
        self._btn_clr = _make_drop_btn_icon('x.png', '#FFE4E4', RED)
        self._btn_clr.setToolTip('Temizle')
        btn_lay.addWidget(self._btn_sel)
        btn_lay.addWidget(self._btn_rec)
        btn_lay.addWidget(self._btn_clr)
        lay.addLayout(btn_lay)

        self._btn_sel.clicked.connect(self._browse)
        self._btn_rec.clicked.connect(lambda: self._recent.show_at(self._btn_rec))
        self._btn_clr.clicked.connect(lambda: self.set_value(''))

    def _set_idle_style(self):
        from core import theme as _t
        bg = _t.DARK_CARD if _t.is_dark() else CARD
        brd = _t.DARK_BORDER if _t.is_dark() else BORDER
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:12px;border:1.5px dashed {brd};}}')

    def _set_hover_style(self):
        from core import theme as _t
        bg = '#0D2010' if _t.is_dark() else '#F0FDF4'
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:12px;border:2px solid {GREEN};}}')

    def _set_filled_style(self):
        from core import theme as _t
        bg = _t.DARK_BLUE_BG if _t.is_dark() else '#F0F9FF'
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:12px;border:1.5px solid {BLUE};}}')

    def _browse(self):
        d = QFileDialog.getExistingDirectory(self, 'Word Klasörünü Seçin')
        if d:
            self.set_value(d)

    def set_value(self, v: str):
        self._path = v
        if v:
            short = v if len(v) <= 40 else '…' + v[-38:]
            self._lbl_path.setText(f'✓  {short}')
            self._lbl_path.setStyleSheet(
                f'font-size:11px;color:{GREEN2};font-weight:600;background:transparent;border:none;'
            )
            self._set_filled_style()
            _recent_add('word_klasor', v)
        else:
            self._lbl_path.setText('Klasör seçin veya buraya sürükleyin...')
            self._lbl_path.setStyleSheet(
                f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
            )
            self._set_idle_style()
        self.changed.emit(v)

    def value(self) -> str:
        return self._path

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            e.acceptProposedAction()
            self._set_hover_style()

    def dragLeaveEvent(self, e):
        self._set_filled_style() if self._path else self._set_idle_style()

    def dropEvent(self, e):
        self._set_filled_style() if self._path else self._set_idle_style()
        for url in e.mimeData().urls():
            p = url.toLocalFile()
            if os.path.isdir(p):
                self.set_value(p)
                break


# ── DropFileZone — sürükle-bırak Excel seçici ────────────────────────────────
class DropFileZone(QFrame):
    changed = Signal(str)
    _ACCEPTED = ('.xlsm', '.xlsx')

    def __init__(self, parent=None):
        super().__init__(parent)
        from PySide6.QtWidgets import QSizePolicy
        self.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.setAcceptDrops(True)
        self._path = ''
        self._set_idle_style()
        self.setFixedHeight(72)

        self._recent = RecentPopup('master_excel', 'Son Kullanılan Excel Dosyaları')
        self._recent.picked.connect(self.set_value)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(10, 0, 10, 0)
        lay.setSpacing(8)

        ico = QLabel()
        ico.setFixedSize(32, 32)
        ico.setAlignment(Qt.AlignCenter)
        ico.setStyleSheet('background:#F0FDF4;border-radius:7px;border:none;')
        _pix_xl = _icon_pix('excel.png', 20)
        ico.setPixmap(_pix_xl) if not _pix_xl.isNull() else ico.setText('📊')
        lay.addWidget(ico)

        text_col = QVBoxLayout()
        text_col.setSpacing(2)
        self._lbl_title = QLabel('MASTER EXCEL')
        self._lbl_title.setStyleSheet(
            f'font-size:11px;font-weight:700;color:{TEXT2};'
            f'background:transparent;border:none;letter-spacing:0.4px;'
        )
        self._lbl_path = QLabel('karsit_master.xlsm seçin...')
        self._lbl_path.setStyleSheet(
            f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
        )
        text_col.addWidget(self._lbl_title)
        text_col.addWidget(self._lbl_path)
        lay.addLayout(text_col, 1)

        btn_lay = QHBoxLayout()
        btn_lay.setSpacing(6)
        self._btn_sel = _make_drop_btn_sel()
        self._btn_rec = _make_drop_btn_icon('clock.png', BLUE_BG, NAVY)
        self._btn_rec.setToolTip('Son kullanılan dosyalar')
        self._btn_clr = _make_drop_btn_icon('x.png', '#FFE4E4', RED)
        self._btn_clr.setToolTip('Temizle')
        btn_lay.addWidget(self._btn_sel)
        btn_lay.addWidget(self._btn_rec)
        btn_lay.addWidget(self._btn_clr)
        lay.addLayout(btn_lay)

        self._btn_sel.clicked.connect(self._browse)
        self._btn_rec.clicked.connect(lambda: self._recent.show_at(self._btn_rec))
        self._btn_clr.clicked.connect(lambda: self.set_value(''))

    def _set_idle_style(self):
        from core import theme as _t
        bg = _t.DARK_CARD if _t.is_dark() else CARD
        brd = _t.DARK_BORDER if _t.is_dark() else BORDER
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:12px;border:1.5px dashed {brd};}}')

    def _set_hover_style(self):
        from core import theme as _t
        bg = '#0D2010' if _t.is_dark() else '#F0FDF4'
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:12px;border:2px solid {GREEN};}}')

    def _set_filled_style(self):
        from core import theme as _t
        bg = '#0D2010' if _t.is_dark() else '#F0FDF4'
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:12px;border:1.5px solid {GREEN};}}')

    def _browse(self):
        p, _ = QFileDialog.getOpenFileName(
            self, 'Master Excel Seçin', '',
            'Excel Dosyaları (*.xlsm *.xlsx);;Tüm Dosyalar (*.*)'
        )
        if p:
            self.set_value(p)

    def set_value(self, v: str):
        self._path = v
        if v:
            short = os.path.basename(v)
            self._lbl_path.setText(f'✓  {short}')
            self._lbl_path.setStyleSheet(
                f'font-size:11px;color:{GREEN2};font-weight:600;background:transparent;border:none;'
            )
            self._set_filled_style()
            _recent_add('master_excel', v)
        else:
            self._lbl_path.setText('karsit_master.xlsm seçin...')
            self._lbl_path.setStyleSheet(
                f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
            )
            self._set_idle_style()
        self.changed.emit(v)

    def value(self) -> str:
        return self._path

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            for url in e.mimeData().urls():
                if url.toLocalFile().lower().endswith(self._ACCEPTED):
                    e.acceptProposedAction()
                    self._set_hover_style()
                    return

    def dragLeaveEvent(self, e):
        self._set_filled_style() if self._path else self._set_idle_style()

    def dropEvent(self, e):
        self._set_filled_style() if self._path else self._set_idle_style()
        for url in e.mimeData().urls():
            p = url.toLocalFile()
            if p.lower().endswith(self._ACCEPTED):
                self.set_value(p)
                break


# ── LogPanel ──────────────────────────────────────────────────────────────────
_LOG_CATS = {
    'TÜMÜ':    ('',      '#94A3B8'),
    'GİRİŞ':  (BLUE,    '#93C5FD'),
    'CAPTCHA': (GOLD,    '#FCD34D'),
    'HATA':    (RED,     '#FCA5A5'),
    'TAMAM':   (GREEN,   '#86EFAC'),
}

class LogPanel(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(
            f'QFrame{{background:{LOG_BG};border-radius:12px;border:1px solid #1E293B;}}'
        )
        self._messages: list = []   # (ts, html_line, category)
        self._active_cat: str = 'TÜMÜ'

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        # Başlık bar
        bar = QFrame()
        bar.setFixedHeight(34)
        bar.setStyleSheet(
            'QFrame{background:#1E293B;border-radius:12px;'
            'border-bottom-left-radius:0;border-bottom-right-radius:0;border:none;}'
        )
        bl = QHBoxLayout(bar)
        bl.setContentsMargins(12, 0, 12, 0)
        bl.setSpacing(6)
        title = QLabel('▶  Terminal')
        title.setStyleSheet('font-size:11px;font-weight:700;color:#94A3B8;background:transparent;border:none;')
        bl.addWidget(title)
        bl.addStretch()

        btn_clear = QPushButton('Temizle')
        btn_clear.setFixedHeight(22)
        btn_clear.setCursor(Qt.PointingHandCursor)
        btn_clear.setStyleSheet(
            'QPushButton{background:#334155;color:#94A3B8;border:none;border-radius:4px;'
            'font-size:10px;padding:0 8px;}'
            'QPushButton:hover{background:#475569;color:#E2E8F0;}'
        )
        btn_clear.clicked.connect(self.clear)
        bl.addWidget(btn_clear)
        lay.addWidget(bar)

        # Filtre butonları bar
        fbar = QFrame()
        fbar.setFixedHeight(28)
        fbar.setStyleSheet('QFrame{background:#0D1B2E;border:none;}')
        fl = QHBoxLayout(fbar)
        fl.setContentsMargins(10, 0, 10, 0)
        fl.setSpacing(4)
        self._filter_btns: dict = {}
        for cat, (_, fc) in _LOG_CATS.items():
            b = QPushButton(cat)
            b.setFixedHeight(20)
            b.setCursor(Qt.PointingHandCursor)
            b.setCheckable(True)
            b.setChecked(cat == 'TÜMÜ')
            b.setStyleSheet(f'''
                QPushButton {{
                    background:transparent; border:1px solid #334155;
                    border-radius:4px; color:#64748B;
                    font-size:9px; font-weight:700; padding:0 6px;
                }}
                QPushButton:checked {{
                    background:#1E3A5F; border-color:{fc}; color:{fc};
                }}
                QPushButton:hover:!checked {{ color:#94A3B8; }}
            ''')
            b.clicked.connect(lambda _, c=cat: self._set_filter(c))
            self._filter_btns[cat] = b
            fl.addWidget(b)
        fl.addStretch()
        lay.addWidget(fbar)

        # Progress bar
        self.prog = QProgressBar()
        self.prog.setFixedHeight(4)
        self.prog.setRange(0, 100)
        self.prog.setValue(0)
        self.prog.setTextVisible(False)
        self.prog.setStyleSheet(f'''
            QProgressBar {{ background:#1E293B; border:none; border-radius:0; }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {BLUE}, stop:1 {GREEN});
            }}
        ''')
        lay.addWidget(self.prog)

        self.txt = QPlainTextEdit()
        self.txt.setReadOnly(True)
        self.txt.setStyleSheet(f'''
            QPlainTextEdit {{
                background: {LOG_BG}; color: #94A3B8; border: none;
                font-family: "Cascadia Code", "Consolas", monospace;
                font-size: 12px; padding: 10px;
            }}
        ''')
        lay.addWidget(self.txt)

    def _cat_from_color(self, color: str, msg: str) -> str:
        if not color:
            return ''
        cu = color.upper()
        if cu == RED.upper():
            return 'HATA'
        if cu == GREEN.upper():
            return 'TAMAM'
        if cu == GOLD.upper() or 'captcha' in msg.lower():
            return 'CAPTCHA'
        if cu == BLUE.upper() or cu == BLUE2.upper():
            return 'GİRİŞ'
        return ''

    def _set_filter(self, cat: str):
        self._active_cat = cat
        for c, b in self._filter_btns.items():
            b.setChecked(c == cat)
        self._rerender()

    def _rerender(self):
        self.txt.clear()
        for (ts, html, cat) in self._messages:
            if self._active_cat == 'TÜMÜ' or cat == self._active_cat:
                self.txt.appendHtml(html)
        self.txt.verticalScrollBar().setValue(self.txt.verticalScrollBar().maximum())

    def set_progress(self, current: int, total: int):
        self.prog.setValue(int(current / total * 100) if total else 0)

    def append(self, msg: str, color: str = ''):
        ts = datetime.now().strftime('%H:%M:%S')
        cat = self._cat_from_color(color, msg)

        # Kategori prefix badge
        prefix_html = ''
        if cat and cat in _LOG_CATS:
            badge_color = _LOG_CATS[cat][1]
            prefix_html = (
                f'<span style="background:{badge_color}22;color:{badge_color};'
                f'font-size:9px;font-weight:700;border-radius:3px;'
                f'padding:1px 4px;">{cat}</span> '
            )

        if color:
            html = (
                f'<span style="color:#475569;">[{ts}]</span> '
                f'{prefix_html}'
                f'<span style="color:{color};">{msg}</span>'
            )
        else:
            html = (
                f'<span style="color:#475569;">[{ts}]</span> '
                f'{prefix_html}'
                f'<span style="color:#94A3B8;">{msg}</span>'
            )

        self._messages.append((ts, html, cat))
        if self._active_cat == 'TÜMÜ' or cat == self._active_cat:
            self.txt.appendHtml(html)
            self.txt.verticalScrollBar().setValue(self.txt.verticalScrollBar().maximum())

    def clear(self):
        self._messages.clear()
        self.txt.clear()


# ── FirmaListPanel ────────────────────────────────────────────────────────────
class FirmaListPanel(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet(
            f'QFrame{{background:{LOG_BG};border-radius:12px;border:1px solid #1E293B;}}'
        )
        self._firmalar: list = []
        self._aktif_idx: int = -1
        self._bilgi_istem: set = set()
        self._statuses: dict = {}   # idx -> 'pending'/'active'/'done'/'error'

        lay = QVBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        bar = QFrame()
        bar.setFixedHeight(34)
        bar.setStyleSheet(
            'QFrame{background:#1E293B;border-radius:12px;'
            'border-bottom-left-radius:0;border-bottom-right-radius:0;border:none;}'
        )
        bl = QHBoxLayout(bar)
        bl.setContentsMargins(12, 0, 12, 0)
        bl.setSpacing(8)
        title_lbl = QLabel('📋  İşlem Bilgileri')
        title_lbl.setStyleSheet(
            'font-size:11px;font-weight:700;color:#94A3B8;background:transparent;border:none;'
        )
        bl.addWidget(title_lbl)
        bl.addStretch()
        self.lbl_counter = QLabel('0 / 0')
        self.lbl_counter.setStyleSheet(
            f'font-size:12px;font-weight:700;color:{GOLD};background:transparent;border:none;'
        )
        bl.addWidget(self.lbl_counter)
        lay.addWidget(bar)

        self.prog = QProgressBar()
        self.prog.setFixedHeight(6)
        self.prog.setRange(0, 100)
        self.prog.setValue(0)
        self.prog.setTextVisible(False)
        self.prog.setStyleSheet(f'''
            QProgressBar {{ background:#1E293B; border:none; }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {BLUE}, stop:1 {GREEN});
            }}
        ''')
        lay.addWidget(self.prog)

        scroll = QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setStyleSheet(f'''
            QScrollArea {{ border:none; background:transparent; }}
            QScrollBar:vertical {{ background:{LOG_BG}; width:4px; border:none; }}
            QScrollBar::handle:vertical {{ background:#334155; border-radius:2px; min-height:20px; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0; }}
        ''')

        self._list_w = QWidget()
        self._list_w.setStyleSheet('background:transparent;')
        self._list_lay = QVBoxLayout(self._list_w)
        self._list_lay.setContentsMargins(8, 8, 8, 8)
        self._list_lay.setSpacing(2)
        self._list_lay.addStretch()

        self._empty_lbl = QLabel('Word klasörü seçilince\nfirmalar burada görünür')
        self._empty_lbl.setAlignment(Qt.AlignCenter)
        self._empty_lbl.setStyleSheet(
            'color:rgba(148,163,184,0.4);font-size:12px;background:transparent;border:none;'
        )
        self._list_lay.insertWidget(0, self._empty_lbl)

        scroll.setWidget(self._list_w)
        lay.addWidget(scroll, 1)

    def set_firmalar(self, firmalar: list, bilgi_istem_set: set = None):
        """firmalar: isim listesi. bilgi_istem_set: kırmızı gösterilecek index set'i."""
        self._firmalar = firmalar
        self._bilgi_istem = bilgi_istem_set or set()
        self._aktif_idx = -1
        self._statuses = {}
        self._rebuild()
        gecerli = len(firmalar) - len(self._bilgi_istem)
        self.lbl_counter.setText(f'0 / {gecerli}')
        self.prog.setValue(0)

    def set_aktif(self, idx: int):
        self._statuses[idx] = 'active'
        self._aktif_idx = idx
        self._rebuild()
        total = len(self._firmalar)
        self.lbl_counter.setText(f'{idx + 1} / {total}')
        self.prog.setValue(int((idx + 1) / total * 100) if total else 0)
        QTimer.singleShot(50, self._scroll_to_aktif)

    def set_firma_done(self, idx: int):
        self._statuses[idx] = 'done'
        self._rebuild()

    def set_firma_error(self, idx: int):
        self._statuses[idx] = 'error'
        self._rebuild()

    def _rebuild(self):
        for i in range(self._list_lay.count() - 1, -1, -1):
            item = self._list_lay.itemAt(i)
            if item and item.widget() and item.widget() is not self._empty_lbl:
                self._list_lay.takeAt(i).widget().deleteLater()

        if not self._firmalar:
            self._empty_lbl.setVisible(True)
            return
        self._empty_lbl.setVisible(False)

        _STATUS_CFG = {
            'done':    (GREEN,  '#22C55E22', '✓', '3px solid ' + GREEN),
            'error':   (RED,    '#EF444422', '✗', '3px solid ' + RED),
            'active':  (ORANGE, 'qlineargradient(x1:0,y1:0,x2:1,y2:0,stop:0 rgba(249,115,22,0.28),stop:1 rgba(249,115,22,0.08))', '▶', '3px solid ' + ORANGE),
            'pending': ('#475569', 'transparent', '⏳', '2px solid #1E293B'),
        }

        for i, firma in enumerate(self._firmalar):
            is_bilgi = i in self._bilgi_istem
            if is_bilgi:
                ico, text_color, bg, border = '⛔', RED, 'rgba(239,68,68,0.1)', f'3px solid {RED}'
                fw = '600'
                lbl = QLabel(f' {ico}  {firma}')
                lbl.setWordWrap(False)
                lbl.setStyleSheet(f'''
                    QLabel {{
                        background: {bg};
                        border-left: {border};
                        border-radius: 4px; color: {text_color};
                        font-size: 11px; font-weight: {fw}; padding: 4px 8px;
                    }}
                ''')
                lbl.setToolTip('Bilgi İsteme belgesi — otomasyonda atlanacak')
            else:
                st = self._statuses.get(i, 'pending')
                text_color, bg, ico_char, border = _STATUS_CFG[st]
                fw = '700' if st == 'active' else '500'
                fs = '12px' if st == 'active' else '11px'
                lbl = QLabel(f' {ico_char}  {firma}')
                lbl.setWordWrap(False)
                lbl.setStyleSheet(f'''
                    QLabel {{
                        background: {bg};
                        border-left: {border};
                        border-radius: {'6px' if st == 'active' else '4px'};
                        color: {text_color};
                        font-size: {fs}; font-weight: {fw}; padding: 5px 8px;
                    }}
                ''')
                if st == 'active':
                    glow = QGraphicsDropShadowEffect()
                    glow.setBlurRadius(14)
                    glow.setColor(QColor(249, 115, 22, 120))
                    glow.setOffset(0, 0)
                    lbl.setGraphicsEffect(glow)
            self._list_lay.insertWidget(self._list_lay.count() - 1, lbl)

    def _scroll_to_aktif(self):
        if self._aktif_idx < 0:
            return
        total = self._list_lay.count() - 1
        if total > 0:
            ratio = self._aktif_idx / total
            for child in self.findChildren(QScrollArea):
                sb = child.verticalScrollBar()
                sb.setValue(int(ratio * sb.maximum()))
                break


# ── OnizlemeDialog ────────────────────────────────────────────────────────────
class OnizlemeDialog(QDialog):
    def __init__(self, firmalar: list, ayarlar: dict,
                 tutanak_baslangic: str, parent=None):
        super().__init__(parent)
        self.setWindowTitle('Önizleme — Başlamadan Önce Kontrol Et')
        self.setModal(True)
        self.setMinimumWidth(500)
        self.setStyleSheet(f'background:{BG};')

        lay = QVBoxLayout(self)
        lay.setContentsMargins(24, 20, 24, 20)
        lay.setSpacing(14)

        title = QLabel('🚀  Otomasyon Başlatılıyor')
        title.setStyleSheet(
            f'font-size:16px;font-weight:800;color:{NAVY};background:transparent;border:none;'
        )
        lay.addWidget(title)

        # Özet mini-kartlar
        def _mini(label, value, color=TEXT):
            card = QFrame()
            card.setStyleSheet(f'QFrame{{background:{CARD};border-radius:10px;border:1px solid {BORDER};}}')
            cl = QVBoxLayout(card)
            cl.setContentsMargins(14, 10, 14, 10)
            cl.setSpacing(2)
            v_lbl = QLabel(str(value))
            v_lbl.setStyleSheet(
                f'font-size:22px;font-weight:800;color:{color};background:transparent;border:none;'
            )
            k_lbl = QLabel(label)
            k_lbl.setStyleSheet(
                f'font-size:10px;font-weight:700;color:{TEXT2};letter-spacing:0.4px;'
                f'background:transparent;border:none;'
            )
            cl.addWidget(v_lbl)
            cl.addWidget(k_lbl)
            return card

        n_firma = len(firmalar)
        tur = ayarlar.get('tasdik_turu', 'KDV')
        kdv_tur = ayarlar.get('kdv_donem_turu', 'NORMAL')
        kdv_tur_kisa = 'İndirimli' if 'İNDİRİMLİ' in kdv_tur else 'Normal'

        grid = QGridLayout()
        grid.setSpacing(10)
        grid.addWidget(_mini('TOPLAM FİRMA', n_firma, NAVY),          0, 0)
        grid.addWidget(_mini('TASDİK TÜRÜ',  tur,     BLUE),          0, 1)
        grid.addWidget(_mini('KDV TÜRÜ', kdv_tur_kisa,
                             GOLD if 'İNDİRİMLİ' in kdv_tur else TEXT), 0, 2)
        lay.addLayout(grid)

        # Tutanak aralığı
        if tutanak_baslangic and '/' in tutanak_baslangic:
            try:
                from .karsit_import import _tutanak_uret
                son = _tutanak_uret(tutanak_baslangic, max(0, n_firma - 1))
                tut_text = f'{tutanak_baslangic}  →  {son}'
            except Exception:
                tut_text = tutanak_baslangic
        elif tutanak_baslangic:
            tut_text = tutanak_baslangic
        else:
            tut_text = "—  (Excel'den okunacak)"

        tut_f = QFrame()
        tut_f.setStyleSheet(
            f'QFrame{{background:{BLUE_BG};border-radius:10px;border:1px solid {BLUE};}}'
        )
        tl = QHBoxLayout(tut_f)
        tl.setContentsMargins(14, 10, 14, 10)
        tl.setSpacing(10)
        tl.addWidget(_icon_lbl('hashtag.png', 16))
        tut_lbl = QLabel(f'<b>Tutanak Sayısı:</b>  {tut_text}')
        tut_lbl.setStyleSheet(
            f'font-size:12px;color:{TEXT};background:transparent;border:none;'
        )
        tl.addWidget(tut_lbl)
        tl.addStretch()
        lay.addWidget(tut_f)

        # Firma listesi önizleme (max 6)
        if firmalar:
            list_lbl = QLabel('İşlenecek Firmalar:')
            list_lbl.setStyleSheet(
                f'font-size:11px;font-weight:700;color:{TEXT2};background:transparent;border:none;'
            )
            lay.addWidget(list_lbl)

            list_f = QFrame()
            list_f.setStyleSheet(
                f'QFrame{{background:{CARD};border-radius:10px;border:1px solid {BORDER};}}'
            )
            ll = QVBoxLayout(list_f)
            ll.setContentsMargins(12, 8, 12, 8)
            ll.setSpacing(3)
            show_n = min(6, n_firma)
            for i, f in enumerate(firmalar[:show_n]):
                r = QLabel(f'  {i + 1}. {f}')
                r.setStyleSheet(
                    f'font-size:11px;color:{TEXT};background:transparent;border:none;'
                )
                ll.addWidget(r)
            if n_firma > show_n:
                more = QLabel(f'  … ve {n_firma - show_n} firma daha')
                more.setStyleSheet(
                    f'font-size:11px;color:{TEXT2};background:transparent;border:none;'
                )
                ll.addWidget(more)
            lay.addWidget(list_f)

        # Ayırıcı + butonlar
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
        lay.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        btn_iptal = QPushButton('İptal')
        btn_iptal.setFixedHeight(42)
        btn_iptal.setCursor(Qt.PointingHandCursor)
        btn_iptal.setStyleSheet(f'''
            QPushButton {{
                background:#F3F4F6; border:1px solid {BORDER};
                border-radius:10px; color:{TEXT2}; font-size:13px; font-weight:600;
            }}
            QPushButton:hover {{ background:{BORDER}; }}
        ''')
        btn_iptal.clicked.connect(self.reject)

        btn_basla = QPushButton('▶  Başlat')
        btn_basla.setFixedHeight(42)
        btn_basla.setCursor(Qt.PointingHandCursor)
        btn_basla.setStyleSheet(f'''
            QPushButton {{
                background:{GREEN}; color:#FFFFFF;
                border:none; border-radius:10px;
                font-size:13px; font-weight:700;
            }}
            QPushButton:hover {{ background:{GREEN2}; }}
        ''')
        btn_basla.clicked.connect(self.accept)

        btn_row.addWidget(btn_iptal, 1)
        btn_row.addWidget(btn_basla, 2)
        lay.addLayout(btn_row)


# ── RaporDialog ────────────────────────────────────────────────────────────────
class RaporDialog(QDialog):
    def __init__(self, data: list, parent=None):
        """data: list of {'firma': str, 'tutanak': str, 'status': str, 'hata': str}"""
        super().__init__(parent)
        self.setWindowTitle('Otomasyon Raporu')
        self.setModal(True)
        self.setMinimumWidth(680)
        self.setMinimumHeight(420)
        self.setStyleSheet(f'background:{BG};')
        self._data = data

        lay = QVBoxLayout(self)
        lay.setContentsMargins(20, 18, 20, 18)
        lay.setSpacing(12)

        # Başlık + özet
        hdr = QHBoxLayout()
        title = QLabel('📊  Otomasyon Raporu')
        title.setStyleSheet(
            f'font-size:16px;font-weight:800;color:{NAVY};background:transparent;border:none;'
        )
        hdr.addWidget(title)
        hdr.addStretch()
        tamamlanan = sum(1 for r in data if r.get('status') == 'tamamlandi')
        hata       = sum(1 for r in data if r.get('status') == 'hata')
        ozet = QLabel(
            f'<span style="color:{GREEN};font-weight:700;">✓ {tamamlanan} Tamamlandı</span>'
            f'&nbsp;&nbsp;&nbsp;'
            f'<span style="color:{RED};font-weight:700;">✗ {hata} Hata</span>'
        )
        ozet.setStyleSheet(f'font-size:13px;background:transparent;border:none;')
        hdr.addWidget(ozet)
        lay.addLayout(hdr)

        # Tablo
        from PySide6.QtWidgets import QTableWidget, QTableWidgetItem, QHeaderView
        tbl = QTableWidget(len(data), 4, self)
        tbl.setHorizontalHeaderLabels(['Firma', 'Tutanak No', 'Durum', 'Açıklama'])
        tbl.horizontalHeader().setStretchLastSection(True)
        tbl.horizontalHeader().setSectionResizeMode(0, QHeaderView.Stretch)
        tbl.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        tbl.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        tbl.setSelectionBehavior(QTableWidget.SelectRows)
        tbl.setAlternatingRowColors(True)
        tbl.setEditTriggers(QTableWidget.NoEditTriggers)
        tbl.verticalHeader().setVisible(False)
        tbl.setStyleSheet(f'''
            QTableWidget {{
                background: {CARD}; border: 1px solid {BORDER};
                border-radius: 10px; gridline-color: {BORDER};
                font-size: 12px; color: {TEXT};
            }}
            QTableWidget::item {{ padding: 6px 10px; }}
            QTableWidget::item:selected {{ background: {BLUE_BG}; color: {BLUE2}; }}
            QHeaderView::section {{
                background: {BG}; color: {TEXT2};
                font-size: 11px; font-weight: 700;
                border: none; border-bottom: 1.5px solid {BORDER};
                padding: 8px 10px;
            }}
            QTableWidget {{ alternate-background-color: #F9FAFB; }}
        ''')

        for i, row in enumerate(data):
            tbl.setItem(i, 0, QTableWidgetItem(row.get('firma', '')))
            tbl.setItem(i, 1, QTableWidgetItem(row.get('tutanak', '')))
            status = row.get('status', '')
            ok = status == 'tamamlandi'
            st_item = QTableWidgetItem('✓ Tamamlandı' if ok else '✗ Hata')
            st_item.setForeground(QColor(GREEN if ok else RED))
            tbl.setItem(i, 2, st_item)
            tbl.setItem(i, 3, QTableWidgetItem(row.get('hata', '')))

        lay.addWidget(tbl, 1)

        # Ayırıcı + butonlar
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
        lay.addWidget(sep)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        btn_excel = QPushButton('📥  Excel Olarak İndir')
        btn_excel.setFixedHeight(42)
        btn_excel.setCursor(Qt.PointingHandCursor)
        btn_excel.setStyleSheet(f'''
            QPushButton {{
                background:{NAVY}; color:{GOLD};
                border:1.5px solid rgba(201,164,106,0.4);
                border-radius:10px; font-size:12px; font-weight:700;
            }}
            QPushButton:hover {{ background:{NAVY2}; border-color:{GOLD}; }}
        ''')
        btn_excel.clicked.connect(self._excel_indir)

        btn_kapat = QPushButton('Kapat')
        btn_kapat.setFixedHeight(42)
        btn_kapat.setCursor(Qt.PointingHandCursor)
        btn_kapat.setStyleSheet(f'''
            QPushButton {{
                background:#F3F4F6; border:1px solid {BORDER};
                border-radius:10px; color:{TEXT2}; font-size:12px;
            }}
            QPushButton:hover {{ background:{BORDER}; }}
        ''')
        btn_kapat.clicked.connect(self.accept)

        btn_row.addWidget(btn_excel, 2)
        btn_row.addWidget(btn_kapat, 1)
        lay.addLayout(btn_row)

    def _excel_indir(self):
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        dest, _ = QFileDialog.getSaveFileName(
            self, 'Raporu Kaydet',
            os.path.join(desktop, f'karsit_rapor_{ts}.xlsx'),
            'Excel Dosyası (*.xlsx)'
        )
        if not dest:
            return
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Karşıt Raporu'

            hdrs = ['Firma Adı', 'Tutanak No', 'Durum', 'Açıklama']
            hdr_fill = PatternFill('solid', fgColor='0B1F3A')
            hdr_font = Font(bold=True, color='C9A46A', size=11)
            for col, h in enumerate(hdrs, 1):
                c = ws.cell(row=1, column=col, value=h)
                c.fill = hdr_fill
                c.font = hdr_font
                c.alignment = Alignment(horizontal='center', vertical='center')

            green_fill = PatternFill('solid', fgColor='D1FAE5')
            red_fill   = PatternFill('solid', fgColor='FEE2E2')
            for i, row in enumerate(self._data, 2):
                ok   = row.get('status', '') == 'tamamlandi'
                fill = green_fill if ok else red_fill
                ws.cell(row=i, column=1, value=row.get('firma',   '')).fill = fill
                ws.cell(row=i, column=2, value=row.get('tutanak', '')).fill = fill
                dc = ws.cell(row=i, column=3,
                             value='✓ Tamamlandı' if ok else '✗ Hata')
                dc.fill = fill
                dc.font = Font(bold=True, color='166534' if ok else 'B91C1C')
                ws.cell(row=i, column=4, value=row.get('hata', '')).fill = fill

            ws.column_dimensions['A'].width = 45
            ws.column_dimensions['B'].width = 18
            ws.column_dimensions['C'].width = 16
            ws.column_dimensions['D'].width = 40
            ws.row_dimensions[1].height = 22

            wb.save(dest)
            QMessageBox.information(self, 'Kaydedildi',
                                    f'Rapor kaydedildi:\n{dest}')
        except Exception as e:
            QMessageBox.critical(self, 'Hata', str(e))


# ── StatusChip — buton satırı yanı tarayıcı durum rozeti ─────────────────────
class StatusChip(QFrame):
    """Buton satırının yanında küçük pill-badge: 🔒 Gizli / 👁 Canlı."""
    clicked = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._running = False
        self._visible = False
        self._pulse_on = True
        self.setCursor(Qt.PointingHandCursor)
        self.setToolTip('Tıkla: Tarayıcıyı göster / gizle')
        self.setFixedHeight(44)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 0, 14, 0)
        lay.setSpacing(6)

        self._ico = QLabel()
        self._ico.setStyleSheet('font-size:15px;background:transparent;border:none;')
        lay.addWidget(self._ico)

        self._txt = QLabel()
        self._txt.setStyleSheet(
            'font-size:11px;font-weight:700;letter-spacing:0.5px;'
            'background:transparent;border:none;'
        )
        lay.addWidget(self._txt)

        self._pulse_timer = QTimer(self)
        self._pulse_timer.setInterval(550)
        self._pulse_timer.timeout.connect(self._pulse_step)

        self._refresh()

    # Her iki modda da iyi duran, _LIGHT_TO_DARK listesinde olmayan renkler
    _STYLE_HIDDEN  = ('rgba(11,31,58,0.85)',  '#334155', '#64748B', '🔒', 'GİZLİ')
    _STYLE_VISIBLE = ('#0F2A1E',             GREEN,      GREEN,     '👁',  'CANLI')

    def _refresh(self):
        bg, border, txt_color, ico, txt = (
            self._STYLE_VISIBLE if self._visible else self._STYLE_HIDDEN
        )
        self.setStyleSheet(f'''
            QFrame {{
                background: {bg};
                border-radius: 22px;
                border: 1.5px solid {border};
            }}
            QFrame:hover {{
                border-color: {GREEN if self._visible else GOLD};
            }}
        ''')
        self._ico.setText(ico)
        self._ico.setStyleSheet(f'font-size:15px;background:transparent;border:none;')
        self._txt.setText(txt)
        self._txt.setStyleSheet(
            f'font-size:11px;font-weight:700;color:{txt_color};'
            'letter-spacing:0.5px;background:transparent;border:none;'
        )

    def _pulse_step(self):
        if not (self._running and self._visible):
            return
        border = GREEN if self._pulse_on else '#1F4A33'
        self.setStyleSheet(f'''
            QFrame {{
                background: #0F2A1E;
                border-radius: 22px;
                border: 2px solid {border};
            }}
        ''')
        self._pulse_on = not self._pulse_on

    def set_state(self, running: bool, visible: bool):
        self._running = running
        self._visible = visible
        self._refresh()
        if running and visible:
            self._pulse_timer.start()
        else:
            self._pulse_timer.stop()
            self._pulse_on = True
            self._refresh()

    def mousePressEvent(self, e):
        self.clicked.emit()
        super().mousePressEvent(e)


# ── StatBar — 3 özet istatistik kartı ────────────────────────────────────────
class _StatKart(QFrame):
    def __init__(self, emoji: str, label: str, color: str, parent=None):
        super().__init__(parent)
        self.setStyleSheet(f'''
            QFrame {{
                background: {CARD};
                border-radius: 14px;
                border: 1px solid {BORDER};
            }}
        ''')
        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(6)

        top = QHBoxLayout()
        top.setSpacing(0)
        self.num = QLabel('0')
        self.num.setStyleSheet(
            f'font-size:30px;font-weight:900;color:{color};'
            f'background:transparent;border:none;'
        )
        top.addWidget(self.num)
        top.addStretch()
        ico = QLabel(emoji)
        ico.setStyleSheet('font-size:22px;background:transparent;border:none;')
        top.addWidget(ico)
        lay.addLayout(top)

        lbl = QLabel(label)
        lbl.setStyleSheet(
            f'font-size:10px;font-weight:700;color:{TEXT2};letter-spacing:0.5px;'
            f'background:transparent;border:none;'
        )
        lay.addWidget(lbl)

    def set_value(self, v: int):
        self.num.setText(str(v))

    def get_value(self) -> int:
        try:
            return int(self.num.text())
        except ValueError:
            return 0


class BrowserStateKart(QFrame):
    """4. stat kart — tarayıcı görünürlük durumu + tıklanabilir toggle."""
    clicked = Signal()

    def __init__(self, parent=None):
        super().__init__(parent)
        self._running = False
        self._visible = False
        self._pulse_on = True
        self.setCursor(Qt.PointingHandCursor)
        self.setToolTip('Tıkla: Tarayıcıyı göster / gizle')

        lay = QVBoxLayout(self)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(6)

        top = QHBoxLayout()
        self._ico = QLabel('🔒')
        self._ico.setStyleSheet('font-size:22px;background:transparent;border:none;')
        top.addWidget(self._ico)
        top.addStretch()
        lay.addLayout(top)

        self._num = QLabel()
        self._num.setStyleSheet('font-size:16px;font-weight:900;background:transparent;border:none;')
        lay.addWidget(self._num)

        self._lbl = QLabel()
        self._lbl.setStyleSheet('font-size:10px;font-weight:700;letter-spacing:0.5px;background:transparent;border:none;')
        lay.addWidget(self._lbl)

        self._pulse_timer = QTimer(self)
        self._pulse_timer.setInterval(600)
        self._pulse_timer.timeout.connect(self._pulse_step)

        self._refresh()

    def _card_bg(self) -> str:
        from core import theme as _t
        return _t.DARK_CARD if _t.is_dark() else CARD

    def _card_border(self) -> str:
        from core import theme as _t
        return _t.DARK_BORDER if _t.is_dark() else BORDER

    def _card_text2(self) -> str:
        from core import theme as _t
        return _t.DARK_TEXT2 if _t.is_dark() else TEXT2

    def _refresh(self):
        if not self._running and not self._visible:
            bg, border, color, icon, text = (
                self._card_bg(), self._card_border(), self._card_text2(), '🔒', 'GİZLİ MOD'
            )
        elif not self._running and self._visible:
            bg, border, color, icon, text = (
                self._card_bg(), GREEN, GREEN, '👁', 'GÖRÜNÜR MOD'
            )
        elif self._running and not self._visible:
            bg, border, color, icon, text = (
                '#0F1E33', BLUE, BLUE, '🔒', 'ARKA PLANDA'
            )
        else:  # running + visible
            bg, border, color, icon, text = (
                '#0F2A1E', GREEN, GREEN, '👁', 'CANLI'
            )
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:14px;border:1.5px solid {border};}}')
        self._ico.setText(icon)
        self._num.setText(icon)
        self._num.setStyleSheet(f'font-size:22px;font-weight:900;color:{color};background:transparent;border:none;')
        self._lbl.setText(text)
        self._lbl.setStyleSheet(f'font-size:10px;font-weight:700;color:{color};letter-spacing:0.5px;background:transparent;border:none;')

    def _pulse_step(self):
        if not self._running:
            return
        bg    = '#0F1E33' if not self._visible else '#0F2A1E'
        color = BLUE      if not self._visible else GREEN
        border = color if self._pulse_on else bg
        self.setStyleSheet(f'QFrame{{background:{bg};border-radius:14px;border:2px solid {border};}}')
        self._pulse_on = not self._pulse_on

    def set_state(self, running: bool, visible: bool):
        self._running = running
        self._visible = visible
        self._refresh()
        if running:
            self._pulse_timer.start()
        else:
            self._pulse_timer.stop()
            self._pulse_on = True

    def mousePressEvent(self, e):
        self.clicked.emit()
        super().mousePressEvent(e)


class StatBar(QFrame):
    browser_toggle = Signal()   # kart tıklandığında

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setStyleSheet('QFrame{background:transparent;border:none;}')
        lay = QHBoxLayout(self)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)
        self._k_toplam      = _StatKart('📋', 'TOPLAM FİRMA',  BLUE,  self)
        self._k_tamamlanan  = _StatKart('✓',  'TAMAMLANAN',    GREEN, self)
        self._k_hata        = _StatKart('✗',  'HATA',          RED,   self)
        self._k_browser     = BrowserStateKart(self)
        self._k_browser.clicked.connect(self.browser_toggle)
        lay.addWidget(self._k_toplam,     1)
        lay.addWidget(self._k_tamamlanan, 1)
        lay.addWidget(self._k_hata,       1)
        lay.addWidget(self._k_browser,    1)

    def reset(self, toplam: int = 0):
        self._k_toplam.set_value(toplam)
        self._k_tamamlanan.set_value(0)
        self._k_hata.set_value(0)

    def set_browser_state(self, running: bool, visible: bool):
        self._k_browser.set_state(running, visible)

    def inc_tamamlanan(self):
        self._k_tamamlanan.set_value(self._k_tamamlanan.get_value() + 1)

    def inc_hata(self):
        self._k_hata.set_value(self._k_hata.get_value() + 1)

    def get_tamamlanan(self) -> int:
        return self._k_tamamlanan.get_value()

    def get_hata(self) -> int:
        return self._k_hata.get_value()


# ── ToastNotification ─────────────────────────────────────────────────────────
class ToastNotification(QFrame):
    _LEVELS = {
        'success': (GREEN,  '#F0FDF4', '✓'),
        'error':   (RED,    '#FEF2F2', '✗'),
        'info':    (BLUE,   '#EFF6FF', 'ℹ'),
        'warning': (ORANGE, '#FFF7ED', '⚠'),
    }

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setFixedWidth(310)
        self.hide()

        self._eff = QGraphicsOpacityEffect(self)
        self._eff.setOpacity(1.0)
        self.setGraphicsEffect(self._eff)

        self._hold_timer = QTimer(self)
        self._hold_timer.setSingleShot(True)
        self._hold_timer.timeout.connect(self._start_fade)

        self._fade_timer = QTimer(self)
        self._fade_timer.setInterval(25)
        self._fade_timer.timeout.connect(self._fade_step)

        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 12)
        lay.setSpacing(10)

        self._ico_lbl = QLabel()
        self._ico_lbl.setFixedSize(30, 30)
        self._ico_lbl.setAlignment(Qt.AlignCenter)
        lay.addWidget(self._ico_lbl)

        self._msg_lbl = QLabel()
        self._msg_lbl.setWordWrap(True)
        self._msg_lbl.setMaximumWidth(230)
        lay.addWidget(self._msg_lbl, 1)

    def show_toast(self, msg: str, level: str = 'info', duration_ms: int = 3500):
        color, bg, icon = self._LEVELS.get(level, self._LEVELS['info'])
        self.setStyleSheet(f'''
            QFrame {{
                background: {bg};
                border-radius: 14px;
                border: 1.5px solid {color};
            }}
        ''')
        self._ico_lbl.setText(icon)
        self._ico_lbl.setStyleSheet(
            f'background:{color}33;border-radius:15px;font-size:16px;'
            f'font-weight:700;color:{color};border:none;'
        )
        self._msg_lbl.setText(msg)
        self._msg_lbl.setStyleSheet(
            f'font-size:12px;font-weight:600;color:{TEXT};background:transparent;border:none;'
        )
        self._eff.setOpacity(1.0)
        self._fade_timer.stop()
        self._hold_timer.stop()
        self.adjustSize()
        self._reposition()
        self.show()
        self.raise_()
        self._hold_timer.start(duration_ms)

    def _reposition(self):
        p = self.parent()
        if p:
            x = p.width()  - self.width()  - 18
            y = p.height() - self.height() - 18
            self.move(x, y)

    def _start_fade(self):
        self._fade_timer.start()

    def _fade_step(self):
        op = max(0.0, self._eff.opacity() - 0.07)
        self._eff.setOpacity(op)
        if op <= 0:
            self._fade_timer.stop()
            self.hide()


# ══════════════════════════════════════════════════════════════════════════════
class MainWindow(QMainWindow):
    def __init__(self, expire_date=None, trial_status=None):
        super().__init__()
        self.expire_date   = expire_date
        self._trial_status = trial_status
        self.worker: KarsitWorker | None = None
        self._is_running   = False
        self._is_paused    = False
        self._batch_id: str | None = None
        self._active_captcha_token: str | None = None
        self._word_sonuclar: list = []
        self._rapor_data: list = []        # [{firma, tutanak, status, hata}]
        self._last_aktif_idx: int = -1     # spinner + rapor için
        self._browser_visible: bool = False  # başlangıç: gizli
        self._spinner_idx: int = 0
        self._spinner_timer: QTimer | None = None

        self._profile_popup = ProfilePopup(self)

        self.setWindowTitle('E-YMM Karşıt İnceleme Otomasyonu')
        screen = QApplication.primaryScreen().availableGeometry()
        h = int(screen.height() * 0.95)
        w = 980
        self.setMinimumSize(820, 600)
        self.setGeometry(
            screen.x() + (screen.width() - w) // 2,
            screen.y() + (screen.height() - h) // 2,
            w, h
        )

        init_db()
        self._cfg = _cfg_load()
        self._build()
        self._fill_from_config()
        from core import theme as _theme
        _theme.register(self._apply_theme)

    def _card(self) -> QFrame:
        f = QFrame()
        f.setStyleSheet(f'QFrame{{background:{CARD};border-radius:16px;border:1px solid {BORDER};}}')
        if not hasattr(self, '_theme_cards'):
            self._theme_cards = []
        self._theme_cards.append(f)
        return f

    def _build(self):
        content = QWidget()
        content.setStyleSheet(f'background:{BG};')
        root = QVBoxLayout(content)
        root.setContentsMargins(14, 14, 14, 10)
        root.setSpacing(10)

        root.addWidget(self._header())
        root.addWidget(self._trial_banner())
        self.stat_bar = StatBar()
        self.stat_bar.browser_toggle.connect(self._on_browser_toggle)
        root.addWidget(self.stat_bar)
        root.addWidget(self._form_card())
        root.addWidget(self._button_row())

        # Inline CAPTCHA panel (başta gizli)
        self.captcha_card = self._build_captcha_card()
        root.addWidget(self.captcha_card)

        bottom = QHBoxLayout()
        bottom.setSpacing(10)
        self.log_panel   = LogPanel()
        self.firma_panel = FirmaListPanel()
        self.log_panel.setMinimumHeight(260)
        self.firma_panel.setMinimumHeight(260)
        bottom.addWidget(self.log_panel, 1)
        bottom.addWidget(self.firma_panel, 1)
        root.addLayout(bottom, 1)
        root.addWidget(self._footer())

        scroll = QScrollArea()
        scroll.setWidget(content)
        scroll.setWidgetResizable(True)
        scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
        scroll.setFocusPolicy(Qt.NoFocus)
        scroll.setVerticalScrollBarPolicy(Qt.ScrollBarAsNeeded)
        scroll.setStyleSheet(f'''
            QScrollArea {{ border:none; background:{BG}; }}
            QScrollBar:vertical {{ background:{BG}; width:6px; border:none; }}
            QScrollBar::handle:vertical {{ background:{BORDER}; border-radius:3px; min-height:30px; }}
            QScrollBar::handle:vertical:hover {{ background:{TEXT3}; }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{ height:0; }}
        ''')
        self._root_widget = scroll
        self.setCentralWidget(scroll)
        # Toast: ana pencere üzerine float (centralWidget değil self)
        self._toast = ToastNotification(self)

    # ── Inline CAPTCHA Paneli ─────────────────────────────────────────────────
    def _build_captcha_card(self) -> QFrame:
        card = QFrame()
        card.setStyleSheet(f'''
            QFrame {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1E3A2E, stop:1 #162D4E);
                border-radius: 14px;
                border: 1.5px solid {GOLD};
            }}
        ''')
        glow = QGraphicsDropShadowEffect()
        glow.setBlurRadius(20)
        glow.setColor(QColor('#C9A46A55'))
        glow.setOffset(0, 2)
        card.setGraphicsEffect(glow)

        lay = QHBoxLayout(card)
        lay.setContentsMargins(18, 12, 18, 12)
        lay.setSpacing(16)

        # Sol: başlık + açıklama
        left = QVBoxLayout()
        left.setSpacing(2)
        ttl = QLabel('🔐  GİB Güvenlik Doğrulaması')
        ttl.setStyleSheet(
            f'font-size:13px;font-weight:700;color:{GOLD};background:transparent;border:none;'
        )
        desc = QLabel('OCR okuyamadı — kodu görüntüden girin')
        desc.setStyleSheet('font-size:11px;color:#94A3B8;background:transparent;border:none;')
        left.addWidget(ttl)
        left.addWidget(desc)

        # CAPTCHA görseli
        self._cap_img = QLabel()
        self._cap_img.setFixedSize(180, 52)
        self._cap_img.setAlignment(Qt.AlignCenter)
        self._cap_img.setStyleSheet(
            f'background:#F8FAFC;border:1.5px solid {BORDER};'
            f'border-radius:8px;color:{TEXT3};font-size:11px;'
        )
        self._cap_img.setText('...')

        # Countdown
        self._cap_countdown = QLabel('120s')
        self._cap_countdown.setFixedWidth(44)
        self._cap_countdown.setAlignment(Qt.AlignCenter)
        self._cap_countdown.setStyleSheet(
            f'font-size:18px;font-weight:700;color:{ORANGE};'
            f'background:transparent;border:none;'
        )

        # Input
        self._cap_inp = QLineEdit()
        self._cap_inp.setPlaceholderText('Kodu girin...')
        self._cap_inp.setFixedSize(130, 42)
        self._cap_inp.setAlignment(Qt.AlignCenter)
        self._cap_inp.setStyleSheet(f'''
            QLineEdit {{
                background:#FFFFFF; border:1.5px solid {BORDER};
                border-radius:10px; font-size:16px; font-weight:700;
                color:{TEXT}; letter-spacing:4px;
            }}
            QLineEdit:focus {{ border-color:{GOLD}; }}
            QLineEdit::placeholder {{
                font-size:12px; font-weight:400; letter-spacing:0; color:{TEXT3};
            }}
        ''')
        self._cap_inp.returnPressed.connect(self._captcha_confirm)

        # Yenile butonu
        btn_yenile = QPushButton('↻  Yenile')
        btn_yenile.setFixedSize(82, 42)
        btn_yenile.setCursor(Qt.PointingHandCursor)
        btn_yenile.setStyleSheet(f'''
            QPushButton {{
                background:{NAVY}; color:{GOLD};
                border:1.5px solid {GOLD}; border-radius:10px;
                font-size:12px; font-weight:700;
            }}
            QPushButton:hover {{ background:#1E3A5F; }}
        ''')
        btn_yenile.clicked.connect(self._captcha_yenile)

        # Onayla butonu
        btn_ok = QPushButton('✓  Onayla')
        btn_ok.setFixedSize(96, 42)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setStyleSheet(f'''
            QPushButton {{
                background:{GREEN}; color:#FFFFFF;
                border:none; border-radius:10px;
                font-size:12px; font-weight:700;
            }}
            QPushButton:hover {{ background:{GREEN2}; }}
        ''')
        btn_ok.clicked.connect(self._captcha_confirm)

        # Layout birleştir
        lay.addLayout(left)
        lay.addStretch()
        lay.addWidget(self._cap_img)
        lay.addWidget(self._cap_countdown)
        lay.addWidget(self._cap_inp)
        lay.addWidget(btn_yenile)
        lay.addWidget(btn_ok)

        # Countdown timer
        self._cap_timer = QTimer(self)
        self._cap_timer.timeout.connect(self._captcha_tick)
        self._cap_remaining = 120

        card.setVisible(False)
        return card

    def _captcha_show(self, img_bytes: bytes, token: str):
        self._active_captcha_token = token
        self._cap_remaining = 120
        self._cap_countdown.setText('120s')
        self._cap_countdown.setStyleSheet(
            f'font-size:18px;font-weight:700;color:{ORANGE};background:transparent;border:none;'
        )
        self._cap_inp.clear()

        pix = QPixmap()
        pix.loadFromData(img_bytes)
        if not pix.isNull():
            self._cap_img.setPixmap(
                pix.scaled(180, 52, Qt.KeepAspectRatio, Qt.SmoothTransformation)
            )
        else:
            self._cap_img.setText('Görüntü yok')

        self.captcha_card.setVisible(True)
        self._cap_timer.start(1000)
        self._cap_inp.setFocus()

    def _captcha_hide(self):
        self._cap_timer.stop()
        self.captcha_card.setVisible(False)
        self._active_captcha_token = None

    def _captcha_yenile(self):
        """Yenile butonuna basılınca boş kod gönder — engine yeni CAPTCHA çeker."""
        token = self._active_captcha_token
        self._captcha_hide()
        if self.worker and token:
            self.worker.submit_captcha('', token)
        self.log_panel.append('CAPTCHA yenileniyor...', ORANGE)

    def _captcha_confirm(self):
        kod = self._cap_inp.text().strip()
        if not kod:
            self._cap_inp.setStyleSheet(
                self._cap_inp.styleSheet() + f'QLineEdit{{border-color:{RED};}}'
            )
            return
        token = self._active_captcha_token
        self._captcha_hide()
        if self.worker and token:
            self.worker.submit_captcha(kod, token)
        self.log_panel.append('CAPTCHA gönderildi', BLUE)

    def _captcha_tick(self):
        self._cap_remaining -= 1
        self._cap_countdown.setText(f'{self._cap_remaining}s')
        if self._cap_remaining <= 30:
            self._cap_countdown.setStyleSheet(
                f'font-size:18px;font-weight:700;color:{RED};background:transparent;border:none;'
            )
        if self._cap_remaining <= 0:
            token = self._active_captcha_token
            self._captcha_hide()
            if self.worker and token:
                self.worker.submit_captcha('', token)
            self.log_panel.append('CAPTCHA zaman aşımı — yenileniyor', ORANGE)

    def _header(self) -> QFrame:
        from core import _icons as _ic_logo
        f = self._card()
        lay = QHBoxLayout(f)
        lay.setContentsMargins(18, 12, 18, 12)
        lay.setSpacing(16)

        # ── SOL: ContraCore logo + başlık ─────────────────────────────────
        left = QHBoxLayout()
        left.setSpacing(12)

        self._lbl_logo = QLabel()
        self._lbl_logo.setStyleSheet('background:transparent;border:none;')
        self._logo_h = 76
        self._update_logo()
        left.addWidget(self._lbl_logo)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.setContentsMargins(0, 0, 0, 0)

        title_row = QHBoxLayout()
        title_row.setSpacing(0)
        title_row.setContentsMargins(0, 0, 0, 0)
        self._cc_lbl = QLabel()
        self._cc_lbl.setStyleSheet('background:transparent;border:none;')
        self._cc_lbl.setPixmap(_cc_pix(24, dark=False))
        title_row.addWidget(self._cc_lbl)
        title_row.addStretch()
        title_col.addLayout(title_row)

        sub_lbl = QLabel('E-YMM Karşıt İnceleme Otomasyonu')
        sub_lbl.setFont(QFont('Segoe UI', 14, QFont.Bold))
        sub_lbl.setStyleSheet(
            f'color:{TEXT};background:transparent;border:none;letter-spacing:0.3px;'
        )
        title_col.addWidget(sub_lbl)
        left.addLayout(title_col)
        lay.addLayout(left)
        lay.addStretch(1)

        # ── SAĞ: Profil ───────────────────────────────────────────────────
        prof = QFrame()
        prof.setStyleSheet('''
            QFrame {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1E3660, stop:1 #162D4E);
                border-radius: 32px;
                border: 1.5px solid #C9A46A;
            }
            QFrame:hover {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #253F6A, stop:1 #1E3660);
                border-color: #E4C285;
            }
        ''')
        _prof_glow = QGraphicsDropShadowEffect()
        _prof_glow.setBlurRadius(18)
        _prof_glow.setColor(QColor('#C9A46A55'))
        _prof_glow.setOffset(0, 2)
        prof.setGraphicsEffect(_prof_glow)

        pl = QHBoxLayout(prof)
        pl.setContentsMargins(14, 7, 14, 7)
        pl.setSpacing(10)

        p_img = QLabel()
        p_img.setFixedSize(46, 46)
        p_img.setStyleSheet('background:transparent;border:none;')
        px2 = _ic_logo.load('profile.png')
        if not px2.isNull():
            canvas = QPixmap(46, 46)
            canvas.fill(Qt.transparent)
            pc = QPainter(canvas)
            pc.setRenderHint(QPainter.Antialiasing)
            scaled = px2.scaled(QSize(40, 40), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            clip = QPainterPath()
            clip.addEllipse(3.0, 3.0, 40.0, 40.0)
            pc.setClipPath(clip)
            pc.drawPixmap(3, 3, scaled)
            pc.setClipping(False)
            pc.setPen(QPen(QColor(GOLD), 2))
            pc.setBrush(Qt.NoBrush)
            pc.drawEllipse(QRectF(1, 1, 44, 44))
            from datetime import datetime as _dt2
            _now2 = _dt2.now()
            _online = (_now2.weekday() <= 5) and (9 <= _now2.hour < 17)
            pc.setPen(QPen(QColor('#FFFFFF'), 1.5))
            pc.setBrush(QColor(GREEN if _online else RED))
            pc.drawEllipse(33, 33, 11, 11)
            pc.end()
            p_img.setPixmap(canvas)
        else:
            p_img.setText('👤')
        pl.addWidget(p_img)

        pt = QVBoxLayout()
        pt.setSpacing(1)
        pt.setContentsMargins(0, 0, 0, 0)
        n = QLabel('Serkan ŞAHİN')
        n.setStyleSheet(
            'font-size:12px;font-weight:700;color:#FFFFFF;background:transparent;border:none;'
        )
        r = QLabel('Geliştirici İletişim')
        r.setStyleSheet(f'font-size:10px;color:{GOLD};background:transparent;border:none;')
        pt.addWidget(n)
        pt.addWidget(r)
        pl.addLayout(pt)

        arr = QLabel('▾')
        arr.setStyleSheet('font-size:11px;color:#C9A46A;background:transparent;border:none;')
        pl.addWidget(arr)

        prof.setCursor(Qt.PointingHandCursor)
        def _toggle_prof(e):
            if self._profile_popup._just_closed:
                return
            if self._profile_popup.isVisible():
                self._profile_popup.close()
            else:
                self._profile_popup.show_at(prof)
        prof.mousePressEvent = _toggle_prof
        lay.addWidget(prof)

        # Dark/Light mod toggle butonu
        self.btn_theme = QPushButton('🌙')
        self.btn_theme.setFixedSize(38, 38)
        self.btn_theme.setCursor(Qt.PointingHandCursor)
        self.btn_theme.setToolTip('Koyu / Açık mod geçişi')
        self.btn_theme.setStyleSheet(f'''
            QPushButton {{
                background: rgba(255,255,255,0.08);
                border: 1px solid rgba(255,255,255,0.15);
                border-radius: 19px;
                font-size: 16px;
            }}
            QPushButton:hover {{
                background: rgba(255,255,255,0.18);
                border-color: rgba(255,255,255,0.35);
            }}
        ''')
        self.btn_theme.clicked.connect(self._toggle_theme)
        lay.addWidget(self.btn_theme)
        return f

    def _trial_banner(self) -> QFrame:
        self.banner_trial = QFrame()
        self.banner_trial.setFixedHeight(38)
        self.banner_trial.setStyleSheet('QFrame{background:#7C3AED;border-radius:10px;border:none;}')
        lay = QHBoxLayout(self.banner_trial)
        lay.setContentsMargins(18, 0, 18, 0)
        lay.setSpacing(12)
        ico = QLabel('⏳')
        ico.setStyleSheet('background:transparent;border:none;font-size:14px;')
        lay.addWidget(ico)
        self.lbl_trial = QLabel()
        self.lbl_trial.setStyleSheet(
            'font-size:12px;font-weight:700;color:#FFFFFF;background:transparent;border:none;'
        )
        lay.addWidget(self.lbl_trial)
        lay.addStretch(1)
        tip = QLabel('Lisans almak için geliştiriciye ulaşın')
        tip.setStyleSheet('font-size:10px;color:#DDD6FE;background:transparent;border:none;')
        lay.addWidget(tip)

        if self._trial_status:
            kalan_gun, islenen, kalan = self._trial_status
            self.lbl_trial.setText(
                f'Deneme Sürümü  —  {kalan_gun} gün kaldı  |  {islenen} tutanak işlendi'
            )
            self.banner_trial.setVisible(True)
        else:
            self.banner_trial.setVisible(False)
        return self.banner_trial

    def _form_card(self) -> QFrame:
        f = self._card()
        lay = QVBoxLayout(f)
        lay.setContentsMargins(18, 16, 18, 16)
        lay.setSpacing(0)

        # ── YMM Profili ───────────────────────────────────────────────────────
        lay.addWidget(_section_hdr('YMM PROFİLİ', 'user.png'))
        lay.addSpacing(10)

        prof_row = QHBoxLayout()
        prof_row.setSpacing(8)

        self.cmb_ymm_profil = QComboBox()
        self.cmb_ymm_profil.setFixedHeight(48)
        self.cmb_ymm_profil.setCursor(Qt.PointingHandCursor)
        self.cmb_ymm_profil.setStyleSheet(f'''
            QComboBox {{
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #1A2E50, stop:1 {NAVY});
                border: 1.5px solid rgba(201,164,106,0.5);
                border-radius: 14px;
                padding: 0 18px;
                font-size: 14px;
                font-weight: 700;
                color: #FFFFFF;
            }}
            QComboBox:hover {{
                border-color: {GOLD};
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #243A5E, stop:1 #162D4E);
            }}
            QComboBox::drop-down {{
                border: none; width: 40px;
            }}
            QComboBox::down-arrow {{
                image: none;
                border-left: 5px solid transparent;
                border-right: 5px solid transparent;
                border-top: 6px solid {GOLD};
                width: 0; height: 0; margin-right: 14px;
            }}
            QComboBox QAbstractItemView {{
                background: {NAVY};
                border: 1.5px solid rgba(201,164,106,0.4);
                border-radius: 12px;
                color: #E2E8F0;
                font-size: 13px;
                font-weight: 600;
                padding: 4px;
                outline: none;
            }}
            QComboBox QAbstractItemView::item {{
                height: 44px;
                padding-left: 16px;
                border-radius: 8px;
                color: #E2E8F0;
            }}
            QComboBox QAbstractItemView::item:hover {{
                background: rgba(201,164,106,0.2);
                color: {GOLD};
            }}
            QComboBox QAbstractItemView::item:selected {{
                background: rgba(201,164,106,0.2);
                color: {GOLD};
            }}
        ''')
        glow_cmb = QGraphicsDropShadowEffect()
        glow_cmb.setBlurRadius(16)
        glow_cmb.setColor(QColor('#C9A46A33'))
        glow_cmb.setOffset(0, 2)
        self.cmb_ymm_profil.setGraphicsEffect(glow_cmb)
        _cmb_no_scroll(self.cmb_ymm_profil)
        self.cmb_ymm_profil.currentIndexChanged.connect(self._on_profil_secildi)
        prof_row.addWidget(self.cmb_ymm_profil, 1)

        btn_kaydet_ic = QPushButton('＋')
        btn_kaydet_ic.setFixedSize(48, 48)
        btn_kaydet_ic.setToolTip('Mevcut bilgileri profil olarak kaydet')
        btn_kaydet_ic.setCursor(Qt.PointingHandCursor)
        btn_kaydet_ic.setStyleSheet(f'''
            QPushButton {{
                background:{NAVY}; border:1.5px solid rgba(201,164,106,0.4);
                border-radius:12px; color:{GOLD}; font-size:20px; font-weight:300;
            }}
            QPushButton:hover {{ background:{NAVY2}; border-color:{GOLD}; }}
            QPushButton:pressed {{ background:#081426; }}
        ''')
        btn_kaydet_ic.clicked.connect(self._profil_kaydet_popup)
        prof_row.addWidget(btn_kaydet_ic)

        btn_sil_ic = QPushButton('🗑')
        btn_sil_ic.setFixedSize(48, 48)
        btn_sil_ic.setToolTip('Seçili profili sil')
        btn_sil_ic.setCursor(Qt.PointingHandCursor)
        btn_sil_ic.setStyleSheet(f'''
            QPushButton {{
                background:#FEF2F2; border:1.5px solid #FECACA;
                border-radius:12px; font-size:15px;
            }}
            QPushButton:hover {{ background:#FEE2E2; border-color:{RED}; }}
            QPushButton:pressed {{ background:#FCA5A5; }}
        ''')
        btn_sil_ic.clicked.connect(self._profil_sil)
        prof_row.addWidget(btn_sil_ic)
        lay.addLayout(prof_row)

        # Profil önizleme kartı
        self.prof_preview = QFrame()
        self.prof_preview.setFixedHeight(0)  # collapsed by default
        self.prof_preview.setStyleSheet(
            f'QFrame{{background:rgba(11,31,58,0.06);border-radius:10px;border:none;}}'
        )
        pv_lay = QHBoxLayout(self.prof_preview)
        pv_lay.setContentsMargins(14, 0, 14, 0)
        pv_lay.setSpacing(20)
        self._pv_kul = QLabel()
        self._pv_kul.setStyleSheet(f'font-size:11px;color:{TEXT2};background:transparent;border:none;')
        self._pv_tel = QLabel()
        self._pv_tel.setStyleSheet(f'font-size:11px;color:{TEXT2};background:transparent;border:none;')
        pv_lay.addWidget(self._pv_kul)
        pv_lay.addWidget(self._pv_tel)
        pv_lay.addStretch()
        lay.addWidget(self.prof_preview)
        lay.addSpacing(16)

        # ── Giriş bilgileri ───────────────────────────────────────────────────
        sep0 = QFrame(); sep0.setFrameShape(QFrame.HLine)
        sep0.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
        lay.addWidget(sep0)
        lay.addSpacing(14)

        lay.addWidget(_section_hdr('GİRİŞ BİLGİLERİ', 'karsi.png'))
        lay.addSpacing(10)

        grid = QGridLayout()
        grid.setSpacing(12)
        grid.setContentsMargins(0, 0, 0, 0)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)

        self.inp_kul     = InputRow('KULLANICI KODU', '68xxxxxxx', icon_file='user.png',
                                    validator=_val_kul_kodu)
        self.inp_sif     = InputRow('ŞİFRE', '••••••••', echo_password=True,
                                    icon_file='password.png')
        self.inp_tel     = InputRow('YMM TELEFON NUMARASI', '(531) 087 93 39',
                                    icon_file='mobile.png', phone_mask=True,
                                    validator=_val_telefon)
        self.inp_muk_tel = InputRow('MÜKELLEF TELEFON NUMARASI', '(531) 087 93 39',
                                    icon_file='mobile.png', phone_mask=True,
                                    validator=_val_telefon)
        self.inp_tutanak_no = InputRow('TUTANAK SAYISI', 'ör: 2026/0012',
                                       icon_file='hashtag.png',
                                       validator=_val_tutanak)

        grid.addWidget(self.inp_kul,        0, 0)
        grid.addWidget(self.inp_sif,        0, 1)
        grid.addWidget(self.inp_tel,        1, 0)
        grid.addWidget(self.inp_muk_tel,    1, 1)
        grid.addWidget(self.inp_tutanak_no, 2, 0, 1, 2)
        lay.addLayout(grid)
        lay.addSpacing(14)

        # ── Dosya seçimi ──────────────────────────────────────────────────────
        sep2 = QFrame(); sep2.setFrameShape(QFrame.HLine)
        sep2.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
        lay.addWidget(sep2)
        lay.addSpacing(14)

        # Dosya Seçimi bölümü — renkli frame
        dosya_frame = QFrame()
        dosya_frame.setStyleSheet(f'QFrame{{background:#EFF6FF;border-radius:12px;border:1.5px solid #BFDBFE;}}')
        dosya_inner = QVBoxLayout(dosya_frame)
        dosya_inner.setContentsMargins(14, 12, 14, 12)
        dosya_inner.setSpacing(10)
        dosya_inner.addWidget(_section_hdr('DOSYA SEÇİMİ', 'klasor.png'))

        self.drop_word = DropFolderZone()
        self.drop_word.changed.connect(self._on_word_klasor_secildi)
        dosya_inner.addWidget(self.drop_word)
        lay.addWidget(dosya_frame)
        lay.addSpacing(14)

        # ── Tasdik Ayarları ───────────────────────────────────────────────────
        sep3 = QFrame(); sep3.setFrameShape(QFrame.HLine)
        sep3.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
        lay.addWidget(sep3)
        lay.addSpacing(14)
        lay.addWidget(_section_hdr('TASDİK AYARLARI', 'karsi.png'))
        lay.addSpacing(10)

        def _cmb_soz(opts):
            c = QComboBox()
            c.addItems(opts)
            c.setFixedHeight(48)
            c.setCursor(Qt.PointingHandCursor)
            c.setStyleSheet('''
                QComboBox {
                    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                        stop:0 #134E4A, stop:1 #0F3D39);
                    border: 1.5px solid rgba(52,211,153,0.5);
                    border-radius: 14px; padding: 0 18px;
                    font-size: 14px; font-weight: 700; color: #FFFFFF;
                }
                QComboBox:hover {
                    border-color: #34D399;
                    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                        stop:0 #1A5C57, stop:1 #134E4A);
                }
                QComboBox::drop-down { border:none; width:40px; }
                QComboBox::down-arrow {
                    border-left:5px solid transparent;
                    border-right:5px solid transparent;
                    border-top:6px solid #34D399;
                    width:0; height:0; margin-right:14px;
                }
                QComboBox QAbstractItemView {
                    background: #0F3D39;
                    border: 1.5px solid #34D399;
                    border-radius: 8px;
                    color: #E2F8F0;
                    font-size: 13px; font-weight: 600;
                    padding: 4px; outline: none;
                    selection-background-color: #059669;
                    selection-color: #FFFFFF;
                }
                QComboBox QAbstractItemView::item {
                    height: 44px; padding-left: 14px;
                    border-radius: 8px; color: #E2F8F0;
                }
                QComboBox QAbstractItemView::item:hover {
                    background: #1A5C57;
                }
                QComboBox QAbstractItemView::item:selected {
                    background: #059669; color: #FFFFFF;
                }
            ''')
            return c

        def _cmb_kdv(opts):
            c = QComboBox()
            c.addItems(opts)
            c.setFixedHeight(48)
            c.setCursor(Qt.PointingHandCursor)
            c.setStyleSheet('''
                QComboBox {
                    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                        stop:0 #f2af20, stop:1 #c98a00);
                    border: 1.5px solid #f2af20;
                    border-radius: 14px; padding: 0 18px;
                    font-size: 14px; font-weight: 700; color: #1A1200;
                }
                QComboBox:hover {
                    border-color: #d99800;
                    background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                        stop:0 #f5bb3a, stop:1 #d99800);
                }
                QComboBox::drop-down { border:none; width:40px; }
                QComboBox::down-arrow {
                    border-left:5px solid transparent;
                    border-right:5px solid transparent;
                    border-top:6px solid #1A1200;
                    width:0; height:0; margin-right:14px;
                }
                QComboBox QAbstractItemView {
                    background: #f2af20;
                    border: 1.5px solid #c98a00;
                    border-radius: 8px;
                    color: #1A1200;
                    font-size: 13px; font-weight: 600;
                    padding: 4px; outline: none;
                    selection-background-color: #c98a00;
                    selection-color: #FFFFFF;
                }
                QComboBox QAbstractItemView::item {
                    height: 44px; padding-left: 14px;
                    border-radius: 8px; color: #1A1200;
                }
                QComboBox QAbstractItemView::item:hover {
                    background: #7B5200; color: #FFFFFF;
                }
                QComboBox QAbstractItemView::item:selected {
                    background: #c98a00; color: #FFFFFF;
                }
            ''')
            return c

        def _two_col(a, b):
            row = QHBoxLayout()
            row.setSpacing(10)
            row.addWidget(a, 1)
            row.addWidget(b, 1)
            return row

        def _sep_line():
            s = QFrame(); s.setFrameShape(QFrame.HLine)
            s.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
            return s

        # ── BÖLÜM 1: SÖZLEŞME TÜRÜ ───────────────────────────────────────────
        soz_tur_lbl = QLabel('📄  Sözleşme Türü')
        soz_tur_lbl.setStyleSheet(f'font-size:15px;font-weight:800;color:{NAVY};background:transparent;border:none;letter-spacing:0.5px;')
        lay.addWidget(soz_tur_lbl)
        lay.addSpacing(8)
        self.cmb_tasdik_tur = _cmb_no_scroll(_cmb_soz(['KDV', 'KDV + Tam Tasdik']))
        self.cmb_tasdik_tur.currentIndexChanged.connect(self._on_tasdik_tur_changed)
        lay.addWidget(self.cmb_tasdik_tur)
        lay.addSpacing(10)

        # ── KDV Sözleşmesi Accordion ──────────────────────────────────────────
        self._acc_kdv_soz = AccordionSection(
            'KDV SÖZLEŞME BİLGİLERİ', icon_file='hashtag.png',
            initial_open=True, accent=BLUE
        )
        self.inp_kdv_soz_tarih = InputRow('KDV SÖZLEŞME TARİHİ',     'GG.AA.YYYY', icon_file='calendar.png',
                                          date_mask=True, validator=_val_tarih)
        self.inp_kdv_soz_no    = InputRow('KDV SÖZLEŞME NO',          'ör: E-65526', icon_file='hashtag.png',
                                          validator=_val_sozlesme_no)
        self.inp_kdv_soz_giris = InputRow('KDV SİSTEME GİRİŞ TARİHİ','GG.AA.YYYY', icon_file='calendar.png',
                                          date_mask=True, validator=_val_tarih)
        self._acc_kdv_soz.add_layout(_two_col(self.inp_kdv_soz_tarih, self.inp_kdv_soz_no))
        self._acc_kdv_soz.add_widget(self.inp_kdv_soz_giris)
        lay.addWidget(self._acc_kdv_soz)
        lay.addSpacing(6)

        # ── Tam Tasdik Accordion (sadece "KDV + Tam Tasdik" seçilince görünür) ─
        self._acc_tam = AccordionSection(
            'TAM TASDİK BİLGİLERİ', icon_file='time.png',
            initial_open=True, accent='#15803D'
        )
        self.inp_tam_bas       = InputRow('TAM TASDİK BAŞLANGIÇ',       'AA.YYYY', icon_file='time.png',
                                          period_mask=True, validator=_val_donem)
        self.inp_tam_bit       = InputRow('TAM TASDİK BİTİŞ',           'AA.YYYY', icon_file='time.png',
                                          period_mask=True, validator=_val_donem)
        self.inp_tam_soz_tarih = InputRow('TAM TASDİK SÖZLEŞME TARİHİ','GG.AA.YYYY', icon_file='calendar.png',
                                          date_mask=True, validator=_val_tarih)
        self.inp_tam_soz_no    = InputRow('TAM TASDİK SÖZLEŞME NO',    'ör: 35-A-32035', icon_file='hashtag.png',
                                          validator=_val_sozlesme_no)
        self.inp_tam_soz_giris = InputRow('TAM TASDİK GİRİŞ TARİHİ',  'GG.AA.YYYY', icon_file='calendar.png',
                                          date_mask=True, validator=_val_tarih)
        self._acc_tam.add_layout(_two_col(self.inp_tam_bas, self.inp_tam_bit))
        self._acc_tam.add_layout(_two_col(self.inp_tam_soz_tarih, self.inp_tam_soz_no))
        self._acc_tam.add_widget(self.inp_tam_soz_giris)
        lay.addWidget(self._acc_tam)
        self._acc_tam.setVisible(False)
        lay.addSpacing(6)

        # _tam_frame uyumluluk referansı (kod başka yerde .setVisible() çağırır)
        self._tam_frame = self._acc_tam

        # ── Ayraç ─────────────────────────────────────────────────────────────
        lay.addWidget(_sep_line())
        lay.addSpacing(10)

        # ── BÖLÜM 2: KDV TÜRÜ ─────────────────────────────────────────────────
        self._kdv_tur_lbl = QLabel('💳  KDV Türü')
        self._kdv_tur_lbl.setStyleSheet(f'font-size:15px;font-weight:800;color:{NAVY};background:transparent;border:none;letter-spacing:0.5px;')
        lay.addWidget(self._kdv_tur_lbl)
        lay.addSpacing(8)
        self.cmb_kdv_tur = _cmb_no_scroll(_cmb_kdv(['Normal KDV', 'İndirimli Oran']))
        self.cmb_kdv_tur.currentIndexChanged.connect(self._on_kdv_tur_changed)
        lay.addWidget(self.cmb_kdv_tur)
        lay.addSpacing(8)

        # ── İndirimli Oran Accordion ──────────────────────────────────────────
        self._acc_indirimli = AccordionSection(
            'İNDİRİMLİ ORAN DETAYLARI', icon_file='time.png',
            initial_open=True, accent='#D97706'
        )
        self.inp_kdv_bas  = InputRow('İADE BAŞLANGIÇ DÖNEMİ', 'AA.YYYY', icon_file='time.png',
                                     period_mask=True, validator=_val_donem)
        self.inp_kdv_bit  = InputRow('İADE BİTİŞ DÖNEMİ',    'AA.YYYY', icon_file='time.png',
                                     period_mask=True, validator=_val_donem)
        self.inp_kdv_iade = InputRow('İADE DÖNEMİ',           'AA.YYYY', icon_file='time.png',
                                     period_mask=True, validator=_val_donem)
        self._acc_indirimli.add_layout(_two_col(self.inp_kdv_bas, self.inp_kdv_bit))
        self._acc_indirimli.add_widget(self.inp_kdv_iade)
        lay.addWidget(self._acc_indirimli)
        self._acc_indirimli.setVisible(False)

        # _indirmli_frame uyumluluk referansı
        self._indirmli_frame = self._acc_indirimli

        # _kdv_soz_frame eski kod uyumluluğu için
        self._kdv_soz_frame = QFrame()
        self._kdv_soz_frame.setVisible(False)

        self._profil_listesi_yenile()
        return f

    def _button_row(self) -> QFrame:
        f = QFrame()
        f.setStyleSheet('QFrame{background:transparent;border:none;}')
        lay = QHBoxLayout(f)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(8)

        def _btn(text, color, hover, text_color='#FFFFFF'):
            b = QPushButton(text)
            b.setFixedHeight(44)
            b.setCursor(Qt.PointingHandCursor)
            b.setStyleSheet(f'''
                QPushButton {{
                    background:{color}; color:{text_color};
                    border:none; border-radius:12px;
                    font-size:12px; font-weight:700; padding:0 18px;
                }}
                QPushButton:hover {{ background:{hover}; }}
                QPushButton:pressed {{ opacity:0.85; }}
                QPushButton:disabled {{ background:#E5E7EB; color:#9CA3AF; }}
            ''')
            return b

        self.btn_start  = _btn('BAŞLAT',   GREEN,    GREEN2)
        self.btn_pause  = _btn('DURAKLAT', ORANGE,   '#EA580C')
        self.btn_stop   = _btn('İPTAL',    RED,      '#DC2626')

        _pix_power = _icon_pix('power.png', 18)
        if not _pix_power.isNull():
            self.btn_start.setIcon(QIcon(_pix_power))
            self.btn_start.setIconSize(QSize(18, 18))
        _pix_pause = _icon_pix('pause.png', 18)
        if not _pix_pause.isNull():
            self.btn_pause.setIcon(QIcon(_pix_pause))
            self.btn_pause.setIconSize(QSize(18, 18))
        _pix_cancel = _icon_pix('cancel.png', 18)
        if not _pix_cancel.isNull():
            self.btn_stop.setIcon(QIcon(_pix_cancel))
            self.btn_stop.setIconSize(QSize(18, 18))

        self.btn_pause.setEnabled(False)
        self.btn_stop.setEnabled(False)
        self._is_paused = False

        self.btn_start.clicked.connect(self._on_start)
        self.btn_pause.clicked.connect(self._on_pause_resume)
        self.btn_stop.clicked.connect(self._on_stop)

        lay.addWidget(self.btn_start, 1)
        lay.addWidget(self.btn_pause, 1)
        lay.addWidget(self.btn_stop,  1)

        self.status_chip = StatusChip()
        self.status_chip.clicked.connect(self._on_browser_toggle)
        lay.addWidget(self.status_chip)
        return f

    def _footer(self) -> QFrame:
        f = QFrame()
        f.setStyleSheet('QFrame{background:transparent;border:none;}')
        lay = QHBoxLayout(f)
        lay.setContentsMargins(4, 0, 4, 0)
        lbl = QLabel('Developed by Serkan ŞAHİN  ©  2026')
        lbl.setStyleSheet(f'font-size:11px;color:{TEXT3};background:transparent;border:none;')
        lay.addWidget(lbl)
        lay.addStretch()
        if self.expire_date:
            exp = self.expire_date.strftime('%d.%m.%Y')
            lic = QLabel(f'Lisans: {exp}')
            lic.setStyleSheet(f'font-size:11px;color:{GOLD};background:transparent;border:none;')
            lay.addWidget(lic)
        return f

    # ── Master Excel İndir ────────────────────────────────────────────────────
    def _master_excel_indir(self):
        src = _MASTER_XLSM_SRC
        if not os.path.exists(src):
            QMessageBox.warning(
                self, 'Şablon Bulunamadı',
                f'karsit_master.xlsm şablon dosyası bulunamadı:\n{src}'
            )
            return
        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        dest, _ = QFileDialog.getSaveFileName(
            self, 'Şablonu Kaydet',
            os.path.join(desktop, 'karsit_master.xlsm'),
            'Excel Dosyası (*.xlsm)'
        )
        if not dest:
            return
        try:
            import zipfile, re as _re
            def _strip_data_rows(xml_bytes: bytes) -> bytes:
                # satır 3+ içindeki <row r="N"> ... </row> bloklarını sil (N >= 3)
                text = xml_bytes.decode('utf-8')
                text = _re.sub(
                    r'<row r="([3-9]|\d{2,})"[^>]*>.*?</row>',
                    '', text, flags=_re.DOTALL
                )
                return text.encode('utf-8')

            with zipfile.ZipFile(src, 'r') as zin:
                with zipfile.ZipFile(dest, 'w', zipfile.ZIP_DEFLATED) as zout:
                    for item in zin.infolist():
                        data = zin.read(item.filename)
                        if item.filename == 'xl/worksheets/sheet1.xml':
                            data = _strip_data_rows(data)
                        zout.writestr(item, data)

            self.log_panel.append(f'Boş şablon indirildi: {os.path.basename(dest)}', GREEN)
        except Exception as e:
            QMessageBox.critical(self, 'İndirme Hatası', str(e))

    # ── Excel seçilince firma önizleme ────────────────────────────────────────
    def _on_word_klasor_secildi(self, path: str):
        """Word klasörü seçilince dosyaları tara, alanları otomatik doldur."""
        if not path or not os.path.exists(path):
            self.firma_panel.set_firmalar([])
            self.stat_bar.reset(0)
            return
        try:
            from .karsit_parser import klasor_tara
            sonuclar = klasor_tara(path)
        except Exception as e:
            self.log_panel.append(f'Klasör tarama hatası: {e}', RED)
            return

        if not sonuclar:
            self.log_panel.append('Word dosyası bulunamadı.', ORANGE)
            return

        # Firma listesi — "Mükellef / Karşıt" formatı (parser'dan), yoksa dosya adı
        import re as _re2
        def _firma_goster(s):
            muk = s.get('mukellef_adi', '')
            kar = s.get('karsit_firma_adi', '')
            if not kar:
                dosya = os.path.basename(s.get('dosya_yolu', ''))
                kar = _re2.sub(r'[-_]\d{2}\.\d{2}\.\d{4}.*$',
                               '', os.path.splitext(dosya)[0]).strip() or '?'
            if muk and muk.upper() != kar.upper():
                return f'{muk}  /  {kar}'
            return kar

        firmalar = [_firma_goster(s) for s in sonuclar]
        bilgi_istem_set = {i for i, s in enumerate(sonuclar) if s.get('bilgi_istem')}
        self._word_sonuclar = sonuclar
        self.firma_panel.set_firmalar(firmalar, bilgi_istem_set)
        gecerli = len(firmalar) - len(bilgi_istem_set)
        self.log_panel.append(f'{len(firmalar)} dosya tarandı — {gecerli} geçerli, {len(bilgi_istem_set)} Bilgi İsteme (atlanacak).', GOLD)

        # İlk dosyadan ortak sözleşme bilgilerini al
        ilk = sonuclar[0]
        if ilk.get('kdv_soz_tarih'):
            self.inp_kdv_soz_tarih.set_value(ilk['kdv_soz_tarih'])
            self.log_panel.append(f'  KDV Sözleşme Tarihi: {ilk["kdv_soz_tarih"]}', GOLD)
        # KDV Sözleşme No manuel girilir (E-65526 formatı), Word'den doldurmuyoruz
        if ilk.get('mukellef_vkn'):
            self.inp_muk_tel.set_value('')  # telefon değil, farklı alan
        if ilk.get('tam_tasdik_var'):
            self.cmb_tasdik_tur.setCurrentIndex(2)  # HER İKİSİ
            if ilk.get('tam_soz_tarih'):
                self.inp_tam_soz_tarih.set_value(ilk['tam_soz_tarih'])
            if ilk.get('tam_soz_no'):
                self.inp_tam_soz_no.set_value(ilk['tam_soz_no'])

        # Boş alanları hemen highlight et
        self._highlight_errors()

    def _on_tasdik_tur_changed(self, idx: int):
        # 0=KDV, 1=KDV+Tam Tasdik
        self._tam_frame.setVisible(idx == 1)

    def _on_kdv_tur_changed(self, idx: int):
        self._indirmli_frame.setVisible(idx == 1)

    def _on_excel_secildi(self, path: str):
        pass  # Excel kaldırıldı — V3'te kullanılmıyor

    # ── Config ────────────────────────────────────────────────────────────────
    def _fill_from_config(self):
        self.drop_word.set_value('')
        self.inp_muk_tel.set_value(self._cfg.get('muk_tel', ''))
        self.inp_tutanak_no.set_value(self._cfg.get('tutanak_no', ''))

    def _save_config(self):
        cfg = _cfg_load()
        cfg['word_klasor']  = self.drop_word.value()
        cfg['muk_tel']      = self.inp_muk_tel.value()
        cfg['tutanak_no']   = self.inp_tutanak_no.value()
        cfg['son_profil']   = self.cmb_ymm_profil.currentIndex()
        _cfg_save(cfg)

    # ── YMM Profil Yönetimi ───────────────────────────────────────────────────
    def _profiller(self) -> list:
        return _cfg_load().get('ymm_profiller', [])

    def _profil_listesi_yenile(self, sec_idx: int = -1):
        self.cmb_ymm_profil.blockSignals(True)
        self.cmb_ymm_profil.clear()
        self.cmb_ymm_profil.addItem('— Profil seçin —')
        for p in self._profiller():
            self.cmb_ymm_profil.addItem(f'  {p.get("ad", "?")}')
        if sec_idx >= 0:
            self.cmb_ymm_profil.setCurrentIndex(sec_idx + 1)
        else:
            son = _cfg_load().get('son_profil', 0)
            self.cmb_ymm_profil.setCurrentIndex(son)
        self.cmb_ymm_profil.blockSignals(False)
        self._on_profil_secildi(self.cmb_ymm_profil.currentIndex())

    def _on_profil_secildi(self, idx: int):
        if idx <= 0:
            self.prof_preview.setFixedHeight(0)
            return
        profiller = self._profiller()
        if idx - 1 >= len(profiller):
            return
        p = profiller[idx - 1]
        self.inp_kul.set_value(p.get('kullanici', ''))
        self.inp_sif.set_value(p.get('sifre', ''))
        self.inp_tel.set_value(p.get('ymm_tel', ''))

        kul = p.get('kullanici', '')
        tel = p.get('ymm_tel', '')
        kul_masked = kul[:3] + '****' + kul[-2:] if len(kul) > 5 else kul
        tel_masked = tel[:3] + '****' + tel[-2:] if len(tel) > 5 else tel
        self._pv_kul.setText(f'Kod: {kul_masked}')
        self._pv_tel.setText(f'Tel: {tel_masked}')
        self.prof_preview.setFixedHeight(32)

        cfg = _cfg_load()
        cfg['son_profil'] = idx
        _cfg_save(cfg)

    def _profil_kaydet_popup(self):
        k = self.inp_kul.value()
        s = self.inp_sif.value()
        if not k or not s:
            QMessageBox.warning(self, 'Eksik Bilgi', 'Önce kullanıcı kodu ve şifre girin.')
            return

        dlg = QDialog(self)
        dlg.setWindowTitle('YMM Profili Kaydet')
        dlg.setFixedSize(320, 140)
        dlg.setStyleSheet(f'background:{BG};')
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(20, 16, 20, 16)
        lay.setSpacing(12)

        lbl = QLabel('Profil adı girin:')
        lbl.setStyleSheet(f'font-size:13px;font-weight:600;color:{TEXT};')
        lay.addWidget(lbl)

        inp = QLineEdit()
        inp.setPlaceholderText('Ör: Ahmet YMM')
        inp.setFixedHeight(38)
        inp.setStyleSheet(f'''
            QLineEdit {{
                background:{CARD}; border:1.5px solid {BORDER};
                border-radius:10px; padding:0 12px; font-size:13px; color:{TEXT};
            }}
            QLineEdit:focus {{ border-color:{BLUE}; background:#FFFFFF; }}
            QLineEdit::placeholder {{ color:{TEXT3}; }}
        ''')
        lay.addWidget(inp)

        btn_row = QHBoxLayout()
        btn_row.setSpacing(8)
        btn_iptl = QPushButton('İptal')
        btn_iptl.setFixedHeight(36)
        btn_iptl.setCursor(Qt.PointingHandCursor)
        btn_iptl.setStyleSheet(f'''
            QPushButton {{
                background:#F3F4F6; border:1px solid {BORDER};
                border-radius:10px; color:{TEXT2}; font-size:12px;
            }}
            QPushButton:hover {{ background:{BORDER}; }}
        ''')
        btn_iptl.clicked.connect(dlg.reject)

        btn_ok = QPushButton('💾  Kaydet')
        btn_ok.setFixedHeight(36)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setStyleSheet(f'''
            QPushButton {{
                background:{NAVY}; color:{GOLD};
                border:1.5px solid rgba(201,164,106,0.4);
                border-radius:10px; font-size:12px; font-weight:700;
            }}
            QPushButton:hover {{ background:{NAVY2}; border-color:{GOLD}; }}
        ''')

        def _kaydet():
            ad = inp.text().strip()
            if not ad:
                return
            self._profil_kaydet(ad)
            dlg.accept()

        btn_ok.clicked.connect(_kaydet)
        inp.returnPressed.connect(_kaydet)
        btn_row.addWidget(btn_iptl, 1)
        btn_row.addWidget(btn_ok, 2)
        lay.addLayout(btn_row)
        dlg.exec()

    def _profil_kaydet(self, ad: str):
        k = self.inp_kul.value()
        s = self.inp_sif.value()
        t = self.inp_tel.value()
        cfg = _cfg_load()
        profiller = cfg.get('ymm_profiller', [])
        for i, p in enumerate(profiller):
            if p.get('ad', '').lower() == ad.lower():
                profiller[i] = {'ad': ad, 'kullanici': k, 'sifre': s, 'ymm_tel': t}
                cfg['ymm_profiller'] = profiller
                _cfg_save(cfg)
                self._profil_listesi_yenile(sec_idx=i)
                self.log_panel.append(f'Profil güncellendi: {ad}', GOLD)
                return
        profiller.append({'ad': ad, 'kullanici': k, 'sifre': s, 'ymm_tel': t})
        cfg['ymm_profiller'] = profiller
        _cfg_save(cfg)
        self._profil_listesi_yenile(sec_idx=len(profiller) - 1)
        self.log_panel.append(f'Profil kaydedildi: {ad}', GREEN)

    def _profil_sil(self):
        idx = self.cmb_ymm_profil.currentIndex()
        if idx <= 0:
            return
        ad = self.cmb_ymm_profil.currentText().strip()
        ret = QMessageBox.question(
            self, 'Profil Sil',
            f'"{ad}" profilini silmek istediğinizden emin misiniz?',
            QMessageBox.Yes | QMessageBox.No,
        )
        if ret != QMessageBox.Yes:
            return
        cfg       = _cfg_load()
        profiller = cfg.get('ymm_profiller', [])
        if idx - 1 < len(profiller):
            profiller.pop(idx - 1)
        cfg['ymm_profiller'] = profiller
        cfg['son_profil']    = 0
        _cfg_save(cfg)
        self._profil_listesi_yenile()
        self.inp_kul.set_value('')
        self.inp_sif.set_value('')
        self.inp_tel.set_value('')
        self.prof_preview.setFixedHeight(0)
        self.log_panel.append(f'Profil silindi: {ad}', RED)

    # ── Kontrol Aksiyonları ───────────────────────────────────────────────────
    def _get_tasdik_ayarlari(self) -> dict:
        """GUI'deki tasdik ayarlarını dict olarak döner."""
        idx = self.cmb_tasdik_tur.currentIndex()
        tur_map = {0: 'KDV', 1: 'HER_IKISI'}
        kdv_tur_map = {0: 'NORMAL', 1: 'İNDİRİMLİ ORAN'}
        return {
            'tasdik_turu':      tur_map.get(idx, 'KDV'),
            'kdv_donem_turu':   kdv_tur_map.get(self.cmb_kdv_tur.currentIndex(), 'NORMAL'),
            'kdv_bas':          self.inp_kdv_bas.value(),
            'kdv_bit':          self.inp_kdv_bit.value(),
            'kdv_iade':         self.inp_kdv_iade.value(),
            'kdv_soz_tarih':    self.inp_kdv_soz_tarih.value(),
            'kdv_soz_no':       self.inp_kdv_soz_no.value(),
            'kdv_soz_giris':    self.inp_kdv_soz_giris.value(),
            'tam_bas':          self.inp_tam_bas.value(),
            'tam_bit':          self.inp_tam_bit.value(),
            'tam_soz_tarih':    self.inp_tam_soz_tarih.value(),
            'tam_soz_no':       self.inp_tam_soz_no.value(),
            'tam_soz_giris':    self.inp_tam_soz_giris.value(),
            # Panel ek verileri (word_klasor_oku için)
            'mukellef_tel':      self.inp_muk_tel.value(),
            'tutanak_baslangic': self.inp_tutanak_no.value(),
        }

    def _highlight_errors(self):
        """Tüm zorunlu InputRow alanlarını kontrol et, boş olanları turuncu yap."""
        ayar = self._get_tasdik_ayarlari()
        # field → zorunlu mu?
        checks = [
            (self.inp_kdv_soz_tarih, True),
            (self.inp_kdv_soz_no,    True),
            (self.inp_kdv_soz_giris, True),
            (self.inp_kdv_bas,  ayar['kdv_donem_turu'] == 'İNDİRİMLİ ORAN'),
            (self.inp_kdv_bit,  ayar['kdv_donem_turu'] == 'İNDİRİMLİ ORAN'),
            (self.inp_kdv_iade, ayar['kdv_donem_turu'] == 'İNDİRİMLİ ORAN'),
            (self.inp_tam_bas,       ayar['tasdik_turu'] == 'HER_IKISI'),
            (self.inp_tam_bit,       ayar['tasdik_turu'] == 'HER_IKISI'),
            (self.inp_tam_soz_tarih, ayar['tasdik_turu'] == 'HER_IKISI'),
            (self.inp_tam_soz_no,    ayar['tasdik_turu'] == 'HER_IKISI'),
            (self.inp_tam_soz_giris, ayar['tasdik_turu'] == 'HER_IKISI'),
        ]
        has_error = False
        for field, zorunlu in checks:
            empty = zorunlu and not field.value()
            field.set_error(empty)
            if empty:
                has_error = True
        return has_error

    def _validate(self) -> bool:
        # Zorunlu temel alanlar
        for field, name in [
            (self.inp_kul,   'Kullanıcı Kodu'),
            (self.inp_sif,   'Şifre'),
            (self.inp_tel,   'YMM Telefon Numarası'),
            (self.drop_word, 'Word Klasörü'),
        ]:
            if not field.value():
                QMessageBox.warning(self, 'Eksik Alan', f'{name} boş bırakılamaz.')
                return False

        # Tasdik alanlarını highlight et
        if self._highlight_errors():
            QMessageBox.warning(
                self, 'Eksik Alanlar',
                'Turuncu işaretli alanlar doldurulmadan otomasyon başlatılamaz.'
            )
            return False
        return True

    def _on_start(self):
        if not self._validate():
            return
        self._save_config()

        word_klasor = self.drop_word.value()
        kullanici   = self.inp_kul.value()
        sifre       = self.inp_sif.value()
        ymm_tel     = self.inp_tel.value()
        ayarlar     = self._get_tasdik_ayarlari()

        self.log_panel.append('Word dosyaları işleniyor...', TEXT2)
        try:
            from .karsit_import import word_klasor_oku
            result = word_klasor_oku(word_klasor, ayarlar)
        except Exception as e:
            QMessageBox.critical(self, 'Okuma Hatası', str(e))
            return

        if result.gecerli == 0:
            hatali_msg = ''
            if result.hatali:
                ornekler = [f'• {r.firma_adi}: {", ".join(r.hatalar)}'
                            for r in result.satirlar if r.hatalar][:5]
                hatali_msg = '\n\nHatalar:\n' + '\n'.join(ornekler)
            QMessageBox.warning(self, 'Veri Yok',
                                f'İşlenecek geçerli dosya bulunamadı.{hatali_msg}')
            return

        # Hatalı satırları loga yaz
        for r in result.satirlar:
            if r.hatalar:
                self.log_panel.append(
                    f'  ⚠ {r.firma_adi}: {", ".join(r.hatalar)}', ORANGE
                )

        # Sadece geçerli firmalar önizlemede ve işlemde
        firmalar = [r.firma_adi for r in result.satirlar if r.gecerli]
        # Tüm firmalar (bilgi_istem dahil) panelde gösterilir — bilgi_istem kırmızı
        tum_firmalar    = [r.firma_adi for r in result.satirlar]
        bilgi_istem_idx = {i for i, r in enumerate(result.satirlar) if not r.gecerli}

        # ── Önizleme Dialogu ──────────────────────────────────────────────────
        onizleme = OnizlemeDialog(firmalar, ayarlar, ayarlar.get('tutanak_baslangic', ''), self)
        if onizleme.exec() != QDialog.Accepted:
            self.log_panel.append('Otomasyon iptal edildi.', ORANGE)
            return

        try:
            ymm_profil = {'ymm_tel': ymm_tel}
            batch_id = kaydet_db(
                result, ymm_profil, kullanici,
                tutanak_baslangic=ayarlar.get('tutanak_baslangic', ''),
                kdv_soz_no_override=ayarlar.get('kdv_soz_no', ''),
            )
        except Exception as e:
            QMessageBox.critical(self, 'Veritabanı Hatası', str(e))
            return

        self._batch_id = batch_id
        self._rapor_data = [
            {'firma': f, 'tutanak': '', 'status': 'bekliyor', 'hata': ''}
            for f in firmalar
        ]
        self._last_aktif_idx = -1

        self.firma_panel.set_firmalar(tum_firmalar, bilgi_istem_idx)
        self.stat_bar.reset(len(firmalar))
        self.log_panel.append(
            f'{result.gecerli} firma yüklendi — otomasyon başlıyor...', GREEN
        )

        self._set_running(True)
        self._start_spinner()
        self._start_worker(batch_id, kullanici, sifre, ymm_tel,
                           mukellef_telefon=ayarlar.get('mukellef_tel', ''))

    def _start_worker(self, batch_id: str, kullanici: str, sifre: str,
                      ymm_telefon: str, mukellef_telefon: str = ""):
        self.worker = KarsitWorker(
            batch_id=batch_id,
            kullanici=kullanici,
            sifre=sifre,
            ymm_telefon=ymm_telefon,
            mukellef_telefon=mukellef_telefon,
            browser_visible=self._browser_visible,
            parent=self,
        )
        self.worker.sig_log.connect(self._on_worker_log)
        self.worker.sig_progress.connect(self._on_worker_progress)
        self.worker.sig_captcha.connect(self._on_worker_captcha)
        self.worker.sig_finished.connect(self._on_worker_finished)
        self.worker.sig_recovery.connect(self._on_worker_recovery)
        self.worker.sig_browser_state.connect(self._on_browser_state)
        self.worker.start()

    def _on_pause_resume(self):
        if not self.worker:
            return
        if not self._is_paused:
            self.worker.request_pause()
            self._is_paused = True
            self.btn_pause.setText('DEVAM ET')
            _pix_devam = _icon_pix('devam.png', 18)
            if not _pix_devam.isNull():
                self.btn_pause.setIcon(QIcon(_pix_devam))
                self.btn_pause.setIconSize(QSize(18, 18))
            self.btn_pause.setStyleSheet(f'''
                QPushButton {{
                    background:{GREEN}; color:#FFFFFF;
                    border:none; border-radius:12px;
                    font-size:12px; font-weight:700; padding:0 18px;
                }}
                QPushButton:hover {{ background:{GREEN2}; }}
                QPushButton:disabled {{ background:#E5E7EB; color:#9CA3AF; }}
            ''')
            self.log_panel.append('Duraklatıldı', ORANGE)
        else:
            self.worker.request_resume()
            self._is_paused = False
            self.btn_pause.setText('DURAKLAT')
            _pix_pause2 = _icon_pix('pause.png', 18)
            if not _pix_pause2.isNull():
                self.btn_pause.setIcon(QIcon(_pix_pause2))
                self.btn_pause.setIconSize(QSize(18, 18))
            self.btn_pause.setStyleSheet(f'''
                QPushButton {{
                    background:{ORANGE}; color:#FFFFFF;
                    border:none; border-radius:12px;
                    font-size:12px; font-weight:700; padding:0 18px;
                }}
                QPushButton:hover {{ background:#EA580C; }}
                QPushButton:disabled {{ background:#E5E7EB; color:#9CA3AF; }}
            ''')
            self.log_panel.append('Devam ediyor...', GREEN)

    def _on_stop(self):
        if self.worker:
            self.worker.request_stop()
            self.worker.terminate()
            self.worker = None
        self.log_panel.append('İptal edildi — tarayıcı kapatılıyor...', RED)
        self._set_running(False)

    # ── Worker Sinyalleri ─────────────────────────────────────────────────────
    def _on_worker_log(self, msg: str, color: str):
        self.log_panel.append(msg, color)

    def _on_worker_progress(self, current: int, total: int, firma: str):
        self.log_panel.set_progress(current, total)
        self.firma_panel.set_aktif(current - 1)
        self._last_aktif_idx = current - 1

        # Önceki firma tamamlandı → stat + durum ikonu güncelle
        prev = current - 2
        if prev >= 0:
            self.stat_bar.inc_tamamlanan()
            self.firma_panel.set_firma_done(prev)
            if prev < len(self._rapor_data):
                self._rapor_data[prev]['status'] = 'tamamlandi'

    def _on_worker_captcha(self, img_bytes: bytes, token: str):
        self.log_panel.append('CAPTCHA — kod girilmesi bekleniyor', GOLD)
        self._captcha_show(img_bytes, token)
        self._play_sound('captcha')
        self._show_toast('CAPTCHA bekleniyor — kodu girin', 'warning')

    def _on_worker_finished(self, success: bool, ozet: str):
        self._captcha_hide()
        self._stop_spinner()
        self._set_running(False)

        if success:
            self.log_panel.append(f'✓  {ozet}', GREEN)
            # Son firma da tamamlandı
            if self._last_aktif_idx >= 0:
                self.stat_bar.inc_tamamlanan()
                self.firma_panel.set_firma_done(self._last_aktif_idx)
                if self._last_aktif_idx < len(self._rapor_data):
                    self._rapor_data[self._last_aktif_idx]['status'] = 'tamamlandi'
            self._play_sound('success')
            self._show_toast(
                f'Tamamlandı!  {self.stat_bar.get_tamamlanan()} firma işlendi.', 'success'
            )
        else:
            self.log_panel.append(f'✕  {ozet}', RED)
            # Son aktif firma hata aldı
            if self._last_aktif_idx >= 0:
                self.stat_bar.inc_hata()
                self.firma_panel.set_firma_error(self._last_aktif_idx)
                if self._last_aktif_idx < len(self._rapor_data):
                    self._rapor_data[self._last_aktif_idx]['status'] = 'hata'
                    self._rapor_data[self._last_aktif_idx]['hata']   = ozet[:80]
            self._play_sound('error')
            self._show_toast(f'Hata: {ozet[:55]}', 'error')

        self.worker = None

        # Rapor dialogunu 500ms sonra aç (UI settle etsin)
        if self._rapor_data:
            QTimer.singleShot(500, self._show_rapor_dialog)

    def _on_worker_recovery(self, report: list):
        if report:
            self.log_panel.append(
                f'Önceki oturumdan {len(report)} job sıfırlandı.', '#F6E05E'
            )

    # ── Durum Yönetimi ────────────────────────────────────────────────────────
    def _set_running(self, running: bool):
        self._is_running = running
        self._is_paused  = False
        self.btn_start.setEnabled(not running)
        self.btn_pause.setEnabled(running)
        self.btn_pause.setText('DURAKLAT')
        self.btn_stop.setEnabled(running)
        self.stat_bar.set_browser_state(running, self._browser_visible)
        if hasattr(self, 'status_chip'):
            self.status_chip.set_state(running, self._browser_visible)

    # ── Spinner (Başlat butonu animasyonu) ───────────────────────────────────
    _SPINNER_FRAMES = ['◐', '◓', '◑', '◒']

    def _start_spinner(self):
        self._spinner_idx = 0
        self._spinner_timer = QTimer(self)
        self._spinner_timer.setInterval(120)
        self._spinner_timer.timeout.connect(self._spinner_step)
        self._spinner_timer.start()

    def _spinner_step(self):
        f = self._SPINNER_FRAMES[self._spinner_idx % len(self._SPINNER_FRAMES)]
        self.btn_start.setText(f'  {f}  İşleniyor...')
        self.btn_start.setIcon(QIcon())
        self._spinner_idx += 1

    def _stop_spinner(self):
        if self._spinner_timer:
            self._spinner_timer.stop()
            self._spinner_timer = None
        self.btn_start.setText('BAŞLAT')
        pix = _icon_pix('power.png', 18)
        if not pix.isNull():
            self.btn_start.setIcon(QIcon(pix))
            self.btn_start.setIconSize(QSize(18, 18))

    # ── Toast ─────────────────────────────────────────────────────────────────
    def _show_toast(self, msg: str, level: str = 'info'):
        self._toast.show_toast(msg, level)

    # ── Ses Bildirimi ─────────────────────────────────────────────────────────
    @staticmethod
    def _play_sound(kind: str = 'info'):
        """
        kind: 'success' | 'error' | 'captcha'
        Windows'ta arka planda sistem sesi çalar; diğer platformlarda sessiz.
        """
        def _beep():
            try:
                import winsound
                if kind == 'success':
                    winsound.MessageBeep(winsound.MB_ICONINFORMATION)
                elif kind == 'error':
                    winsound.MessageBeep(winsound.MB_ICONHAND)
                elif kind == 'captcha':
                    winsound.MessageBeep(winsound.MB_ICONEXCLAMATION)
                else:
                    winsound.MessageBeep(winsound.MB_OK)
            except Exception:
                pass  # Windows değil veya ses devre dışı
        import threading
        threading.Thread(target=_beep, daemon=True).start()

    # ── Rapor Dialog ──────────────────────────────────────────────────────────
    def _show_rapor_dialog(self):
        if not self._rapor_data:
            return
        dlg = RaporDialog(self._rapor_data, self)
        dlg.exec()

    # ── Tarayıcı toggle ───────────────────────────────────────────────────────
    def _on_browser_toggle(self):
        """Chip veya stat kart tıklanınca."""
        if self.worker and self.worker.isRunning():
            self.worker.request_toggle_browser()
        else:
            self._browser_visible = not self._browser_visible
            self._sync_browser_ui()

    def _on_browser_state(self, visible: bool):
        """Worker'dan gelen durum güncellemesi."""
        self._browser_visible = visible
        self._sync_browser_ui()

    def _sync_browser_ui(self):
        self.stat_bar.set_browser_state(self._is_running, self._browser_visible)
        if hasattr(self, 'status_chip'):
            self.status_chip.set_state(self._is_running, self._browser_visible)

    # ── Dark / Light Mod ─────────────────────────────────────────────────────
    def _toggle_theme(self):
        from core import theme as _theme
        _theme.toggle()

    def _apply_theme(self):
        from core import theme as _theme
        dark = _theme.is_dark()

        if hasattr(self, 'btn_theme'):
            self.btn_theme.setText('☀️' if dark else '🌙')

        self._update_logo()

        root = getattr(self, '_root_widget', None) or self.centralWidget()
        _theme.apply_to_widget(root, exclude_types=(BrowserStateKart, StatusChip))

        # InputRow widget'larını tema-farkındalıklı metodlarla yenile
        if root:
            for inp in root.findChildren(InputRow):
                inp._apply_normal_style()

        if hasattr(self, 'stat_bar') and hasattr(self.stat_bar, '_k_browser'):
            self.stat_bar._k_browser._refresh()
        if hasattr(self, 'status_chip'):
            self.status_chip._refresh()

    def _update_logo(self):
        """Temaya göre doğru logoyu yükler."""
        if not hasattr(self, '_lbl_logo'):
            return
        from core import theme as _theme
        _LOGO_DIR = os.path.normpath(
            os.path.join(os.path.dirname(__file__), '..', '..', 'Logom', 'big_logo')
        )
        fname = 'ContraCoreBeyaz.png' if _theme.is_dark() else 'ContraCore.png'
        pix = QPixmap(os.path.join(_LOGO_DIR, fname))
        if pix.isNull():
            pix = QPixmap(os.path.join(_LOGO_DIR, 'ContraCore.png'))
        if not pix.isNull():
            h = self._logo_h
            w = int(pix.width() * h / pix.height())
            self._lbl_logo.setPixmap(pix.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            self._lbl_logo.setText('CC')

        # ContraCORE yazı pixmap'ini de güncelle
        if hasattr(self, '_cc_lbl'):
            self._cc_lbl.setPixmap(_cc_pix(24, dark=_theme.is_dark()))

    # ── resizeEvent — toast yeniden konumlanır ────────────────────────────────
    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, '_toast'):
            self._toast._reposition()

    # ── ContraCore Lifecycle Hooks ────────────────────────────────────────────
    def on_module_activated(self):
        pass

    def on_module_deactivated(self):
        self._captcha_hide()
        if self.worker and self.worker.isRunning():
            self.worker.request_stop()
            self.worker.wait(3000)


# ── Standalone çalıştırma ─────────────────────────────────────────────────────
if __name__ == '__main__':
    app = QApplication(sys.argv)
    app.setStyle('Fusion')
    win = MainWindow(expire_date=None, trial_status=None)
    win.show()
    sys.exit(app.exec())

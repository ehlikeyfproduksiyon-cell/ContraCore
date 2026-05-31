#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
191 Muavin Karşılaştırma v1.0
Developed by Serkan ŞAHİN © 2026
"""
import sys, os, threading, time, json
from datetime import datetime
from decimal import Decimal

from PySide6.QtWidgets import *
from PySide6.QtCore    import *
from PySide6.QtGui     import *

from core import _icons as _ic

def ip(name):
    return _ic.load(name)

VERSION = '1.0'

_APPDATA_DIR = os.path.join(os.environ.get('APPDATA', os.path.expanduser('~')), 'ContraCore', 'compare-191')
os.makedirs(_APPDATA_DIR, exist_ok=True)
RECENT_FILE  = os.path.join(_APPDATA_DIR, 'recent_files.json')

def load_recent():
    try:
        with open(RECENT_FILE, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}

def save_recent(data):
    try:
        with open(RECENT_FILE, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False)
    except Exception:
        pass

def add_recent(tag, path):
    d = load_recent()
    lst = d.get(tag, [])
    if path in lst:
        lst.remove(path)
    lst.insert(0, path)
    d[tag] = lst[:5]
    save_recent(d)

# ── Renk Paleti ──────────────────────────────────────────────────────────────
BG      = '#F2F4F7'
CARD    = '#FFFFFF'
NAVY    = '#0B1F3A'
NAVY2   = '#162D4E'
GREEN   = '#22C55E'
GREEN2  = '#16A34A'
RED     = '#EF4444'
RED2    = '#DC2626'
GOLD    = '#C9A46A'
BORDER  = '#E5E7EB'
TEXT    = '#111827'
TEXT2   = '#6B7280'
TEXT3   = '#9CA3AF'
BLUE_BG = '#EFF6FF'
ORG_BG  = '#FFF7ED'
GRN_BG  = '#F0FDF4'
RED_BG  = '#FFF1F2'
YEL_BG  = '#FEFCE8'
LOG_BG  = '#0F172A'
BLUE    = '#3B82F6'
BLUE2   = '#2563EB'
ORANGE  = '#F97316'


# ── Worker Sinyalleri ────────────────────────────────────────────────────────
class WorkerSig(QObject):
    log      = Signal(str, str)
    progress = Signal(float)
    stats    = Signal(int, int, int, int)
    done     = Signal(int, int, int, int)


class KarsilastirWorker(QThread):
    def __init__(self, ikdvl_path, muavin_path, stop_flag, max_muavin_rows=None):
        super().__init__()
        self.sig             = WorkerSig()
        self.ikdvl_path      = ikdvl_path
        self.muavin_path     = muavin_path
        self.stop_flag       = stop_flag
        self.max_muavin_rows = max_muavin_rows
        self._result         = None

    def get_result(self):
        return self._result

    def run(self):
        try:
            from cc_modules.compare_191 import karsilastir
            result = karsilastir.compare(
                ikdvl_path=self.ikdvl_path,
                muavin_path=self.muavin_path,
                stop_flag=self.stop_flag,
                cb_log=self.sig.log.emit,
                cb_progress=self.sig.progress.emit,
                cb_stats=self.sig.stats.emit,
                cb_done=self.sig.done.emit,
                max_muavin_rows=self.max_muavin_rows,
            )
            self._result = result
        except Exception as e:
            import traceback
            self.sig.log.emit(f'❌ Hata: {e}', 'err')
            self.sig.log.emit(traceback.format_exc(), 'err')
            self.sig.done.emit(0, 0, 0, 0)


class DuzeltWorkerSig(QObject):
    log      = Signal(str, str)
    progress = Signal(float)
    done     = Signal(str)   # output_path veya '' (hata)


class DuzeltWorker(QThread):
    def __init__(self, ikdvl_path, result, output_path, mode='mevcut'):
        super().__init__()
        self.sig         = DuzeltWorkerSig()
        self.ikdvl_path  = ikdvl_path
        self.result      = result
        self.output_path = output_path
        self.mode        = mode

    def run(self):
        try:
            from cc_modules.compare_191 import karsilastir
            karsilastir.duzelt(
                ikdvl_path=self.ikdvl_path,
                result=self.result,
                output_path=self.output_path,
                cb_log=self.sig.log.emit,
                cb_progress=self.sig.progress.emit,
                mode=self.mode,
            )
            self.sig.done.emit(self.output_path)
        except Exception as e:
            import traceback
            self.sig.log.emit(f'❌ Düzeltme hatası: {e}', 'err')
            self.sig.log.emit(traceback.format_exc(), 'err')
            self.sig.done.emit('')


# ── Son Kullanılan Dosyalar Popup ─────────────────────────────────────────────
class RecentPopup(QFrame):
    picked = Signal(str)

    def __init__(self, tag, parent=None):
        super().__init__(parent, Qt.Popup | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.tag  = tag
        self.setFixedWidth(320)
        self._lay = QVBoxLayout(self)
        self._lay.setContentsMargins(0, 0, 0, 0)

    def show_at(self, widget):
        self._refresh()
        self.adjustSize()
        p = widget.mapToGlobal(QPoint(0, widget.height() + 4))
        screen = QApplication.screenAt(p) or QApplication.primaryScreen()
        sg = screen.availableGeometry()
        x = max(sg.left(), min(p.x(), sg.right()  - self.width()))
        y = max(sg.top(),  min(p.y(), sg.bottom() - self.height()))
        self.move(x, y)
        self.show()

    def _refresh(self):
        while self._lay.count():
            item = self._lay.takeAt(0)
            if item.widget():
                item.widget().deleteLater()

        card = QFrame()
        card.setStyleSheet(
            f'QFrame{{background:{CARD};border-radius:12px;border:1px solid {BORDER};}}'
        )
        cl = QVBoxLayout(card)
        cl.setContentsMargins(12, 12, 12, 12)
        cl.setSpacing(4)

        ttl = QLabel('Son Kullanılan Dosyalar')
        ttl.setStyleSheet(
            f'font-size:11px;font-weight:700;color:{TEXT};background:transparent;border:none;'
        )
        cl.addWidget(ttl)

        paths = load_recent().get(self.tag, [])
        paths = [p for p in paths if os.path.exists(p)]
        if not paths:
            l = QLabel('Henüz dosya seçilmedi.')
            l.setStyleSheet(
                f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
            )
            cl.addWidget(l)
        else:
            for p in paths:
                b = QPushButton(f'  {os.path.basename(p)}')
                b.setToolTip(p)
                b.setCursor(Qt.PointingHandCursor)
                b.setStyleSheet(f'''
                    QPushButton {{
                        background:#F9FAFB;border:1px solid {BORDER};
                        border-radius:8px;color:{TEXT2};font-size:11px;
                        padding:6px 10px;text-align:left;
                    }}
                    QPushButton:hover {{
                        background:{BLUE_BG};border-color:{NAVY};color:{TEXT};
                    }}
                ''')
                b.clicked.connect(lambda _, x=p: self._pick(x))
                cl.addWidget(b)

        self._lay.addWidget(card)

    def _pick(self, path):
        self.picked.emit(path)
        self.close()


# ── Profil Popup ──────────────────────────────────────────────────────────────
class ProfilePopup(QFrame):
    def __init__(self, parent=None):
        super().__init__(parent, Qt.Popup | Qt.FramelessWindowHint)
        self.setAttribute(Qt.WA_TranslucentBackground)
        self.setFixedWidth(290)
        self._just_closed = False

        outer = QVBoxLayout(self)
        outer.setContentsMargins(6, 6, 6, 6)

        card = QFrame()
        card.setStyleSheet('''
            QFrame {
                background: qlineargradient(x1:0,y1:0,x2:0,y2:1,
                    stop:0 #162D4E, stop:1 #0B1F3A);
                border-radius: 16px;
                border: 1px solid #C9A46A;
            }
        ''')
        glow = QGraphicsDropShadowEffect()
        glow.setBlurRadius(28)
        glow.setColor(QColor('#C9A46A44'))
        glow.setOffset(0, 4)
        card.setGraphicsEffect(glow)

        cl = QVBoxLayout(card)
        cl.setContentsMargins(0, 0, 0, 16)
        cl.setSpacing(0)

        # ── Gradient başlık şeridi ──────────────────────────────────────────
        header = QFrame()
        header.setFixedHeight(64)
        header.setStyleSheet('''
            QFrame {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1E3660, stop:1 #0B1F3A);
                border-radius: 16px;
                border-bottom-left-radius: 0px;
                border-bottom-right-radius: 0px;
                border: none;
            }
        ''')
        hl = QHBoxLayout(header)
        hl.setContentsMargins(16, 10, 16, 10)
        hl.setSpacing(12)

        av_lbl = QLabel()
        av_lbl.setFixedSize(42, 42)
        av_lbl.setStyleSheet('background:transparent;border:none;')
        av_pix = QPixmap(ip('profile.png'))
        if not av_pix.isNull():
            canvas = QPixmap(42, 42)
            canvas.fill(Qt.transparent)
            pc = QPainter(canvas)
            pc.setRenderHint(QPainter.Antialiasing)
            sc = av_pix.scaled(38, 38, Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            path = QPainterPath()
            path.addEllipse(2.0, 2.0, 38.0, 38.0)
            pc.setClipPath(path)
            pc.drawPixmap(2, 2, sc)
            pc.setClipping(False)
            pc.setPen(QPen(QColor(GOLD), 2))
            pc.setBrush(Qt.NoBrush)
            pc.drawEllipse(QRectF(1, 1, 40, 40))
            pc.end()
            av_lbl.setPixmap(canvas)
        else:
            av_lbl.setText('👤')
        hl.addWidget(av_lbl)

        name_col = QVBoxLayout()
        name_col.setSpacing(1)
        name_lbl = QLabel('Serkan ŞAHİN')
        name_lbl.setStyleSheet(
            'font-size:13px;font-weight:700;color:#FFFFFF;background:transparent;border:none;'
        )
        role_lbl = QLabel('Yazılım Geliştirici')
        role_lbl.setStyleSheet(
            f'font-size:10px;color:{GOLD};background:transparent;border:none;'
        )
        name_col.addWidget(name_lbl)
        name_col.addWidget(role_lbl)
        hl.addLayout(name_col)
        hl.addStretch()
        cl.addWidget(header)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setFixedHeight(1)
        sep.setStyleSheet('background:#C9A46A44;border:none;max-height:1px;')
        cl.addWidget(sep)
        cl.addSpacing(12)

        for ico, label, val in [
            ('📱', 'WhatsApp', '0531 087 93 39'),
            ('✉️', 'Mail',     'serkan.opsiyonymm@gmail.com'),
        ]:
            row_frame = QFrame()
            row_frame.setStyleSheet('''
                QFrame {
                    background: rgba(255,255,255,0.04);
                    border-radius: 8px;
                    border: 1px solid rgba(201,164,106,0.15);
                }
                QFrame:hover {
                    background: rgba(201,164,106,0.10);
                    border: 1px solid rgba(201,164,106,0.35);
                }
            ''')
            row_lay = QHBoxLayout(row_frame)
            row_lay.setContentsMargins(12, 8, 12, 8)
            row_lay.setSpacing(10)

            ico_lbl = QLabel(ico)
            ico_lbl.setFixedWidth(20)
            ico_lbl.setStyleSheet('background:transparent;border:none;font-size:14px;')

            info_col = QVBoxLayout()
            info_col.setSpacing(0)
            lbl_top = QLabel(label)
            lbl_top.setStyleSheet(
                f'font-size:9px;font-weight:600;color:{GOLD};background:transparent;border:none;'
            )
            lbl_val = QLabel(val)
            lbl_val.setStyleSheet(
                'font-size:11px;color:#E2E8F0;background:transparent;border:none;'
            )
            lbl_val.setTextInteractionFlags(Qt.TextSelectableByMouse)
            info_col.addWidget(lbl_top)
            info_col.addWidget(lbl_val)

            row_lay.addWidget(ico_lbl)
            row_lay.addLayout(info_col)
            row_lay.addStretch()

            cl.addWidget(row_frame)
            cl.addSpacing(6)

        outer.addWidget(card)

    def hideEvent(self, event):
        super().hideEvent(event)
        self._just_closed = True
        QTimer.singleShot(200, lambda: setattr(self, '_just_closed', False))

    def show_at(self, widget):
        p = widget.mapToGlobal(QPoint(0, widget.height() + 4))
        x = p.x() - self.width() + widget.width()
        screen = QApplication.screenAt(p) or QApplication.primaryScreen()
        sg = screen.availableGeometry()
        x = max(sg.left(), min(x,    sg.right()  - self.width()))
        y = max(sg.top(),  min(p.y(), sg.bottom() - self.height()))
        self.move(x, y)
        self.show()


# ── Dosya Seçim Kartı ─────────────────────────────────────────────────────────
class FileCard(QFrame):
    changed = Signal(str)

    def __init__(self, label, icon_file, ico_bg, filter_str, recent_tag=''):
        super().__init__()
        self._path       = ''
        self._filter     = filter_str
        self._label_text = label
        self._ico_bg     = ico_bg
        self._icon_file  = icon_file
        self._recent_tag = recent_tag or label.lower().replace(' ', '_')
        self._recent_popup = RecentPopup(self._recent_tag)
        self._recent_popup.picked.connect(self.set_path)
        self.setAcceptDrops(True)
        self._set_normal_style()
        self.setFixedHeight(90)
        self._build()

    def _set_normal_style(self):
        self.setStyleSheet(f'''
            QFrame {{
                background: {CARD}; border-radius: 14px;
                border: 1.5px solid {BORDER};
            }}
        ''')

    def _set_drag_style(self):
        self.setStyleSheet(f'''
            QFrame {{
                background: {BLUE_BG}; border-radius: 14px;
                border: 2px dashed {NAVY};
            }}
        ''')

    def _build(self):
        lay = QHBoxLayout(self)
        lay.setContentsMargins(14, 0, 14, 0)
        lay.setSpacing(12)

        ico_box = QLabel()
        ico_box.setFixedSize(52, 52)
        ico_box.setAlignment(Qt.AlignCenter)
        ico_box.setStyleSheet(
            f'background:{self._ico_bg};border-radius:12px;border:none;'
        )
        pix = QPixmap(ip(self._icon_file))
        if not pix.isNull():
            ico_box.setPixmap(pix.scaled(34, 34, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            ico_box.setText('📄')
            ico_box.setStyleSheet(
                f'background:{self._ico_bg};border-radius:12px;border:none;font-size:24px;'
            )
        lay.addWidget(ico_box)

        ml = QVBoxLayout()
        ml.setSpacing(2)
        self.lbl_name = QLabel(self._label_text)
        self.lbl_name.setStyleSheet(
            f'font-size:11px;font-weight:700;color:{TEXT};'
            f'background:transparent;border:none;letter-spacing:0.3px;'
        )
        self.lbl_path = QLabel('Dosya seçilmedi...')
        self.lbl_path.setStyleSheet(
            f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
        )
        self.lbl_info = QLabel('')
        self.lbl_info.setStyleSheet(
            f'font-size:11px;color:{GREEN};font-weight:600;'
            f'background:transparent;border:none;'
        )
        ml.addWidget(self.lbl_name)
        ml.addWidget(self.lbl_path)
        ml.addWidget(self.lbl_info)
        lay.addLayout(ml, 1)

        bl = QHBoxLayout()
        bl.setSpacing(6)

        btn_pick = QPushButton('📂  Seç')
        btn_pick.setFixedSize(80, 34)
        btn_pick.setCursor(Qt.PointingHandCursor)
        btn_pick.setStyleSheet(f'''
            QPushButton {{
                background:{NAVY};color:#FFF;border-radius:8px;
                border:none;font-size:11px;font-weight:700;
            }}
            QPushButton:hover{{background:{NAVY2};}}
        ''')
        btn_pick.clicked.connect(self._pick)

        btn_hist = QPushButton()
        btn_hist.setFixedSize(34, 34)
        btn_hist.setCursor(Qt.PointingHandCursor)
        pix_clock = QPixmap(ip('clock.png'))
        if not pix_clock.isNull():
            btn_hist.setIcon(QIcon(pix_clock.scaled(18, 18, Qt.KeepAspectRatio, Qt.SmoothTransformation)))
            btn_hist.setIconSize(QSize(18, 18))
        btn_hist.setStyleSheet(f'''
            QPushButton {{
                background:#F3F4F6;border-radius:8px;
                border:1px solid {BORDER};
            }}
            QPushButton:hover{{background:{BLUE_BG};border-color:{BLUE};}}
        ''')
        btn_hist.clicked.connect(lambda: self._recent_popup.show_at(btn_hist))
        self._hist_btn = btn_hist

        btn_clear = QPushButton()
        btn_clear.setFixedSize(34, 34)
        btn_clear.setCursor(Qt.PointingHandCursor)
        pix_x = QPixmap(ip('x.png'))
        if not pix_x.isNull():
            btn_clear.setIcon(QIcon(pix_x.scaled(14, 14, Qt.KeepAspectRatio, Qt.SmoothTransformation)))
            btn_clear.setIconSize(QSize(14, 14))
        else:
            btn_clear.setText('✕')
        btn_clear.setStyleSheet(f'''
            QPushButton {{
                background:#F3F4F6;border-radius:8px;
                border:1px solid {BORDER};
            }}
            QPushButton:hover{{background:{RED_BG};border-color:{RED};}}
        ''')
        btn_clear.clicked.connect(self.clear)

        bl.addWidget(btn_pick)
        bl.addWidget(btn_hist)
        bl.addWidget(btn_clear)
        lay.addLayout(bl)

    def _pick(self):
        recent = load_recent().get(self._recent_tag, [])
        recent = [r for r in recent if os.path.exists(r)]
        start_dir = os.path.dirname(recent[0]) if recent else ''
        path, _ = QFileDialog.getOpenFileName(
            self, f'{self._label_text} Seç', start_dir, self._filter
        )
        if path:
            self.set_path(path)

    def set_path(self, path):
        self._path = path
        name = os.path.basename(path)
        self.lbl_path.setText(name)
        self.lbl_path.setStyleSheet(
            f'font-size:11px;color:{TEXT};background:transparent;border:none;'
        )
        self._try_read_info(path)
        add_recent(self._recent_tag, path)
        self.changed.emit(path)

    def _try_read_info(self, path):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = ws.max_row or 0
            wb.close()
            self.lbl_info.setText(f'✓  Hazır  ·  ~{max(0, rows - 6)} kayıt')
        except Exception:
            self.lbl_info.setText('✓  Dosya seçildi')

    def clear(self):
        self._path = ''
        self.lbl_path.setText('Dosya seçilmedi...')
        self.lbl_path.setStyleSheet(
            f'font-size:11px;color:{TEXT3};background:transparent;border:none;'
        )
        self.lbl_info.setText('')
        self.changed.emit('')

    def get_path(self):
        return self._path

    def dragEnterEvent(self, e):
        if e.mimeData().hasUrls():
            self._set_drag_style()
            e.acceptProposedAction()

    def dragLeaveEvent(self, e):
        self._set_normal_style()

    def dropEvent(self, e):
        self._set_normal_style()
        for url in e.mimeData().urls():
            p = url.toLocalFile()
            if p.lower().endswith(('.xlsx', '.xls')):
                self.set_path(p)
                break


# ── Halka Grafik ─────────────────────────────────────────────────────────────
class DonutChart(QWidget):
    _ANIM_FPS  = 30     # frame/s
    _ANIM_MS   = 600    # animasyon süresi ms

    def __init__(self, parent=None):
        super().__init__(parent)
        self._pct        = 0.0    # gösterilen değer
        self._target     = 0.0    # hedef değer
        self._start_pct  = 0.0    # animasyon başlangıç değeri
        self._anim_t0    = 0.0    # animasyon başlangıç zamanı
        self._color      = QColor('#CBD5E1')
        self._timer      = QTimer(self)
        self._timer.setInterval(1000 // self._ANIM_FPS)
        self._timer.timeout.connect(self._anim_tick)

    def _target_color(self, pct: float) -> QColor:
        if pct >= 95:
            return QColor('#22C55E')
        if pct >= 80:
            return QColor('#F59E0B')
        return QColor('#EF4444')

    def set_pct(self, pct: float):
        self._start_pct = self._pct
        self._target    = max(0.0, min(100.0, pct))
        self._anim_t0   = time.monotonic()
        if abs(self._target - self._pct) > 0.001:
            self._timer.start()
        else:
            self._pct   = self._target
            self._color = self._target_color(self._pct)
            self.update()

    def _anim_tick(self):
        elapsed = (time.monotonic() - self._anim_t0) * 1000
        t       = min(1.0, elapsed / self._ANIM_MS)
        t_eased = 1.0 - (1.0 - t) ** 3   # ease-out cubic
        self._pct   = self._start_pct + (self._target - self._start_pct) * t_eased
        self._color = self._target_color(self._pct)
        self.update()
        if t >= 1.0:
            self._pct = self._target
            self._timer.stop()

    def reset(self):
        self._timer.stop()
        self._pct    = 0.0
        self._target = 0.0
        self._color  = QColor('#CBD5E1')
        self.update()

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)

        side = min(self.width(), self.height())
        margin = 10
        r = QRectF((self.width() - side) / 2 + margin,
                   (self.height() - side) / 2 + margin,
                   side - 2 * margin, side - 2 * margin)

        # Arka plan halkası
        pen_bg = QPen(QColor('#E5E7EB'), side * 0.13)
        pen_bg.setCapStyle(Qt.FlatCap)
        p.setPen(pen_bg)
        p.drawArc(r, 0, 360 * 16)

        # Dolu yay
        if self._pct > 0:
            pen_fg = QPen(self._color, side * 0.13)
            pen_fg.setCapStyle(Qt.FlatCap)
            p.setPen(pen_fg)
            span = int(self._pct / 100 * 360 * 16)
            p.drawArc(r, 90 * 16, -span)

        # Orta metin — font boyutunu iç alana sığacak şekilde dinamik ayarla
        from core import theme as _t
        p.setPen(QColor(_t.DARK_TEXT if _t.is_dark() else TEXT))
        txt = f'%{self._pct:.1f}'   # her zaman sayısal göster (%0.0 dahil)
        inner_w = (side - 2 * margin) * (1 - 0.13) * 0.80   # ring kalınlığı çıkarılmış iç çap
        base_pt = int(side * 0.17)
        font = QFont('Segoe UI', base_pt, QFont.Bold)
        from PySide6.QtGui import QFontMetrics as _QFM
        fm = _QFM(font)
        while fm.horizontalAdvance(txt) > inner_w and base_pt > 7:
            base_pt -= 1
            font.setPointSize(base_pt)
            fm = _QFM(font)
        p.setFont(font)
        p.drawText(QRectF(self.rect()), Qt.AlignCenter, txt)

        p.end()


# ── Özet Stat Kartı ───────────────────────────────────────────────────────────
class SummaryCard(QFrame):
    def __init__(self, icon_file, title, bg_color, title_color, num_color):
        super().__init__()
        self._num_color = num_color
        self.setStyleSheet(
            f'QFrame{{background:{bg_color};border-radius:14px;border:1px solid {BORDER};}}'
        )
        lay = QVBoxLayout(self)
        lay.setContentsMargins(14, 12, 14, 12)
        lay.setSpacing(6)

        # İkon + başlık satırı
        top = QHBoxLayout()
        top.setSpacing(8)

        ico = QLabel()
        ico.setFixedSize(32, 32)
        ico.setAlignment(Qt.AlignCenter)
        ico.setStyleSheet('background:transparent;border:none;')
        pix_ic = QPixmap(ip(icon_file))
        if not pix_ic.isNull():
            ico.setPixmap(pix_ic.scaled(28, 28, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        top.addWidget(ico)

        t = QLabel(title)
        t.setStyleSheet(
            f'font-size:10px;font-weight:700;color:{title_color};'
            f'background:transparent;border:none;letter-spacing:0.5px;'
        )
        t.setWordWrap(True)
        top.addWidget(t, 1)
        lay.addLayout(top)

        # Sayı
        self.lbl_count = QLabel('—')
        self.lbl_count.setStyleSheet(
            f'font-size:22px;font-weight:800;color:{num_color};'
            f'background:transparent;border:none;'
        )
        lay.addWidget(self.lbl_count)

        # Tutar
        self.lbl_tutar = QLabel('—')
        self.lbl_tutar.setStyleSheet(
            f'font-size:11px;font-weight:600;color:#8491ac;'
            f'background:transparent;border:none;'
        )
        lay.addWidget(self.lbl_tutar)

    def set_values(self, count, tutar=None):
        self.lbl_count.setText(str(count))
        if tutar is not None:
            self.lbl_tutar.setText(f'{tutar:,.2f} TL'.replace(',', '.'))
        else:
            self.lbl_tutar.setText('—')

    def reset(self):
        self.lbl_count.setText('—')
        self.lbl_tutar.setText('—')


# ── Ana Pencere ────────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    def __init__(self, expire_date=None, trial_status=None):
        super().__init__()
        self.expire_date    = expire_date
        self._trial_status  = trial_status
        self._running       = False
        self._stop_flag     = threading.Event()
        self.worker         = None
        self._t0            = None
        self._last_result   = None
        self._islem_baslama = None
        self._islem_sure    = None
        self._profile_popup = ProfilePopup(self)
        self._spin_frames   = ['◐', '◓', '◑', '◒']
        self._spin_idx      = 0
        self._spin_timer    = QTimer(self)
        self._spin_timer.setInterval(120)
        self._spin_timer.timeout.connect(self._spin_tick)
        # ContraCore lifecycle state
        self._cc_paused          = False
        self._cc_log_buffer      = []
        self._cc_spin_was_active = False
        self._ikdvl_months  = set()
        self._muavin_months = set()
        self._detected_period = ''

        exp = expire_date.strftime('%d.%m.%Y') if expire_date else ''
        self.setWindowTitle(
            f'191 Muavin Karşılaştırma — Lisans: {exp}' if exp
            else '191 Muavin Karşılaştırma'
        )

        screen = QApplication.primaryScreen().availableGeometry()
        w = 920
        h = min(int(screen.height() * 0.95), 960)
        x = screen.x() + (screen.width()  - w) // 2
        y = screen.y() + (screen.height() - h) // 2
        self.setMinimumSize(740, 620)
        self.setGeometry(x, y, w, h)

        self._tray = QSystemTrayIcon(self)
        for ico_f in ('muavin.ico', 'logo.png'):
            _pix = ip(ico_f)
            if not _pix.isNull():
                self._tray.setIcon(QIcon(_pix))
                break
        self._tray.show()

        self._build()

    # ── Kart çerçevesi ────────────────────────────────────────────────────────
    def _card(self):
        f = QFrame()
        f.setStyleSheet(
            f'QFrame{{background:{CARD};border-radius:16px;border:1px solid {BORDER};}}'
        )
        return f

    # ── Ana Layout ────────────────────────────────────────────────────────────
    def _build(self):
        content = QWidget()
        content.setStyleSheet(f'background:{BG};')

        root = QVBoxLayout(content)
        root.setContentsMargins(14, 14, 14, 10)
        root.setSpacing(10)

        root.addWidget(self._header())
        root.addWidget(self._trial_banner())
        root.addWidget(self._file_cards())
        root.addWidget(self._stats_row())
        root.addWidget(self._buttons())
        root.addLayout(self._bottom_panels(), 1)

        screen_h = QApplication.primaryScreen().availableGeometry().height()
        if screen_h < 768:
            scroll = QScrollArea()
            scroll.setWidget(content)
            scroll.setWidgetResizable(True)
            scroll.setHorizontalScrollBarPolicy(Qt.ScrollBarAlwaysOff)
            scroll.setStyleSheet(f'QScrollArea{{border:none;background:{BG};}}')
            self._root_widget = scroll
            self.setCentralWidget(scroll)
        else:
            self._root_widget = content
            self.setCentralWidget(content)

        from core import theme as _theme
        _theme.register(self._apply_theme)

        # Klavye kısayolları
        QShortcut(QKeySequence('F5'),      self, self._on_start)
        QShortcut(QKeySequence('F9'),      self, self._on_report)
        QShortcut(QKeySequence('Ctrl+D'),  self, self._on_duzelt)

    # ── Header ────────────────────────────────────────────────────────────────
    def _header(self):
        import os as _os
        f = self._card()
        lay = QHBoxLayout(f)
        lay.setContentsMargins(18, 12, 18, 12)
        lay.setSpacing(16)

        # ── SOL: ContraCore logo + başlık ─────────────────────────────────
        left = QHBoxLayout()
        left.setSpacing(12)

        self._lbl_logo = QLabel()
        self._lbl_logo.setStyleSheet('background:transparent;border:none;')
        self._logo_h = 76
        self._update_logo()
        left.addWidget(self._lbl_logo)

        title_col = QVBoxLayout()
        title_col.setSpacing(2)
        title_col.setContentsMargins(0, 0, 0, 0)

        title_row = QHBoxLayout()
        title_row.setSpacing(0)
        title_row.setContentsMargins(0, 0, 0, 0)

        from PySide6.QtGui import QPainterPath as _QPP2, QLinearGradient as _QLG2, QBrush as _QB2, QFontMetrics as _QFM2

        def _cc_pix(sz):
            """'Contra' navy gradient + 'CORE' gold gradient — tek pixmap."""
            font = QFont('Coolvetica', sz, QFont.Bold)
            fm   = _QFM2(font)
            cw   = fm.horizontalAdvance('Contra')
            ow   = fm.horizontalAdvance('CORE')
            h    = fm.height() + 4
            pix  = QPixmap(cw + ow + 4, h)
            pix.fill(Qt.transparent)
            p    = QPainter(pix)
            p.setRenderHint(QPainter.Antialiasing)
            pa1 = _QPP2(); pa1.addText(2, fm.ascent() + 2, font, 'Contra')
            g1 = _QLG2(0, 2, 0, h - 2)
            g1.setColorAt(0, QColor('#0a1e43')); g1.setColorAt(1, QColor('#081631'))
            p.setBrush(_QB2(g1)); p.setPen(Qt.NoPen); p.drawPath(pa1)
            pa2 = _QPP2(); pa2.addText(2 + cw, fm.ascent() + 2, font, 'CORE')
            g2 = _QLG2(0, 2, 0, h - 2)
            g2.setColorAt(0, QColor('#c8a45b')); g2.setColorAt(1, QColor('#96732d'))
            p.setBrush(_QB2(g2)); p.drawPath(pa2)
            p.end()
            return pix

        self._cc_lbl = QLabel()
        self._cc_lbl.setStyleSheet('background:transparent;border:none;')
        self._cc_lbl.setPixmap(_cc_pix(24))
        title_row.addWidget(self._cc_lbl)
        title_row.addStretch()

        title_col.addLayout(title_row)

        sub_lbl = QLabel('191 Muavin Karşılaştırma Otomasyonu')
        sub_lbl.setFont(QFont('Segoe UI', 14, QFont.Bold))
        sub_lbl.setStyleSheet(
            f'color:{TEXT}; background:transparent;border:none;letter-spacing:0.3px;'
        )
        title_col.addWidget(sub_lbl)
        left.addLayout(title_col)

        lay.addLayout(left)
        lay.addStretch(1)

        # ── SAĞ: Profil ───────────────────────────────────────────────────
        prof = QFrame()
        prof.setStyleSheet('''
            QFrame {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #1E3660, stop:1 #162D4E);
                border-radius: 32px;
                border: 1.5px solid #C9A46A;
            }
            QFrame:hover {
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 #253F6A, stop:1 #1E3660);
                border-color: #E4C285;
            }
        ''')
        _prof_glow = QGraphicsDropShadowEffect()
        _prof_glow.setBlurRadius(18)
        _prof_glow.setColor(QColor('#C9A46A55'))
        _prof_glow.setOffset(0, 2)
        prof.setGraphicsEffect(_prof_glow)

        pl = QHBoxLayout(prof)
        pl.setContentsMargins(14, 7, 14, 7)
        pl.setSpacing(10)

        p_img = QLabel()
        p_img.setFixedSize(46, 46)
        p_img.setStyleSheet('background:transparent;border:none;')
        px2 = QPixmap(ip('profile.png'))
        if not px2.isNull():
            canvas = QPixmap(46, 46)
            canvas.fill(Qt.transparent)
            pc = QPainter(canvas)
            pc.setRenderHint(QPainter.Antialiasing)
            scaled = px2.scaled(QSize(40, 40), Qt.KeepAspectRatioByExpanding, Qt.SmoothTransformation)
            clip = QPainterPath()
            clip.addEllipse(3.0, 3.0, 40.0, 40.0)
            pc.setClipPath(clip)
            pc.drawPixmap(3, 3, scaled)
            pc.setClipping(False)
            pc.setPen(QPen(QColor(GOLD), 2))
            pc.setBrush(Qt.NoBrush)
            pc.drawEllipse(QRectF(1, 1, 44, 44))
            _now = datetime.now()
            _online = (_now.weekday() <= 5) and (9 <= _now.hour < 17)
            pc.setPen(QPen(QColor('#FFFFFF'), 1.5))
            pc.setBrush(QColor(GREEN if _online else RED))
            pc.drawEllipse(33, 33, 11, 11)
            pc.end()
            p_img.setPixmap(canvas)
        else:
            p_img.setText('👤')
        pl.addWidget(p_img)

        pt = QVBoxLayout()
        pt.setSpacing(1)
        pt.setContentsMargins(0, 0, 0, 0)
        n = QLabel('Serkan ŞAHİN')
        n.setStyleSheet('font-size:12px;font-weight:700;color:#FFFFFF;background:transparent;border:none;')
        r = QLabel('Geliştirici İletişim')
        r.setStyleSheet(f'font-size:10px;color:{GOLD};background:transparent;border:none;')
        pt.addWidget(n); pt.addWidget(r)
        pl.addLayout(pt)

        arr = QLabel('▾')
        arr.setStyleSheet('font-size:11px;color:#C9A46A;background:transparent;border:none;')
        pl.addWidget(arr)

        prof.setCursor(Qt.PointingHandCursor)
        def _tog(e):
            if self._profile_popup._just_closed:
                return
            if self._profile_popup.isVisible():
                self._profile_popup.close()
            else:
                self._profile_popup.show_at(prof)
        prof.mousePressEvent = _tog
        lay.addWidget(prof)

        from core import theme as _theme
        self.btn_theme = _theme.make_toggle_button()
        self.btn_theme.clicked.connect(lambda: _theme.toggle())
        lay.addWidget(self.btn_theme)

        return f

    def _update_logo(self):
        import os as _os
        from core import theme as _t
        _LOGO_DIR = _os.path.join(_os.path.dirname(__file__), '..', '..', 'Logom', 'big_logo')
        fname = 'ContraCoreBeyaz.png' if _t.is_dark() else 'ContraCore.png'
        pix = QPixmap(_os.path.join(_LOGO_DIR, fname))
        if pix.isNull():
            pix = QPixmap(_os.path.join(_LOGO_DIR, 'ContraCore.png'))
        if not pix.isNull() and hasattr(self, '_lbl_logo'):
            h = self._logo_h
            w = int(pix.width() * h / pix.height())
            self._lbl_logo.setPixmap(pix.scaled(w, h, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        if hasattr(self, '_cc_lbl'):
            from PySide6.QtGui import QPainterPath as _QPP, QLinearGradient as _QLG, QBrush as _QB, QFontMetrics as _QFM
            def _cc_pix(sz):
                font = QFont('Coolvetica', sz, QFont.Bold)
                fm = _QFM(font)
                cw = fm.horizontalAdvance('Contra'); ow = fm.horizontalAdvance('CORE')
                h2 = fm.height() + 4
                p2 = QPixmap(cw + ow + 4, h2); p2.fill(Qt.transparent)
                pr = QPainter(p2); pr.setRenderHint(QPainter.Antialiasing)
                pa1 = _QPP(); pa1.addText(2, fm.ascent() + 2, font, 'Contra')
                g1 = _QLG(0, 2, 0, h2 - 2)
                if _t.is_dark():
                    g1.setColorAt(0, QColor('#FFFFFF')); g1.setColorAt(1, QColor('#E2E8F0'))
                else:
                    g1.setColorAt(0, QColor('#0a1e43')); g1.setColorAt(1, QColor('#081631'))
                pr.setBrush(_QB(g1)); pr.setPen(Qt.NoPen); pr.drawPath(pa1)
                pa2 = _QPP(); pa2.addText(2 + cw, fm.ascent() + 2, font, 'CORE')
                g2 = _QLG(0, 2, 0, h2 - 2)
                g2.setColorAt(0, QColor('#c8a45b')); g2.setColorAt(1, QColor('#96732d'))
                pr.setBrush(_QB(g2)); pr.drawPath(pa2); pr.end()
                return p2
            self._cc_lbl.setPixmap(_cc_pix(24))

    def _apply_theme(self):
        from core import theme as _theme
        if hasattr(self, 'btn_theme'):
            self.btn_theme.setText('☀️' if _theme.is_dark() else '🌙')
        self._update_logo()
        root = getattr(self, '_root_widget', None) or self.centralWidget()
        _theme.apply_to_widget(root)
        # Gauge'ı yeniden çiz (% rengi tema değişince)
        if hasattr(self, 'donut'):
            self.donut.update()

    def _trial_banner(self):
        """Deneme modunda üstte sabit bilgi şeridi. Lisanslıda gizlidir."""
        self.banner_trial = QFrame()
        self.banner_trial.setFixedHeight(38)
        self.banner_trial.setStyleSheet(
            'QFrame{background:#7C3AED;border-radius:10px;border:none;}'
        )

        lay = QHBoxLayout(self.banner_trial)
        lay.setContentsMargins(18, 0, 18, 0)
        lay.setSpacing(12)

        ico = QLabel('⏳')
        ico.setStyleSheet('background:transparent;border:none;font-size:14px;')
        lay.addWidget(ico)

        self.lbl_trial = QLabel()
        self.lbl_trial.setStyleSheet(
            'font-size:12px;font-weight:700;color:#FFFFFF;background:transparent;border:none;'
        )
        lay.addWidget(self.lbl_trial)
        lay.addStretch(1)

        tip = QLabel('Lisans almak için geliştiriciye ulaşın')
        tip.setStyleSheet(
            'font-size:10px;color:#DDD6FE;background:transparent;border:none;'
        )
        lay.addWidget(tip)

        if self._trial_status:
            kalan_gun, islenen, kalan = self._trial_status
            kota = islenen + kalan
            self.lbl_trial.setText(
                f'Deneme Sürümü  —  {kalan_gun} gün kaldı  |  '
                f'{islenen} / {kota} muavin satırı kullanıldı'
            )
            self.banner_trial.setVisible(True)
        else:
            self.banner_trial.setVisible(False)

        return self.banner_trial

    # ── Dosya Kartları ────────────────────────────────────────────────────────
    def _file_cards(self):
        f = self._card()
        outer = QVBoxLayout(f)
        outer.setContentsMargins(14, 14, 14, 14)
        outer.setSpacing(10)

        hdr = QHBoxLayout()
        hdr.setSpacing(6)
        hdr_ico = QLabel()
        hdr_ico.setStyleSheet('background:transparent;border:none;')
        pix_karsi = QPixmap(ip('karsi.png'))
        if not pix_karsi.isNull():
            hdr_ico.setPixmap(pix_karsi.scaled(16, 16, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            hdr_ico.setText('📋')
        hdr.addWidget(hdr_ico)
        lbl = QLabel('DOSYA SEÇİMİ')
        lbl.setStyleSheet(
            f'font-size:14px;font-weight:700;color:{TEXT};'
            f'background:transparent;border:none;letter-spacing:0.5px;'
        )
        hdr.addWidget(lbl)
        hdr.addStretch()

        btn_clear_all = QPushButton('🗑  Temizle')
        btn_clear_all.setCursor(Qt.PointingHandCursor)
        btn_clear_all.setStyleSheet(f'''
            QPushButton {{
                background:#F3F4F6;color:{TEXT2};border:1px solid {BORDER};
                border-radius:8px;font-size:11px;padding:5px 12px;
            }}
            QPushButton:hover{{background:{RED_BG};border-color:{RED};color:{RED};}}
        ''')
        btn_clear_all.clicked.connect(lambda: [self.fc_ikdvl.clear(), self.fc_muavin.clear()])
        hdr.addWidget(btn_clear_all)
        outer.addLayout(hdr)

        cards_row = QHBoxLayout()
        cards_row.setSpacing(0)

        self.fc_ikdvl  = FileCard(
            'İNDİRİLECEK KDV LİSTESİ', 'ikdvl.png', BLUE_BG,
            'Excel Dosyası (*.xlsx *.xls)'
        )
        self.fc_muavin = FileCard(
            '191 MUAVİN LİSTESİ', '191m.png', GRN_BG,
            'Excel Dosyası (*.xlsx *.xls)'
        )
        self.fc_ikdvl.changed.connect(lambda p: self._on_file_changed(p, 'ikdvl'))
        self.fc_muavin.changed.connect(lambda p: self._on_file_changed(p, 'muavin'))

        vs_lbl = QLabel('VS')
        vs_lbl.setFixedSize(40, 40)
        vs_lbl.setAlignment(Qt.AlignCenter)
        vs_lbl.setStyleSheet(
            f'background:{NAVY};color:#FFF;border-radius:20px;'
            f'font-size:12px;font-weight:900;border:none;'
        )

        cards_row.addWidget(self.fc_ikdvl, 1)
        cards_row.addSpacing(6)
        cards_row.addWidget(vs_lbl, 0, Qt.AlignVCenter)
        cards_row.addSpacing(6)
        cards_row.addWidget(self.fc_muavin, 1)
        outer.addLayout(cards_row)
        return f

    # ── Karşılaştırma Özeti ───────────────────────────────────────────────────
    def _stats_row(self):
        f = self._card()
        outer = QVBoxLayout(f)
        outer.setContentsMargins(14, 12, 14, 12)
        outer.setSpacing(10)

        # Başlık
        sec_row = QHBoxLayout()
        sec_row.setSpacing(6)
        sec_ico = QLabel()
        sec_ico.setStyleSheet('background:transparent;border:none;')
        pix_karsi2 = QPixmap(ip('karsi.png'))
        if not pix_karsi2.isNull():
            sec_ico.setPixmap(pix_karsi2.scaled(16, 16, Qt.KeepAspectRatio, Qt.SmoothTransformation))
        else:
            sec_ico.setText('📊')
        sec_lbl = QLabel('KARŞILAŞTIRMA ÖZETİ')
        sec_lbl.setStyleSheet(
            f'font-size:14px;font-weight:700;color:{TEXT};'
            f'background:transparent;border:none;letter-spacing:0.5px;'
        )
        self.btn_detay = QPushButton('📊  Detaylı Gör')
        self.btn_detay.setFixedHeight(26)
        self.btn_detay.setEnabled(False)
        self.btn_detay.setCursor(Qt.PointingHandCursor)
        self.btn_detay.setStyleSheet(f'''
            QPushButton{{background:{NAVY};color:#FFF;border-radius:6px;
            border:none;font-size:10px;font-weight:700;padding:0 10px;}}
            QPushButton:hover{{background:{NAVY2};}}
            QPushButton:disabled{{background:#ccc;color:#999;}}
        ''')
        self.btn_detay.clicked.connect(self._on_detail)

        sec_row.addWidget(sec_ico)
        sec_row.addWidget(sec_lbl)
        sec_row.addStretch()
        sec_row.addWidget(self.btn_detay)
        outer.addLayout(sec_row)

        cards = QHBoxLayout()
        cards.setSpacing(8)

        self.sc_eslesen  = SummaryCard('check.png',   'EŞLEŞEN\nFATURA',        '#cffddd', '#1ca957', '#35354d')
        self.sc_tutar    = SummaryCard('warning.png',  'TUTAR FARKI\nOLAN',       '#ffeacd', '#fca42b', '#88481a')
        self.sc_ikdvl    = SummaryCard('yukari.png',   'SADECE\nİKDVL\'DE OLAN',  '#dbe9fd', '#005ee6', '#1e3151')
        self.sc_muavin   = SummaryCard('asagi.png',    'SADECE\nMUAVİNDE OLAN',   '#ffe8e8', '#ff0d30', '#740f10')
        self.sc_toplam   = SummaryCard('toplam.png',   'TOPLAM\nFATURA',          '#ead5ff', '#5a3eab', '#291c4e')

        for sc in (self.sc_eslesen, self.sc_tutar, self.sc_ikdvl, self.sc_muavin, self.sc_toplam):
            cards.addWidget(sc, 1)

        outer.addLayout(cards)
        return f

    # ── Butonlar ──────────────────────────────────────────────────────────────
    def _buttons(self):
        f = QFrame()
        f.setStyleSheet('QFrame{background:transparent;border:none;}')
        lay = QHBoxLayout(f)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(10)

        self.btn_start = QPushButton('▶   KARŞILAŞTIRMAYI BAŞLAT\nDosyaları karşılaştır')
        self.btn_start.setFixedHeight(62)
        self.btn_start.setCursor(Qt.PointingHandCursor)
        self._style_start_idle()
        self.btn_start.clicked.connect(self._on_start)

        self.btn_duzelt = QPushButton('FARKLARI GÖRÜNTÜLE\nDetayları ve düzeltmeleri gör')
        pix_d = QPixmap(ip('disli.png'))
        if not pix_d.isNull():
            self.btn_duzelt.setIcon(QIcon(pix_d))
            self.btn_duzelt.setIconSize(QSize(22, 22))
        self.btn_duzelt.setFixedHeight(62)
        self.btn_duzelt.setEnabled(False)
        self.btn_duzelt.setCursor(Qt.PointingHandCursor)
        self.btn_duzelt.setStyleSheet('''
            QPushButton {
                background: #ffcc00; color: #532c02;
                border-radius: 14px; border: none;
                font-size: 12px; font-weight: 700; text-align: center;
            }
            QPushButton:hover { background: #f0be00; }
            QPushButton:disabled { background: #fff0a0; color: #c4a030; }
        ''')
        self.btn_duzelt.clicked.connect(self._on_duzelt)

        self.btn_report = QPushButton('RAPOR OLUŞTUR\nDetaylı Raporu Hazırla')
        pix_r = QPixmap(ip('rapor.png'))
        if not pix_r.isNull():
            self.btn_report.setIcon(QIcon(pix_r))
            self.btn_report.setIconSize(QSize(22, 22))
        self.btn_report.setFixedHeight(62)
        self.btn_report.setEnabled(False)
        self.btn_report.setCursor(Qt.PointingHandCursor)
        self.btn_report.setStyleSheet('''
            QPushButton {
                background: #1354e6; color: #fefefe;
                border-radius: 14px; border: none;
                font-size: 12px; font-weight: 700; text-align: center;
            }
            QPushButton:hover { background: #0f47cc; }
            QPushButton:disabled { background: #8aabf5; color: #d0dcfb; }
        ''')
        self.btn_report.clicked.connect(self._on_report)

        lay.addWidget(self.btn_start, 1)
        lay.addWidget(self.btn_duzelt, 1)
        lay.addWidget(self.btn_report, 1)
        return f

    def _style_start_idle(self):
        self.btn_start.setStyleSheet(f'''
            QPushButton {{
                background: {GREEN}; color: #FFF;
                border-radius: 14px; border: none;
                font-size: 12px; font-weight: 700; text-align: center;
            }}
            QPushButton:hover {{ background: {GREEN2}; }}
        ''')

    def _style_start_stop(self):
        self.btn_start.setStyleSheet(f'''
            QPushButton {{
                background: {RED}; color: #FFF;
                border-radius: 14px; border: none;
                font-size: 12px; font-weight: 700; text-align: center;
            }}
            QPushButton:hover {{ background: {RED2}; }}
        ''')

    # ── Progress ──────────────────────────────────────────────────────────────
    def _init_progress_widgets(self):
        """Progress bar ve ilgili etiketleri oluşturur (log kartına gömülecek)."""
        self.lbl_spinner = QLabel('')
        self.lbl_spinner.setStyleSheet(
            f'font-size:13px;color:{GREEN};background:transparent;border:none;'
        )
        self.lbl_status = QLabel('Hazır')   # sadece backward-compat için, görünmez
        self.lbl_pct = QLabel('')
        self.lbl_pct.setStyleSheet(
            f'font-size:10px;font-weight:700;color:{GREEN};background:transparent;border:none;'
        )
        self.pbar = QProgressBar()
        self.pbar.setRange(0, 1000)
        self.pbar.setValue(0)
        self.pbar.setFixedHeight(3)
        self.pbar.setTextVisible(False)
        self.pbar.setStyleSheet(f'''
            QProgressBar {{
                background:#1E293B; border-radius:0; border:none;
            }}
            QProgressBar::chunk {{
                background: qlineargradient(x1:0,y1:0,x2:1,y2:0,
                    stop:0 {GREEN}, stop:1 #16A34A);
                border-radius:0;
            }}
        ''')

    # ── Alt Paneller ──────────────────────────────────────────────────────────
    def _bottom_panels(self):
        self._init_progress_widgets()

        lay = QHBoxLayout()
        lay.setSpacing(10)

        # ── Sol: İşlem Günlükleri ──────────────────────────────────────────
        log_card = QFrame()
        log_card.setStyleSheet(
            f'QFrame{{background:{LOG_BG};border-radius:16px;border:1px solid #1E293B;}}'
        )
        log_lay = QVBoxLayout(log_card)
        log_lay.setContentsMargins(0, 0, 0, 0)
        log_lay.setSpacing(0)

        log_hdr = QFrame()
        log_hdr.setFixedHeight(38)
        log_hdr.setStyleSheet(
            f'QFrame{{background:{NAVY};border-top-left-radius:16px;border-top-right-radius:16px;border:none;}}'
        )
        lh = QHBoxLayout(log_hdr)
        lh.setContentsMargins(14, 0, 10, 0)
        lh.setSpacing(6)

        dot = QLabel('●')
        dot.setStyleSheet('font-size:10px;color:#4DCC78;background:transparent;border:none;')
        ttl = QLabel('İŞLEM GÜNLÜKLERİ')
        ttl.setStyleSheet(
            'font-size:11px;font-weight:700;color:#CBD5E1;'
            'background:transparent;border:none;letter-spacing:0.5px;'
        )
        lh.addWidget(dot)
        lh.addWidget(self.lbl_spinner)
        lh.addWidget(ttl, 1)
        self.lbl_sys = QLabel('● Hazır')
        self.lbl_sys.setStyleSheet(
            f'font-size:10px;color:{GREEN};background:transparent;border:none;'
        )
        lh.addWidget(self.lbl_sys)
        lh.addWidget(self.lbl_pct)

        btn_clear_log = QPushButton('🗑')
        btn_clear_log.setFixedSize(28, 26)
        btn_clear_log.setCursor(Qt.PointingHandCursor)
        btn_clear_log.setStyleSheet(
            'QPushButton{background:rgba(255,255,255,0.08);color:#9CA3AF;'
            'border-radius:6px;border:none;font-size:12px;}'
            'QPushButton:hover{background:rgba(239,68,68,0.3);color:#FCA5A5;}'
        )
        btn_clear_log.clicked.connect(lambda: self.log_box.clear())
        lh.addWidget(btn_clear_log)

        btn_save_log = QPushButton('💾')
        btn_save_log.setFixedSize(28, 26)
        btn_save_log.setCursor(Qt.PointingHandCursor)
        btn_save_log.setStyleSheet(
            'QPushButton{background:rgba(255,255,255,0.08);color:#9CA3AF;'
            'border-radius:6px;border:none;font-size:12px;}'
            'QPushButton:hover{background:rgba(34,197,94,0.3);color:#86EFAC;}'
        )
        btn_save_log.clicked.connect(self._save_log)
        lh.addWidget(btn_save_log)
        log_lay.addWidget(log_hdr)
        log_lay.addWidget(self.pbar)

        self.log_box = QTextEdit()
        self.log_box.setReadOnly(True)
        self.log_box.setStyleSheet(f'''
            QTextEdit {{
                background:{LOG_BG};color:#CBD5E1;
                border:none;border-radius:0;
                font-family:Consolas,monospace;font-size:11px;
                padding:10px 12px;
            }}
            QScrollBar:vertical {{
                background:#1E293B;width:8px;border-radius:4px;margin:2px;
            }}
            QScrollBar::handle:vertical {{
                background:#334155;border-radius:4px;min-height:24px;
            }}
            QScrollBar::add-line:vertical, QScrollBar::sub-line:vertical {{
                height:0px;
            }}
        ''')
        log_lay.addWidget(self.log_box, 1)
        # Alt köşe yuvarlama için boşluk
        log_lay.addSpacing(12)

        # ── Sağ: Dosya Bilgileri + KDV Özeti ─────────────────────────────
        info_card = self._card()
        info_lay  = QVBoxLayout(info_card)
        info_lay.setContentsMargins(16, 14, 16, 14)
        info_lay.setSpacing(0)

        def _sec_ttl(txt):
            w = QFrame()
            w.setStyleSheet('QFrame{background:transparent;border:none;}')
            wl = QHBoxLayout(w)
            wl.setContentsMargins(0, 0, 0, 0)
            wl.setSpacing(8)
            accent = QFrame()
            accent.setFixedSize(4, 16)
            accent.setStyleSheet(f'background:{NAVY};border-radius:2px;border:none;')
            lbl = QLabel(txt)
            lbl.setStyleSheet(
                f'font-size:12px;font-weight:800;color:{TEXT};'
                f'background:transparent;border:none;letter-spacing:1.2px;'
            )
            wl.addWidget(accent)
            wl.addWidget(lbl)
            wl.addStretch()
            return w

        def _sep():
            f = QFrame(); f.setFrameShape(QFrame.HLine)
            f.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;margin:8px 0;')
            return f

        # ── Dosya Bilgileri ───────────────────────────────────────────────
        info_lay.addWidget(_sec_ttl('DOSYA BİLGİLERİ'))
        info_lay.addSpacing(8)

        def _file_status_row(label, dot_attr):
            frm = QFrame()
            frm.setStyleSheet(
                f'QFrame{{background:#F8FAFC;border-radius:10px;border:1px solid {BORDER};}}'
            )
            fl = QHBoxLayout(frm)
            fl.setContentsMargins(12, 8, 12, 8)
            fl.setSpacing(10)
            dot = QLabel('●')
            dot.setFixedWidth(14)
            dot.setStyleSheet(f'font-size:12px;color:{BORDER};background:transparent;border:none;')
            lk = QLabel(label)
            lk.setStyleSheet(
                f'font-size:12px;font-weight:600;color:{TEXT};background:transparent;border:none;'
            )
            fl.addWidget(dot)
            fl.addWidget(lk, 1)
            setattr(self, dot_attr, dot)
            return frm

        info_lay.addWidget(_file_status_row('İndirilecek KDV Listesi', 'dot_ikdvl'))
        info_lay.addSpacing(6)
        info_lay.addWidget(_file_status_row('191 Muavin Listesi', 'dot_muavin'))
        info_lay.addSpacing(8)

        # Dönem uyumsuzluğu uyarısı
        self.frm_donem_warn = QFrame()
        self.frm_donem_warn.setStyleSheet(
            'QFrame{background:#FFFBEB;border-radius:10px;border:1px solid #FDE68A;}'
        )
        wl = QHBoxLayout(self.frm_donem_warn)
        wl.setContentsMargins(10, 8, 10, 8)
        wl.setSpacing(8)
        warn_ico = QLabel('⚠️')
        warn_ico.setStyleSheet('font-size:15px;background:transparent;border:none;')
        self.lbl_donem_warn = QLabel('Dönem uyumsuzluğu!')
        self.lbl_donem_warn.setStyleSheet(
            'font-size:11px;font-weight:600;color:#92400E;background:transparent;border:none;'
        )
        self.lbl_donem_warn.setWordWrap(True)
        wl.addWidget(warn_ico, 0, Qt.AlignTop)
        wl.addWidget(self.lbl_donem_warn, 1)
        self.frm_donem_warn.hide()
        info_lay.addWidget(self.frm_donem_warn)

        info_lay.addWidget(_sep())

        # ── KDV Özeti ─────────────────────────────────────────────────────
        info_lay.addWidget(_sec_ttl('KDV ÖZETİ'))
        info_lay.addSpacing(8)

        def _kdv_card(label, attr_name, bg, val_color, left_border_color):
            frm = QFrame()
            frm.setStyleSheet(
                f'QFrame{{background:{bg};border-radius:10px;'
                f'border:1px solid {BORDER};'
                f'border-left:4px solid {left_border_color};}}'
            )
            fl = QVBoxLayout(frm)
            fl.setContentsMargins(12, 8, 12, 8)
            fl.setSpacing(2)
            lk = QLabel(label)
            lk.setStyleSheet(
                f'font-size:10px;font-weight:700;color:{TEXT3};'
                f'background:transparent;border:none;letter-spacing:0.5px;'
            )
            lv = QLabel('—')
            lv.setStyleSheet(
                f'font-size:15px;font-weight:800;color:{val_color};'
                f'background:transparent;border:none;'
            )
            fl.addWidget(lk)
            fl.addWidget(lv)
            setattr(self, attr_name, lv)
            return frm

        # KDV kartları sol, donut sağ yan yana
        kdv_donut_row = QHBoxLayout()
        kdv_donut_row.setSpacing(10)

        cards_col = QVBoxLayout()
        cards_col.setSpacing(5)
        cards_col.addWidget(_kdv_card('IKDVL TOPLAM KDV', 'lbl_kdv_ikdvl', '#EFF6FF', NAVY, NAVY))
        cards_col.addWidget(_kdv_card('MUAVİN TOPLAM KDV', 'lbl_kdv_muavin', '#F0FDF4', '#15803D', '#16A34A'))
        cards_col.addWidget(_kdv_card('FARK', 'lbl_kdv_fark', '#FEF2F2', RED, RED))

        self.donut = DonutChart()
        self.donut.setFixedSize(100, 100)

        donut_col = QVBoxLayout()
        donut_col.setSpacing(4)
        donut_lbl = QLabel('EŞLEŞME\nORANI')
        donut_lbl.setAlignment(Qt.AlignCenter)
        donut_lbl.setStyleSheet(
            f'font-size:10px;font-weight:800;color:{TEXT};'
            f'background:transparent;border:none;letter-spacing:1px;'
        )
        donut_col.addStretch()
        donut_col.addWidget(donut_lbl, 0, Qt.AlignHCenter)
        donut_col.addWidget(self.donut, 0, Qt.AlignHCenter)
        donut_col.addStretch()

        kdv_donut_row.addLayout(cards_col, 1)
        kdv_donut_row.addLayout(donut_col)
        info_lay.addLayout(kdv_donut_row)

        info_lay.addStretch()

        # Backward-compat: _detail_rows gizli etiketler olarak tutuluyor
        self._detail_rows = {}
        for key in ('kars_sure', 'islenen_kayit', 'duzeltilecek', 'durum'):
            lv = QLabel('—')
            lv.hide()
            self._detail_rows[key] = lv

        # Backward-compat: fname labels (artık görünmüyor, dot_* ile değiştirildi)
        self.lbl_ikdvl_fname = QLabel()
        self.lbl_ikdvl_fname.hide()
        self.lbl_muavin_fname = QLabel()
        self.lbl_muavin_fname.hide()

        lay.addWidget(log_card, 3)
        lay.addWidget(info_card, 2)
        return lay

    def _info_box(self, emoji, label, value):
        box = QFrame()
        box.setStyleSheet(
            f'QFrame{{background:#F8FAFC;border-radius:10px;border:1px solid {BORDER};}}'
        )
        bl = QVBoxLayout(box)
        bl.setContentsMargins(12, 8, 12, 8)
        bl.setSpacing(2)

        top = QHBoxLayout()
        ico = QLabel(emoji)
        ico.setStyleSheet('font-size:14px;background:transparent;border:none;')
        lbl = QLabel(label)
        lbl.setStyleSheet(f'font-size:9px;font-weight:700;color:{TEXT3};background:transparent;border:none;letter-spacing:0.5px;')
        top.addWidget(ico); top.addWidget(lbl, 1)
        bl.addLayout(top)

        val_lbl = QLabel(value)
        val_lbl.setStyleSheet(f'font-size:13px;font-weight:700;color:{NAVY};background:transparent;border:none;')
        val_lbl.setObjectName(label)
        bl.addWidget(val_lbl)

        # Referansı sakla
        box._val_lbl = val_lbl
        return box

    def _detail_row(self, label, value):
        row = QHBoxLayout()
        row.setSpacing(6)
        lbl_k = QLabel(label)
        lbl_k.setStyleSheet(f'font-size:10px;color:{TEXT2};background:transparent;')
        lbl_v = QLabel(value)
        lbl_v.setStyleSheet(f'font-size:10px;font-weight:700;color:{TEXT};background:transparent;')
        lbl_v.setAlignment(Qt.AlignRight)
        row.addWidget(lbl_k); row.addStretch(); row.addWidget(lbl_v)
        return row, lbl_v

    def _premium_row(self, ico_txt, label, value):
        """Düz bilgi satırı — başlık solda, değer sağda, ikisi de bold siyah."""
        box = QFrame()
        box.setStyleSheet('QFrame{background:transparent;border:none;}')
        bl = QHBoxLayout(box)
        bl.setContentsMargins(4, 6, 4, 6)
        bl.setSpacing(8)

        lbl_key = QLabel(label)
        lbl_key.setStyleSheet(
            f'font-size:13px;font-weight:700;color:{TEXT};'
            f'background:transparent;border:none;'
        )
        val_lbl = QLabel(value)
        val_lbl.setStyleSheet(
            f'font-size:13px;font-weight:700;color:{TEXT};'
            f'background:transparent;border:none;'
        )
        val_lbl.setAlignment(Qt.AlignRight | Qt.AlignVCenter)

        bl.addWidget(lbl_key)
        bl.addStretch()
        bl.addWidget(val_lbl)

        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')

        wrapper = QFrame()
        wrapper.setStyleSheet('QFrame{background:transparent;border:none;}')
        wl = QVBoxLayout(wrapper)
        wl.setContentsMargins(0, 0, 0, 0)
        wl.setSpacing(0)
        wl.addWidget(box)
        wl.addWidget(sep)

        return wrapper, val_lbl

    # ── Olaylar ───────────────────────────────────────────────────────────────
    def _on_file_changed(self, path, file_type=''):
        self.btn_report.setEnabled(False)
        self.btn_duzelt.setEnabled(False)
        self._last_result = None

        if file_type == 'ikdvl':
            if path:
                self._ikdvl_months = self._detect_months(path, 'ikdvl')
                self.dot_ikdvl.setStyleSheet(f'font-size:12px;color:{GREEN};background:transparent;border:none;')
            else:
                self._ikdvl_months = set()
                self.dot_ikdvl.setStyleSheet(f'font-size:12px;color:{BORDER};background:transparent;border:none;')
        elif file_type == 'muavin':
            if path:
                self._muavin_months = self._detect_months(path, 'muavin')
                self.dot_muavin.setStyleSheet(f'font-size:12px;color:{GREEN};background:transparent;border:none;')
            else:
                self._muavin_months = set()
                self.dot_muavin.setStyleSheet(f'font-size:12px;color:{BORDER};background:transparent;border:none;')

        self._check_period_match()
        # KDV özetini sıfırla
        for lbl in (self.lbl_kdv_ikdvl, self.lbl_kdv_muavin, self.lbl_kdv_fark):
            lbl.setText('—')
        self.donut.reset()

    _AYLAR_TR_DICT = {1:'OCAK',2:'ŞUBAT',3:'MART',4:'NİSAN',5:'MAYIS',6:'HAZİRAN',
                      7:'TEMMUZ',8:'AĞUSTOS',9:'EYLÜL',10:'EKİM',11:'KASIM',12:'ARALIK'}

    def _detect_months(self, path, file_type) -> set:
        """Dosyadaki tüm benzersiz (yil, ay) çiftlerini döndürür."""
        months = set()
        try:
            if file_type == 'ikdvl':
                if path.lower().endswith('.xls'):
                    import xlrd
                    xwb = xlrd.open_workbook(path)
                    xws = next((xwb.sheet_by_name(s) for s in xwb.sheet_names() if 'KDV' in s.upper()), xwb.sheet_by_index(0))
                    for ri in range(4, xws.nrows):
                        val = xws.cell_value(ri, 16)
                        if val:
                            s = str(val).strip().split('.')[0]
                            if len(s) == 6 and s.isdigit():
                                months.add((int(s[:4]), int(s[4:])))
                else:
                    import openpyxl
                    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                    ws = wb.active
                    for row in ws.iter_rows(min_row=5, min_col=17, max_col=17, values_only=True):
                        val = row[0]
                        if val:
                            s = str(val).strip()
                            if len(s) == 6 and s.isdigit():
                                months.add((int(s[:4]), int(s[4:])))
                            elif '.' in s:
                                parts = s.split('.')
                                if len(parts) == 2:
                                    try: months.add((int(parts[1]), int(parts[0])))
                                    except ValueError: pass
                    wb.close()
            elif file_type == 'muavin':
                import openpyxl
                from datetime import date as _date
                wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
                ws = wb['HAM'] if 'HAM' in wb.sheetnames else wb.active
                for row in ws.iter_rows(min_row=7, min_col=1, max_col=1, values_only=True):
                    val = row[0]
                    if val:
                        if isinstance(val, _date):
                            months.add((val.year, val.month))
                        else:
                            s = str(val).strip()
                            if '.' in s:
                                parts = s.split('.')
                                if len(parts) >= 3:
                                    try: months.add((int(parts[2]), int(parts[1])))
                                    except ValueError: pass
                wb.close()
        except Exception:
            pass
        return months

    def _check_period_match(self):
        m1 = getattr(self, '_ikdvl_months', set())
        m2 = getattr(self, '_muavin_months', set())
        if m1 and m2 and not m1.intersection(m2):
            def _fmt(months):
                return ', '.join(
                    f'{self._AYLAR_TR_DICT.get(ay, str(ay))} {yil}'
                    for yil, ay in sorted(months)
                )
            self.lbl_donem_warn.setText(
                f'Ortak dönem yok — IKDVL: {_fmt(m1)} | Muavin: {_fmt(m2)}'
            )
            self.frm_donem_warn.show()
        else:
            self.frm_donem_warn.hide()

    def _on_start(self):
        if self._running:
            self._stop_flag.set()
            self._log_append('⛔  Durdurma isteği gönderildi...', 'warn')
            self.btn_start.setEnabled(False)
            return

        p1 = self.fc_ikdvl.get_path()
        p2 = self.fc_muavin.get_path()
        if not p1 or not os.path.exists(p1):
            QMessageBox.warning(self, 'Dosya Eksik', 'Lütfen İndirilecek KDV Listesi dosyasını seçin.')
            return
        if not p2 or not os.path.exists(p2):
            QMessageBox.warning(self, 'Dosya Eksik', 'Lütfen 191 Muavin dosyasını seçin.')
            return

        # Deneme kotası kontrolü — kalan 0 ise başlatma
        if self._trial_status:
            try:
                from license import get_trial_status
                aktif, _, _, kalan = get_trial_status()
                if not aktif or kalan <= 0:
                    QMessageBox.warning(self, 'Deneme Kotası Doldu',
                        'Deneme muavin satır limitine ulaşıldı veya deneme süresi doldu.\n'
                        'Devam etmek için lisans satın alın.')
                    return
            except Exception:
                pass

        self._running = True
        self._stop_flag.clear()
        self._t0      = time.time()
        self._islem_baslama = datetime.now()
        self.btn_report.setEnabled(False)
        self.btn_duzelt.setEnabled(False)
        self._last_result = None
        self.btn_start.setText('⏹   DURDUR\nİşlemi durdur')
        self._style_start_stop()
        self.pbar.setValue(0)
        self.lbl_pct.setText('')
        self.lbl_sys.setText('● Çalışıyor...')
        self.lbl_sys.setStyleSheet(
            f'font-size:10px;color:{GOLD};background:transparent;border:none;'
        )
        self._spin_timer.start()

        for sc in (self.sc_eslesen, self.sc_tutar, self.sc_ikdvl, self.sc_muavin, self.sc_toplam):
            sc.reset()

        self._detail_rows['kars_sure'].setText('...')
        self._detail_rows['islenen_kayit'].setText('...')
        self._detail_rows['duzeltilecek'].setText('...')
        self._detail_rows['durum'].setText('Çalışıyor...')

        # Trial modundaysa kalan satır limitini worker'a ilet
        max_rows = None
        if self._trial_status:
            try:
                from license import get_trial_status
                _, _, _, kalan = get_trial_status()
                max_rows = kalan if kalan > 0 else 0
            except Exception:
                pass

        self.worker = KarsilastirWorker(p1, p2, self._stop_flag, max_muavin_rows=max_rows)
        self.worker.sig.log.connect(self._on_log)
        self.worker.sig.progress.connect(self._on_progress)
        self.worker.sig.stats.connect(self._on_stats)
        self.worker.sig.done.connect(self._on_done)
        self.worker.start()

    # ─────────────────────────────────────────────────────────────────────────
    def _on_detail(self):
        if not self._last_result:
            return
        res = self._last_result

        _AYLAR_TR = {1:'OCAK',2:'ŞUBAT',3:'MART',4:'NİSAN',5:'MAYIS',6:'HAZİRAN',
                     7:'TEMMUZ',8:'AĞUSTOS',9:'EYLÜL',10:'EKİM',11:'KASIM',12:'ARALIK'}

        def _ay_label(ay, yil):
            return f'{_AYLAR_TR.get(ay, str(ay))} {yil}' if ay and yil else '—'

        def _fmt_tutar(v):
            try: return f'{float(v):,.2f} ₺'
            except: return str(v or '—')

        # ── Veri hazırlığı ────────────────────────────────────────────────────
        def _srt(lst, kfn): return sorted(lst, key=lambda x: (kfn(x)[0] or 0, kfn(x)[1] or 0))
        def _key_es(p):  return (p['ikdvl'].get('yil'), p['ikdvl'].get('ay'))
        def _key_ir(x):  return (x.get('yil'), x.get('ay'))

        es_all = _srt(res.get('eslesen',    []), _key_es)
        tf_all = _srt(res.get('tutar_farki',[]), _key_es)
        if_all = _srt(res.get('ikdvl_fazla',[]), _key_ir)
        mf_all = _srt(res.get('muavin_fazla',[]),_key_ir)

        # IKDVL aylarını alı (sekme yapısı için)
        _ikdvl_months = sorted(
            {(ir.get('yil'), ir.get('ay')) for ir in res.get('ikdvl_rows', [])
             if ir.get('yil') and ir.get('ay')},
            key=lambda k: (k[0], k[1])
        )

        def _es_row(p):
            ir = p['ikdvl']
            return [ir.get('fatura_no',''), ir.get('tarih',''),
                    (ir.get('satici') or '')[:40],
                    _fmt_tutar(ir.get('toplam_kdv')), _fmt_tutar(p.get('muavin_tutar'))]
        def _tf_row(p):
            ir = p['ikdvl']
            return [ir.get('fatura_no',''), ir.get('tarih',''),
                    (ir.get('satici') or '')[:40],
                    _fmt_tutar(ir.get('toplam_kdv')), _fmt_tutar(p.get('muavin_tutar')),
                    _fmt_tutar(p.get('fark'))]
        def _if_row(ir):
            return [ir.get('fatura_no',''), ir.get('tarih',''),
                    (ir.get('satici') or '')[:40], _fmt_tutar(ir.get('toplam_kdv'))]
        def _mf_row(mr):
            return [mr.get('fatura_no',''), mr.get('tarih',''),
                    (mr.get('firma') or '')[:40], _fmt_tutar(mr.get('toplam_kdv'))]

        # Tümü kategorisi için birleşik veri
        TUM_HDRS = ['FATURA NO','TARİH','FİRMA / SATICI','TUTAR','TİP']
        TUM_COLORS = {
            'es': '#d1fae5', 'tf': '#fef3c7',
            'if': '#dbeafe', 'mf': '#fee2e2',
        }
        def _tum_rows_for(month_filter=None):
            rows = []
            for p in es_all:
                k = _key_es(p)
                if month_filter and k != month_filter: continue
                ir = p['ikdvl']
                rows.append(('es', [ir.get('fatura_no',''), ir.get('tarih',''),
                    (ir.get('satici') or '')[:40], _fmt_tutar(ir.get('toplam_kdv')),
                    'Eşleşen']))
            for p in tf_all:
                k = _key_es(p)
                if month_filter and k != month_filter: continue
                ir = p['ikdvl']
                rows.append(('tf', [ir.get('fatura_no',''), ir.get('tarih',''),
                    (ir.get('satici') or '')[:40], _fmt_tutar(ir.get('toplam_kdv')),
                    'Tutar Farkı']))
            for x in if_all:
                k = _key_ir(x)
                if month_filter and k != month_filter: continue
                rows.append(('if', [x.get('fatura_no',''), x.get('tarih',''),
                    (x.get('satici') or '')[:40], _fmt_tutar(x.get('toplam_kdv')),
                    'IKDVL Fazlası']))
            for x in mf_all:
                k = _key_ir(x)
                if month_filter and k != month_filter: continue
                rows.append(('mf', [x.get('fatura_no',''), x.get('tarih',''),
                    (x.get('firma') or '')[:40], _fmt_tutar(x.get('toplam_kdv')),
                    'Muavin Fazlası']))
            return rows

        # (label, accent, row_bg, data, headers, row_fn, key_fn)
        CATS = [
            ('🔷  Tümü',           '#7c3aed', '#f5f3ff', None,
             TUM_HDRS, None, None),
            ('✅  Eşleşen',        '#1ca957', '#d1fae5', es_all,
             ['FATURA NO','TARİH','FİRMA','IKDVL KDV','MUAVİN KDV'], _es_row, _key_es),
            ('⚠️  Tutar Farkı',    '#b45309', '#fef3c7', tf_all,
             ['FATURA NO','TARİH','FİRMA','IKDVL','MUAVİN','FARK'],  _tf_row, _key_es),
            ('🔵  IKDVL Fazlası',  '#1d4ed8', '#dbeafe', if_all,
             ['FATURA NO','TARİH','SATICI','TUTAR'],                   _if_row, _key_ir),
            ('🔴  Muavin Fazlası', '#dc2626', '#fee2e2', mf_all,
             ['FATURA NO','TARİH','FİRMA','TUTAR'],                    _mf_row, _key_ir),
        ]

        # ── Dialog ────────────────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle('KARŞILAŞTIRMA DETAYLARI')
        dlg.resize(1150, 700)
        dlg.setStyleSheet(f'QDialog{{background:{BG};}}')
        dl = QVBoxLayout(dlg)
        dl.setContentsMargins(14, 12, 14, 12)
        dl.setSpacing(8)

        ttl = QLabel('📊  KARŞILAŞTIRMA DETAYLARI')
        ttl.setStyleSheet(f'font-size:13px;font-weight:800;color:{NAVY};background:transparent;')
        dl.addWidget(ttl)

        body = QHBoxLayout()
        body.setSpacing(10)

        # ── Sol sidebar ───────────────────────────────────────────────────────
        sidebar = QFrame()
        sidebar.setFixedWidth(205)
        sidebar.setStyleSheet(
            f'QFrame{{background:{CARD};border-radius:10px;border:1px solid #dde3ec;}}'
            f'QFrame QLabel{{background:transparent;border:none;}}'
            f'QFrame QPushButton{{border:none;}}'
        )
        sb_lay = QVBoxLayout(sidebar)
        sb_lay.setContentsMargins(8, 12, 8, 12)
        sb_lay.setSpacing(5)
        sb_lbl = QLabel('KATEGORİLER')
        sb_lbl.setStyleSheet(f'font-size:9px;font-weight:700;color:{TEXT2};letter-spacing:1px;')
        sb_lay.addWidget(sb_lbl)

        # ── Sağ panel ─────────────────────────────────────────────────────────
        stack = QStackedWidget()

        # ── Ortak stiller ─────────────────────────────────────────────────────
        HDR_STYLE = (
            f'QHeaderView::section{{background:{NAVY};color:#FFF;font-weight:700;'
            f'font-size:11px;padding:7px 8px;border:none;border-right:1px solid #1a3555;}}'
        )
        SRCH_STYLE = (
            f'QLineEdit{{background:#fff;border:1px solid #c5cfe0;border-radius:6px;'
            f'padding:0 10px;font-size:11px;color:{TEXT};}}'
        )
        NAV_STYLE = (
            'QTabBar QToolButton{'
            f'background:{NAVY};color:#FFF;border-radius:4px;'
            f'min-width:22px;min-height:22px;}}'
            'QTabBar QToolButton:hover{'
            f'background:{NAVY2};}}'
            'QTabBar::scroller{width:26px;}'
        )

        def _show_copy_toast(dlg_ref, txt):
            toast = QLabel(f'📋  Kopyalandı: {txt}', dlg_ref)
            toast.setStyleSheet(
                f'background:{NAVY};color:#FFF;border-radius:8px;'
                f'font-size:11px;font-weight:600;padding:6px 14px;border:none;'
            )
            toast.adjustSize()
            toast.move(
                (dlg_ref.width()  - toast.width())  // 2,
                dlg_ref.height() - toast.height() - 20
            )
            toast.show()
            QTimer.singleShot(1800, toast.deleteLater)

        def _make_tbl(headers, row_bg):
            t = QTableWidget()
            t.setColumnCount(len(headers))
            t.setHorizontalHeaderLabels(headers)
            t.setEditTriggers(QAbstractItemView.NoEditTriggers)
            t.setSelectionBehavior(QAbstractItemView.SelectRows)
            t.setAlternatingRowColors(False)
            t.verticalHeader().setVisible(False)
            t.horizontalHeader().setStretchLastSection(True)
            t.setStyleSheet(
                f'QTableWidget{{background:{row_bg};gridline-color:#d1d9e6;'
                f'border:none;font-size:11px;outline:0;}}'
                f'QTableWidget::item{{color:{TEXT};padding:5px 8px;}}'
                f'QTableWidget::item:selected{{background:#c7d2fe;color:{TEXT};}}'
                + HDR_STYLE
            )
            def _on_dbl(item):
                # Sütun 0 = Fatura No; diğer sütunlarda da kopyalanabilir
                val = item.text().strip()
                if val and not val.startswith(' '):  # ay header satırlarını atla
                    QApplication.clipboard().setText(val)
                    _show_copy_toast(dlg, val[:40])
            t.itemDoubleClicked.connect(_on_dbl)
            return t

        def _fill_tbl_cat(tbl, rows_data, row_fn, key_fn, accent, row_bg):
            tbl.setRowCount(0)
            cur_k = object()
            for item in rows_data:
                k = key_fn(item)
                if k != cur_k:
                    cur_k = k
                    sr = tbl.rowCount(); tbl.insertRow(sr)
                    it0 = QTableWidgetItem(f'   {_ay_label(*k)}')
                    it0.setBackground(QColor(accent))
                    it0.setForeground(QColor('#FFFFFF'))
                    fnt = it0.font(); fnt.setBold(True); fnt.setPointSize(10); it0.setFont(fnt)
                    tbl.setItem(sr, 0, it0)
                    for cc in range(1, tbl.columnCount()):
                        itx = QTableWidgetItem('')
                        itx.setBackground(QColor(accent))
                        tbl.setItem(sr, cc, itx)
                    tbl.setSpan(sr, 0, 1, tbl.columnCount())
                    tbl.setRowHeight(sr, 28)
                rr = tbl.rowCount(); tbl.insertRow(rr)
                for ci2, v in enumerate(row_fn(item)):
                    ci2t = QTableWidgetItem(str(v))
                    ci2t.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                    ci2t.setBackground(QColor(row_bg))
                    tbl.setItem(rr, ci2, ci2t)
                tbl.setRowHeight(rr, 26)
            tbl.resizeColumnsToContents()

        def _fill_tbl_tum(tbl, typed_rows):
            tbl.setRowCount(0)
            for typ, vals in typed_rows:
                rr = tbl.rowCount(); tbl.insertRow(rr)
                bg = TUM_COLORS.get(typ, '#fff')
                for ci2, v in enumerate(vals):
                    it = QTableWidgetItem(str(v))
                    it.setTextAlignment(Qt.AlignCenter | Qt.AlignVCenter)
                    it.setBackground(QColor(bg))
                    tbl.setItem(rr, ci2, it)
                tbl.setRowHeight(rr, 26)
            tbl.resizeColumnsToContents()

        def _search_wrap(tbl):
            w = QWidget(); w.setStyleSheet('background:transparent;border:none;')
            wl = QVBoxLayout(w)
            wl.setContentsMargins(8, 8, 8, 8); wl.setSpacing(6)
            srch = QLineEdit()
            srch.setPlaceholderText('🔍  Fatura no, firma veya tutar ara...')
            srch.setFixedHeight(32); srch.setStyleSheet(SRCH_STYLE)
            wl.addWidget(srch); wl.addWidget(tbl, 1)
            def _flt(txt):
                t = txt.lower()
                for row in range(tbl.rowCount()):
                    tbl.setRowHidden(row, bool(t) and not any(
                        tbl.item(row, c) and t in tbl.item(row, c).text().lower()
                        for c in range(tbl.columnCount())
                    ))
            srch.textChanged.connect(_flt)
            return w

        def _make_month_tabs(accent, tab_fn):
            """IKDVL aylarına göre Tümü + ay sekmeleri oluşturur."""
            tw = QTabWidget()
            tw.setUsesScrollButtons(True)
            tw.setElideMode(Qt.ElideNone)
            tw.setStyleSheet(
                f'QTabWidget::pane{{border:none;background:transparent;}}'
                f'QTabBar::tab{{background:#e8edf5;color:{TEXT};padding:7px 14px;'
                f'border-radius:5px 5px 0 0;font-size:11px;font-weight:600;'
                f'margin-right:2px;min-width:70px;}}'
                f'QTabBar::tab:selected{{background:{accent};color:#FFF;font-weight:700;}}'
                f'QTabBar::tab:hover{{background:#d4dce8;}}'
                + NAV_STYLE
            )
            # Tümü sekmesi
            w_all, tbl_all = tab_fn(None)
            tw.addTab(_search_wrap(tbl_all), 'TÜMÜ')
            # Aylık sekmeler (IKDVL ayları)
            for (yil, ay) in _ikdvl_months:
                w_m, tbl_m = tab_fn((yil, ay))
                tw.addTab(_search_wrap(tbl_m), _ay_label(ay, yil))
            return tw

        # Sidebar butonları
        cat_btns = []

        # Lazy-populate: her kategori için içerik ilk tıklamada oluşturulur
        _cat_populated = [False] * len(CATS)
        _cat_widgets   = [None]  * len(CATS)

        def _build_cat_widget(idx):
            label, accent, row_bg, data, headers, row_fn, key_fn = CATS[idx]
            is_tum = (data is None)
            if is_tum:
                def _tum_tab_fn(month_key):
                    tbl = _make_tbl(TUM_HDRS, row_bg)
                    typed = _tum_rows_for(month_key)
                    _fill_tbl_tum(tbl, typed)
                    return None, tbl
                return _make_month_tabs(accent, _tum_tab_fn)
            else:
                def _cat_tab_fn(month_key, _data=data, _rfn=row_fn, _kfn=key_fn,
                                _ac=accent, _rb=row_bg, _hdr=headers):
                    tbl = _make_tbl(_hdr, _rb)
                    rows = [x for x in _data if _kfn(x) == month_key] if month_key else _data
                    _fill_tbl_cat(tbl, rows, _rfn, _kfn, _ac, _rb)
                    return None, tbl
                return _make_month_tabs(accent, _cat_tab_fn)

        def _select_cat(idx):
            # İlk kez seçiliyorsa içeriği şimdi oluştur
            if not _cat_populated[idx]:
                tw = _build_cat_widget(idx)
                _cat_widgets[idx] = tw
                stack.insertWidget(idx, tw)
                _cat_populated[idx] = True
            stack.setCurrentIndex(idx)
            for i, b in enumerate(cat_btns):
                ac = CATS[i][1]
                if i == idx:
                    b.setStyleSheet(
                        f'QPushButton{{background:{ac};color:#FFF;border-radius:8px;'
                        f'border:none;font-size:11px;font-weight:700;'
                        f'text-align:left;padding:8px 12px;}}')
                else:
                    b.setStyleSheet(
                        f'QPushButton{{background:#f1f5f9;color:{TEXT};border-radius:8px;'
                        f'border:none;font-size:11px;font-weight:600;'
                        f'text-align:left;padding:8px 12px;}}'
                        f'QPushButton:hover{{background:#e2e8f0;}}')

        for idx, (label, accent, row_bg, data, headers, row_fn, key_fn) in enumerate(CATS):
            is_tum = (data is None)
            count  = (len(es_all)+len(tf_all)+len(if_all)+len(mf_all)) if is_tum else len(data)

            btn = QPushButton(f'{label}\n{count} kayıt')
            btn.setCursor(Qt.PointingHandCursor); btn.setFixedHeight(52)
            btn.clicked.connect(lambda _=False, i=idx: _select_cat(i))
            sb_lay.addWidget(btn); cat_btns.append(btn)

            # Placeholder widget — içerik ilk tıklamada doldurulur
            ph = QWidget()
            stack.addWidget(ph)
            _cat_widgets[idx] = ph

        sb_lay.addStretch()
        body.addWidget(sidebar)
        body.addWidget(stack, 1)
        dl.addLayout(body, 1)

        _select_cat(0)

        btn_kapat = QPushButton('Kapat')
        btn_kapat.setFixedSize(100, 36); btn_kapat.setCursor(Qt.PointingHandCursor)
        btn_kapat.setStyleSheet(
            f'QPushButton{{background:{NAVY};color:#FFF;border-radius:8px;'
            f'border:none;font-size:11px;font-weight:700;}}'
            f'QPushButton:hover{{background:{NAVY2};}}')
        btn_kapat.clicked.connect(dlg.reject)
        brow = QHBoxLayout(); brow.addStretch(); brow.addWidget(btn_kapat)
        dl.addLayout(brow)
        dlg.exec()

    # ─────────────────────────────────────────────────────────────────────────
    def _on_duzelt(self):
        if not self._last_result:
            return
        result   = self._last_result
        ikdvl_f  = result.get('ikdvl_fazla', [])
        muavin_f = result.get('muavin_fazla', [])
        tutar_f  = result.get('tutar_farki', [])

        if not ikdvl_f and not muavin_f and not tutar_f:
            QMessageBox.information(self, 'Mükemmel!', 'Hiçbir fark bulunamadı. Tüm kayıtlar eşleşiyor.')
            return

        # ── Dialog ────────────────────────────────────────────────────────────
        dlg = QDialog(self)
        dlg.setWindowTitle('FARKLARI GÖRÜNTÜLE')
        dlg.resize(900, 580)
        dlg.setStyleSheet(f'QDialog{{background:{BG};}}')
        dl = QVBoxLayout(dlg)
        dl.setContentsMargins(16, 12, 16, 14)
        dl.setSpacing(8)

        # Başlık
        ttl = QLabel('⚠️  FARK DETAYLARI — Düzeltmek için aşağıdaki butonları kullanın')
        ttl.setStyleSheet(f'font-size:12px;font-weight:800;color:{NAVY};background:transparent;')
        dl.addWidget(ttl)

        # ── Arama kutusu ──────────────────────────────────────────────────────
        srch_row = QHBoxLayout()
        srch_row.setSpacing(6)
        lbl_srch = QLabel('🔍')
        lbl_srch.setStyleSheet('font-size:14px;background:transparent;')
        srch_box = QLineEdit()
        srch_box.setPlaceholderText('Fatura no, satıcı veya tutar ile ara...')
        srch_box.setFixedHeight(32)
        srch_box.setStyleSheet(f'''
            QLineEdit{{background:{CARD};border:1.5px solid {BORDER};border-radius:8px;
            padding:0 10px;font-size:11px;color:{TEXT};}}
            QLineEdit:focus{{border-color:{NAVY};}}
        ''')
        srch_row.addWidget(lbl_srch)
        srch_row.addWidget(srch_box, 1)
        dl.addLayout(srch_row)

        # ── Tablar ────────────────────────────────────────────────────────────
        tabs = QTabWidget()
        tabs.setStyleSheet(f'''
            QTabWidget::pane {{border:1px solid {BORDER};border-radius:10px;background:{CARD};}}
            QTabBar::tab {{
                background:#F3F4F6;color:{TEXT2};border-radius:6px 6px 0 0;
                padding:6px 14px;font-size:11px;font-weight:600;margin-right:2px;
            }}
            QTabBar::tab:selected {{background:{NAVY};color:#FFF;}}
        ''')

        MAX_ROWS = 2000   # UI donmasını önlemek için üst sınır
        _all_tables: list[QTableWidget] = []

        def make_tab(title, all_rows, headers):
            rows = all_rows[:MAX_ROWS]
            w  = QWidget()
            wl = QVBoxLayout(w)
            wl.setContentsMargins(4, 4, 4, 4)
            wl.setSpacing(0)

            if len(all_rows) > MAX_ROWS:
                warn = QLabel(f'  İlk {MAX_ROWS} kayıt gösteriliyor (toplam {len(all_rows)})')
                warn.setStyleSheet(f'font-size:10px;color:{TEXT2};background:transparent;')
                wl.addWidget(warn)

            tbl = QTableWidget()
            tbl.setColumnCount(len(headers))
            tbl.setRowCount(len(rows))
            tbl.setHorizontalHeaderLabels(headers)
            tbl.setStyleSheet(f'''
                QTableWidget{{
                    background:{CARD};border:none;font-size:11px;color:{TEXT};
                    gridline-color:#E5E7EB;alternate-background-color:#F9FAFB;
                }}
                QTableWidget::item{{color:{TEXT};padding:2px 6px;}}
                QTableWidget::item:selected{{background:{NAVY};color:#FFF;}}
                QHeaderView::section{{background:{NAVY};color:#FFF;font-weight:700;
                                      font-size:10px;padding:5px 6px;border:none;}}
            ''')
            tbl.setAlternatingRowColors(True)
            tbl.horizontalHeader().setStretchLastSection(True)
            tbl.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeToContents)
            tbl.verticalHeader().setVisible(False)
            tbl.setEditTriggers(QAbstractItemView.NoEditTriggers)
            tbl.setSelectionBehavior(QAbstractItemView.SelectRows)
            tbl.setSelectionMode(QAbstractItemView.SingleSelection)
            tbl.setSortingEnabled(True)

            for ri, row_data in enumerate(rows):
                tbl.setRowHeight(ri, 22)
                for ci, val in enumerate(row_data):
                    item = QTableWidgetItem(str(val))
                    item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                    tbl.setItem(ri, ci, item)

            wl.addWidget(tbl)
            tabs.addTab(w, title)
            _all_tables.append(tbl)

        # Veri hazırla
        tutar_rows = [
            [p['ikdvl']['fatura_no'],
             p['ikdvl']['tarih'],
             p['ikdvl']['satici'][:40],
             f"{float(p['ikdvl']['toplam_kdv']):.2f}",
             f"{float(p.get('muavin_tutar', p['muavin'].get('toplam_kdv', 0))):.2f}",
             f"{float(p['fark']):.2f}",
             str(len(p.get('muavin_rows', [p['muavin']])))]
            for p in tutar_f
        ]
        ikdvl_rows_data = [
            [r['fatura_no'], r['tarih'], r['satici'][:40],
             r['vkn'], f"{float(r['toplam_kdv']):.2f}"]
            for r in ikdvl_f
        ]
        muavin_rows_data = []
        for r in muavin_f:
            try:
                kdv = f"{float(r['toplam_kdv']):.2f}"
            except Exception:
                kdv = str(r.get('toplam_kdv', ''))
            muavin_rows_data.append([
                r.get('fatura_no', ''),
                r.get('tarih', ''),
                str(r.get('firma', ''))[:40],
                kdv,
            ])

        if tutar_f:
            make_tab(f'🟡  Tutar Farkı  ({len(tutar_f)})', tutar_rows,
                     ['Fatura No', 'Tarih', 'Satıcı', 'IKDVL KDV', 'Muavin KDV', 'Fark', 'Sat.'])
        if ikdvl_f:
            make_tab(f'🔴  IKDVL Fazlası  ({len(ikdvl_f)})', ikdvl_rows_data,
                     ['Fatura No', 'Tarih', 'Satıcı', 'VKN', 'KDV'])
        if muavin_f:
            make_tab(f'🟠  Muavin Fazlası  ({len(muavin_f)})', muavin_rows_data,
                     ['Fatura No', 'Tarih', 'Satıcı/Açıklama', 'KDV'])

        dl.addWidget(tabs, 1)

        # ── Arama filtresi ────────────────────────────────────────────────────
        def _tr_norm(s: str) -> str:
            """Türkçe karakterleri normalize edip büyük harfe çevirir."""
            return s.upper().translate(str.maketrans('İIĞÜŞÖÇığüşöç', 'IIĞÜŞÖÇIĞÜŞÖÇ')
                                       ).translate(str.maketrans('İ', 'I'))

        def _filter(text):
            txt = _tr_norm(text.strip())
            idx = tabs.currentIndex()
            if idx < 0 or idx >= len(_all_tables):
                return
            tbl = _all_tables[idx]
            for r in range(tbl.rowCount()):
                match = not txt
                if not match:
                    for c in range(tbl.columnCount()):
                        item = tbl.item(r, c)
                        if item and txt in _tr_norm(item.text()):
                            match = True
                            break
                tbl.setRowHidden(r, not match)

        srch_box.textChanged.connect(_filter)
        tabs.currentChanged.connect(lambda _: _filter(srch_box.text()))

        # ── Alt buton satırı ──────────────────────────────────────────────────
        btn_row = QHBoxLayout()
        btn_row.setSpacing(10)

        btn_mevcut_k = QPushButton('📝  Mevcut IKDVL\'yi Düzelt')
        btn_yeni_k   = QPushButton('✨  Yeni Düzenlenmiş Liste Al')
        for _b, _bg, _hv in [
            (btn_mevcut_k, NAVY,      NAVY2),
            (btn_yeni_k,   '#22C55E', '#16A34A'),
        ]:
            _b.setFixedHeight(42)
            _b.setCursor(Qt.PointingHandCursor)
            _b.setStyleSheet(f'''
                QPushButton{{background:{_bg};color:#FFF;border-radius:10px;
                border:none;font-size:12px;font-weight:700;}}
                QPushButton:hover{{background:{_hv};}}
            ''')

        btn_kapat = QPushButton('Kapat')
        btn_kapat.setFixedHeight(42)
        btn_kapat.setFixedWidth(100)
        btn_kapat.setCursor(Qt.PointingHandCursor)
        btn_kapat.setStyleSheet(f'''
            QPushButton{{background:#6B7280;color:#FFF;border-radius:10px;
            border:none;font-size:12px;font-weight:700;}}
            QPushButton:hover{{background:#4B5563;}}
        ''')
        btn_kapat.clicked.connect(dlg.reject)

        btn_row.addWidget(btn_mevcut_k, 1)
        btn_row.addWidget(btn_yeni_k, 1)
        btn_row.addWidget(btn_kapat)
        dl.addLayout(btn_row)

        # ── Düzelt ve Kaydet aksiyonu ─────────────────────────────────────────
        def _do_kaydet(mode):
            ikdvl_path = result.get('ikdvl_path', '')
            if not ikdvl_path or not os.path.isfile(ikdvl_path):
                QMessageBox.warning(dlg, 'Hata', 'IKDVL dosya yolu bulunamadı.')
                return

            if mode == 'mevcut':
                # Mevcut dosyayı yerinde düzelt — kaydet diyaloğu yok
                if ikdvl_path.lower().endswith('.xls'):
                    out_path = ikdvl_path[:-4] + '_duzeltilmis.xlsx'
                else:
                    out_path = ikdvl_path
            else:
                # Yeni liste → kullanıcı konum seçer
                desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
                default = os.path.join(
                    desktop,
                    f'ikdvl_duzeltilmis_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
                )
                out_path, _ = QFileDialog.getSaveFileName(
                    dlg, 'Yeni Düzenlenmiş IKDVL\'yi Kaydet', default, 'Excel Dosyası (*.xlsx)'
                )
                if not out_path:
                    return

            for _b in (btn_mevcut_k, btn_yeni_k):
                _b.setEnabled(False)
            btn_mevcut_k.setText('⏳  Düzeltiliyor...')
            dlg.setCursor(Qt.WaitCursor)

            self._duzelt_worker = DuzeltWorker(ikdvl_path, result, out_path, mode=mode)
            self._duzelt_worker.sig.log.connect(self._on_log)
            self._duzelt_worker.sig.progress.connect(self._on_duzelt_progress)
            self._duzelt_worker.sig.done.connect(lambda p: self._on_duzelt_done(p, dlg))
            self._duzelt_worker.start()

        btn_mevcut_k.clicked.connect(lambda: _do_kaydet('mevcut'))
        btn_yeni_k.clicked.connect(lambda: _do_kaydet('yeni'))
        dlg.exec()

    @Slot(float)
    def _on_duzelt_progress(self, pct):
        self.pbar.setValue(int(pct * 10))
        self.lbl_pct.setText(f'{int(pct)}%')

    def _show_done_popup(self, baslik: str, satirlar: list[tuple], dosya_yolu: str):
        """
        Referans gui tarzında 'İşlem Tamamlandı' popup'u.
        satirlar: [(label, value), ...]
        dosya_yolu: 'Düzenlenmiş Exceli Aç' butonu için açılacak dosya
        """
        dlg = QDialog(self)
        dlg.setWindowTitle(baslik)
        dlg.setFixedWidth(400)
        dlg.setStyleSheet(f'background:{CARD};')
        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(28, 24, 28, 20)
        lay.setSpacing(12)

        # İkon
        ico = QLabel('✅')
        ico.setAlignment(Qt.AlignCenter)
        ico.setStyleSheet('font-size:44px;background:transparent;border:none;')
        lay.addWidget(ico)

        # Başlık
        ttl = QLabel('İŞLEM TAMAMLANDI')
        ttl.setAlignment(Qt.AlignCenter)
        ttl.setStyleSheet(
            f'font-size:15px;font-weight:800;color:{NAVY};background:transparent;border:none;'
        )
        lay.addWidget(ttl)

        # Ayırıcı çizgi
        sep = QFrame()
        sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
        lay.addWidget(sep)

        # Bilgi satırları
        for label, value in satirlar:
            row = QHBoxLayout()
            lbl_k = QLabel(label)
            lbl_k.setStyleSheet(
                f'font-size:11px;color:{TEXT2};background:transparent;border:none;'
            )
            lbl_v = QLabel(value)
            lbl_v.setStyleSheet(
                f'font-size:11px;font-weight:700;color:{TEXT};background:transparent;border:none;'
            )
            row.addWidget(lbl_k)
            row.addStretch()
            row.addWidget(lbl_v)
            lay.addLayout(row)

        # Dosya yolu (kısaltılmış)
        if dosya_yolu:
            fname = os.path.basename(dosya_yolu)
            lbl_f = QLabel(fname)
            lbl_f.setAlignment(Qt.AlignCenter)
            lbl_f.setWordWrap(True)
            lbl_f.setStyleSheet(
                f'font-size:10px;color:{TEXT2};background:#F3F4F6;'
                f'border-radius:6px;padding:4px 8px;border:none;'
            )
            lay.addWidget(lbl_f)

        # Butonlar
        btn_ac = QPushButton('📂  Düzenlenmiş Exceli Aç')
        btn_ac.setFixedHeight(40)
        btn_ac.setCursor(Qt.PointingHandCursor)
        btn_ac.setStyleSheet(f'''
            QPushButton{{background:{GREEN};color:#FFF;border-radius:10px;
            border:none;font-size:12px;font-weight:700;}}
            QPushButton:hover{{background:{GREEN2};}}
        ''')

        btn_kapat = QPushButton('Kapat')
        btn_kapat.setFixedHeight(36)
        btn_kapat.setCursor(Qt.PointingHandCursor)
        btn_kapat.setStyleSheet(f'''
            QPushButton{{background:#F3F4F6;color:{TEXT2};border-radius:10px;
            border:none;font-size:11px;font-weight:600;}}
            QPushButton:hover{{background:{BORDER};}}
        ''')

        def _ac():
            import subprocess
            subprocess.Popen(['start', '', dosya_yolu], shell=True)
            dlg.accept()

        if dosya_yolu:
            btn_ac.clicked.connect(_ac)
        else:
            btn_ac.setVisible(False)

        btn_kapat.clicked.connect(dlg.accept)

        lay.addWidget(btn_ac)
        lay.addWidget(btn_kapat)
        dlg.exec()

    @Slot(str)
    def _on_duzelt_done(self, out_path, dlg):
        dlg.setCursor(Qt.ArrowCursor)
        if out_path:
            result = self._last_result or {}
            tf = len(result.get('tutar_farki',  []))
            mf = len(result.get('muavin_fazla', []))
            if_ = len(result.get('ikdvl_fazla',  []))
            self._log_append(f'Düzeltilmiş dosya hazır: {out_path}', 'ok')
            dlg.accept()
            self._show_done_popup(
                'Düzeltme Tamamlandı',
                [
                    ('Tutar güncellenen',   f'{tf} fatura'),
                    ('Muavinden eklenen',   f'{mf} fatura'),
                    ('Alta taşınan',        f'{if_} fatura'),
                    ('Tamamlanma saati',    datetime.now().strftime('%H:%M:%S')),
                ],
                out_path,
            )
        else:
            self._log_append('Düzeltme başarısız. Log kutusunu kontrol edin.', 'err')

    def _on_report(self):
        if not self._last_result:
            QMessageBox.information(self, 'Bilgi', 'Önce karşılaştırma yapın.')
            return

        desktop = os.path.join(os.path.expanduser('~'), 'Desktop')
        period_part = f'_{self._detected_period}' if self._detected_period else ''
        default = os.path.join(
            desktop, f'Karsilastirma_Raporu{period_part}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        path, _ = QFileDialog.getSaveFileName(
            self, 'Raporu Kaydet', default, 'Excel Dosyası (*.xlsx)'
        )
        if not path:
            return

        try:
            from cc_modules.compare_191 import karsilastir
            karsilastir.create_report(self._last_result, path, cb_log=self._log_append)
            result = self._last_result
            self._show_done_popup(
                'Rapor Tamamlandı',
                [
                    ('Eşleşen fatura',    f'{len(result.get("eslesen", []))}'),
                    ('IKDVL fazlası',      f'{len(result.get("ikdvl_fazla", []))}'),
                    ('Muavin fazlası',     f'{len(result.get("muavin_fazla", []))}'),
                    ('Tutar farkı',        f'{len(result.get("tutar_farki", []))}'),
                    ('Tamamlanma saati',   datetime.now().strftime('%H:%M:%S')),
                ],
                path,
            )
        except Exception as e:
            import traceback
            self._log_append(f'Rapor hatası: {e}', 'err')
            self._log_append(traceback.format_exc(), 'err')

    def _open_activation(self):
        from activation import ActivationWindow
        win = ActivationWindow(parent=self)
        win.exec()
        if win.activated:
            from license import check_license
            valid, msg, expire = check_license()
            if valid:
                self.expire_date   = expire
                self._trial_status = None

    # ── Worker Sinyalleri ─────────────────────────────────────────────────────
    @Slot(str, str)
    def _on_log(self, msg, tag='info'):
        self._log_append(msg, tag)

    @Slot(float)
    def _on_progress(self, pct):
        self.pbar.setValue(int(pct * 10))
        self.lbl_pct.setText(f'{int(pct)}%')
        self.lbl_status.setText(f'İşleniyor...  {int(pct)}%')

    @Slot(int, int, int, int)
    def _on_stats(self, ikdvl_cnt, muavin_cnt, eslesen, fark):
        self._detail_rows['islenen_kayit'].setText(str(ikdvl_cnt))

    @Slot(int, int, int, int)
    def _on_done(self, eslesen, ikdvl_fazla, muavin_fazla, tutar_farki):
        elapsed = time.time() - (self._t0 or time.time())
        m, sec  = divmod(int(elapsed), 60)
        self._running       = False
        self._islem_sure    = elapsed

        # Sonucu al
        if self.worker:
            self._last_result = self.worker.get_result()

        # Tutarları hesapla
        def _sum_kdv(rows):
            return float(sum(r.get('toplam_kdv', 0) for r in rows))

        def _sum_muavin(rows):
            return float(sum(r.get('borc', 0) or r.get('alacak', 0) for r in rows))

        if self._last_result:
            r = self._last_result
            es_rows  = r.get('eslesen',     [])
            ik_rows  = r.get('ikdvl_fazla', [])
            mu_rows  = r.get('muavin_fazla',[])
            tf_rows  = r.get('tutar_farki', [])
            ik_all   = r.get('ikdvl_rows',  [])

            es_tutar  = sum(float(p['ikdvl'].get('toplam_kdv', 0)) for p in es_rows)
            tf_tutar  = sum(float(p.get('fark', 0)) for p in tf_rows)
            ik_tutar  = _sum_kdv(ik_rows)
            mu_tutar  = _sum_muavin(mu_rows)
            top_tutar = _sum_kdv(ik_all)

            self.sc_eslesen.set_values(eslesen,    es_tutar)
            self.sc_tutar.set_values(tutar_farki,  tf_tutar)
            self.sc_ikdvl.set_values(ikdvl_fazla,  ik_tutar)
            self.sc_muavin.set_values(muavin_fazla, mu_tutar)
            self.sc_toplam.set_values(len(ik_all),  top_tutar)

            fark_say = ikdvl_fazla + muavin_fazla + tutar_farki
            self._detail_rows['islenen_kayit'].setText(str(len(ik_all)))
            self._detail_rows['duzeltilecek'].setText(str(fark_say))

            # KDV Özeti güncelle
            muavin_es_top = sum(float(p.get('muavin_tutar', 0)) for p in es_rows)
            muavin_mf_top = _sum_kdv(mu_rows)
            muavin_tf_top = sum(float(p.get('muavin_tutar', 0)) for p in tf_rows)
            muavin_top = muavin_es_top + muavin_mf_top + muavin_tf_top
            toplam_fark = ik_tutar + sum(abs(float(p.get('fark', 0))) for p in tf_rows)
            esleme_pct = (eslesen / len(ik_all) * 100) if ik_all else 0
            self.lbl_kdv_ikdvl.setText(f'{top_tutar:,.2f} ₺')
            self.lbl_kdv_muavin.setText(f'{muavin_top:,.2f} ₺' if muavin_top else '—')
            fark_clr = RED if toplam_fark > 0 else GREEN
            self.lbl_kdv_fark.setStyleSheet(
                f'font-size:15px;font-weight:800;color:{fark_clr};'
                f'background:transparent;border:none;'
            )
            self.lbl_kdv_fark.setText(f'{toplam_fark:,.2f} ₺' if toplam_fark else '0,00 ₺')
            self.donut.set_pct(esleme_pct)
            # Detected period for report filename
            months = self._ikdvl_months or self._muavin_months
            if months:
                yil, ay = sorted(months)[0]
                self._detected_period = f'{self._AYLAR_TR_DICT.get(ay, str(ay))}_{yil}'
            else:
                self._detected_period = ''
        else:
            self.sc_eslesen.set_values(eslesen)
            self.sc_tutar.set_values(tutar_farki)
            self.sc_ikdvl.set_values(ikdvl_fazla)
            self.sc_muavin.set_values(muavin_fazla)

        # UI sıfırla
        self.btn_start.setText('▶   KARŞILAŞTIRMAYI BAŞLAT\nDosyaları karşılaştır')
        self._style_start_idle()
        self.btn_start.setEnabled(True)

        if self._last_result:
            self.btn_report.setEnabled(True)
            self.btn_duzelt.setEnabled(True)
            self.btn_detay.setEnabled(True)

        # Deneme kullanımı — muavin satır sayısı sayılır
        if self._trial_status and (eslesen + ikdvl_fazla + muavin_fazla) > 0:
            try:
                from license import add_trial_usage, get_trial_status
                muavin_row_count = len(
                    self._last_result.get('muavin_rows', [])
                ) if self._last_result else 1
                add_trial_usage(max(1, muavin_row_count))
                _, kalan_gun, islenen2, kalan2 = get_trial_status()
                kota2 = islenen2 + kalan2
                if hasattr(self, 'lbl_trial'):
                    self.lbl_trial.setText(
                        f'Deneme Sürümü  —  {kalan_gun} gün kaldı  |  '
                        f'{islenen2} / {kota2} muavin satırı kullanıldı'
                    )
                # Kota dolmuşsa uyar
                if kalan2 <= 0:
                    from PySide6.QtWidgets import QMessageBox
                    QMessageBox.warning(self, 'Deneme Kotası Doldu',
                        'Deneme muavin satır limitine ulaşıldı.\n'
                        'Devam etmek için lisans satın alın.')
            except Exception:
                pass

        self._spin_timer.stop()
        self.lbl_spinner.setText('')
        self.pbar.setValue(1000)
        self.lbl_pct.setText('100%')
        self.lbl_status.setText('Tamamlandı')
        self.lbl_sys.setText('● Tamamlandı')
        self.lbl_sys.setStyleSheet(
            f'font-size:10px;color:{GREEN};background:transparent;border:none;'
        )

        # İşlem bilgileri güncelle
        self._detail_rows['kars_sure'].setText(f'{m:02d}dk {sec:02d}sn')
        fark_say2 = ikdvl_fazla + muavin_fazla + tutar_farki
        self._detail_rows['durum'].setText('✅ Tamamlandı' if fark_say2 == 0 else f'⚠️ {fark_say2} fark bulundu')

        # Tamamlanma popup
        dlg = QDialog(self)
        dlg.setWindowTitle('Karşılaştırma Tamamlandı')
        dlg.setFixedWidth(380)
        dlg.setStyleSheet(f'background:{CARD};')
        d_lay = QVBoxLayout(dlg)
        d_lay.setContentsMargins(28, 24, 28, 20)
        d_lay.setSpacing(10)

        fark_toplam = ikdvl_fazla + muavin_fazla + tutar_farki
        ico_lbl = QLabel('✅' if fark_toplam == 0 else '⚠️')
        ico_lbl.setAlignment(Qt.AlignCenter)
        ico_lbl.setStyleSheet('font-size:42px;background:transparent;border:none;')
        d_lay.addWidget(ico_lbl)

        ttl_lbl = QLabel('KARŞILAŞTIRMA TAMAMLANDI')
        ttl_lbl.setAlignment(Qt.AlignCenter)
        ttl_lbl.setStyleSheet(
            f'font-size:14px;font-weight:800;color:{NAVY};background:transparent;border:none;'
        )
        d_lay.addWidget(ttl_lbl)

        sep = QFrame(); sep.setFrameShape(QFrame.HLine)
        sep.setStyleSheet(f'background:{BORDER};border:none;max-height:1px;')
        d_lay.addWidget(sep)

        for label, value, clr in [
            ('Eşleşen Fatura',       f'{eslesen} adet',       GREEN),
            ('Tutar Farkı Olan',     f'{tutar_farki} adet',   '#D97706' if tutar_farki else TEXT2),
            ('Sadece IKDVL\'de',     f'{ikdvl_fazla} adet',   RED if ikdvl_fazla else TEXT2),
            ('Sadece Muavinde',      f'{muavin_fazla} adet',  ORANGE if muavin_fazla else TEXT2),
            ('Toplam Süre',          f'{m:02d}dk {sec:02d}sn',TEXT),
        ]:
            row = QHBoxLayout()
            lk = QLabel(label)
            lk.setStyleSheet(f'font-size:11px;color:{TEXT2};background:transparent;border:none;')
            lv = QLabel(value)
            lv.setStyleSheet(f'font-size:11px;font-weight:700;color:{clr};background:transparent;border:none;')
            row.addWidget(lk); row.addStretch(); row.addWidget(lv)
            d_lay.addLayout(row)

        btn_ok = QPushButton('Tamam')
        btn_ok.setFixedHeight(40)
        btn_ok.setCursor(Qt.PointingHandCursor)
        btn_ok.setStyleSheet(f'''
            QPushButton{{background:{GREEN};color:#FFF;border-radius:10px;
            border:none;font-size:13px;font-weight:700;}}
            QPushButton:hover{{background:{GREEN2};}}
        ''')
        btn_ok.clicked.connect(dlg.accept)
        d_lay.addWidget(btn_ok)
        dlg.exec()

        self._tray.showMessage(
            '✅ Tamamlandı' if fark_toplam == 0 else '⚠️ Farklılıklar Var',
            f'Eşleşen: {eslesen}  Fark: {fark_toplam}',
            QSystemTrayIcon.Information, 4000
        )

    # ── Log ───────────────────────────────────────────────────────────────────
    LOG_CLR = {
        'ok':   '#4DCC78',
        'err':  '#FF5C5C',
        'warn': '#FFB84D',
        'info': '#7EB3D8',
        'head': '#C9A84C',
    }

    def _log_append(self, msg, tag='info'):
        c    = self.LOG_CLR.get(tag, '#CBD5E1')
        ts   = datetime.now().strftime('%H:%M:%S')
        html = (
            f'<span style="color:#475569;">[{ts}]</span>'
            f'&nbsp;&nbsp;<span style="color:{c};">{msg}</span>'
        )
        if getattr(self, '_cc_paused', False):
            self._cc_log_buffer.append(html)
            return
        self.log_box.append(html)
        self.log_box.ensureCursorVisible()

    # ── ContraCore lifecycle hooks ────────────────────────────────────────────

    def on_module_activated(self):
        """Shell tarafından modül görünür hale geldiğinde çağrılır."""
        self._cc_paused = False
        for html in self._cc_log_buffer:
            self.log_box.append(html)
        if self._cc_log_buffer:
            self.log_box.ensureCursorVisible()
        self._cc_log_buffer.clear()
        if self._cc_spin_was_active and not self._spin_timer.isActive():
            self._spin_timer.start()

    def on_module_deactivated(self):
        """Shell tarafından modül gizlendiğinde çağrılır."""
        self._cc_paused          = True
        self._cc_spin_was_active = self._spin_timer.isActive()
        if self._spin_timer.isActive():
            self._spin_timer.stop()

    def _save_log(self):
        content = self.log_box.toPlainText().strip()
        if not content:
            QMessageBox.information(self, 'Bilgi', 'İşlem günlüğü boş.')
            return
        default = f'Gunluk_{datetime.now().strftime("%Y%m%d_%H%M%S")}.txt'
        path, _ = QFileDialog.getSaveFileName(
            self, 'Günlüğü Kaydet',
            os.path.join(os.path.expanduser('~'), 'Desktop', default),
            'Metin Dosyası (*.txt)'
        )
        if not path:
            return
        with open(path, 'w', encoding='utf-8') as f:
            f.write(f'191 Muavin Karşılaştırma — İşlem Günlüğü\n')
            f.write(f'{datetime.now().strftime("%d.%m.%Y %H:%M")}\n')
            f.write('=' * 60 + '\n\n')
            f.write(content)
        self._log_append(f'✓ Günlük kaydedildi: {path}', 'ok')

    # ── Spinner ───────────────────────────────────────────────────────────────
    def _spin_tick(self):
        self.lbl_spinner.setText(self._spin_frames[self._spin_idx % 4])
        self._spin_idx += 1


# ── Giriş ─────────────────────────────────────────────────────────────────────
def main():
    app = QApplication(sys.argv)
    try:
        app.setAttribute(Qt.AA_EnableHighDpiScaling, True)
        app.setAttribute(Qt.AA_UseHighDpiPixmaps,    True)
    except AttributeError:
        pass
    app.setFont(QFont('Segoe UI', 10))

    from license import check_license, get_trial_status
    valid, msg, expire = check_license()
    trial_status = None

    if not valid:
        aktif, kalan_gun, islenen, kalan = get_trial_status()
        if aktif:
            trial_status = (kalan_gun, islenen, kalan)
        else:
            trial_expired = not aktif and islenen > 0
            from activation import ActivationWindow
            win = ActivationWindow(
                expired=trial_expired or ('doldu' in msg.lower()),
                expire_msg=msg if trial_expired else ''
            )
            win.show()
            app.exec()
            if win.activated:
                valid, msg, expire = check_license()
                if not valid:
                    sys.exit(0)
            elif win.trial_started:
                aktif2, kalan_gun2, islenen2, kalan2 = get_trial_status()
                trial_status = (kalan_gun2, islenen2, kalan2)
            else:
                sys.exit(0)

    ico_file = None
    for f in ('muavin.ico', '191.png', 'logo.png'):
        _pix = ip(f)
        if not _pix.isNull():
            ico_file = _pix
            break

    w = MainWindow(expire_date=expire if valid else None, trial_status=trial_status)
    if ico_file:
        w.setWindowIcon(QIcon(ico_file))
    w.show()
    sys.exit(app.exec())


if __name__ == '__main__':
    main()


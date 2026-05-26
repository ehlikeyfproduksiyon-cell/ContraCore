"""
191 Muavin Karşılaştırma — Çekirdek Mantık v3 (Universal Reader)
Developed by Serkan ŞAHİN © 2026

Desteklenen 191 muavin formatları:
  - Ayrı EVRAK NO sütunlu (191_08, 191 12, ADENİS)
  - Açıklamada birleşik (muavin_cinaralti, Ak Zafer, 191 MUAVİN)
  - .xls ve .xlsx dosyaları
  - Çoklu KDV oranı bölümleri (%1, %10, %20 ayrı tablolar)
"""

import re
from datetime import datetime
from decimal import Decimal, InvalidOperation
from collections import defaultdict


# ── E-fatura deseni ───────────────────────────────────────────────────────────
# Türk e-fatura: 3 alfanümerik prefix + 4 rakam yıl (20xx) + en az 7 rakam sıra
# Örnek: BRJ2025000000045, O062025000000048, 0012025000136437
_EFATURA_RE = re.compile(r'(?<!\w)([A-Z0-9]{3}20[2-9][0-9][0-9]{7,12})(?!\w)')

_AYLAR = {
    'OCAK': 1, 'SUBAT': 2, 'MART': 3, 'NISAN': 4, 'MAYIS': 5,
    'HAZIRAN': 6, 'TEMMUZ': 7, 'AGUSTOS': 8, 'EYLUL': 9,
    'EKIM': 10, 'KASIM': 11, 'ARALIK': 12,
}

_SKIP_TEXT = {
    'TOPLAM', 'ARA TOPLAM', 'GENEL TOPLAM', 'ALT HESAP TOPLAMI',
    'NAKLI YEKUN', 'NAKLİ YEKÜN', 'NAKLI YEKUN', 'DEVIR', 'DEVİR', '_DEVIR_',
    'KDV TAHAKKUK', 'TAHAKKUK', 'MUHASEBE FISI', 'ALT HESAP',
}


# ══════════════════════════════════════════════════════════════════════════════
# YARDIMCI FONKSİYONLAR
# ══════════════════════════════════════════════════════════════════════════════

def _to_decimal(val) -> Decimal:
    if val is None:
        return Decimal('0')
    s = str(val).replace(',', '.').replace('\xa0', '').replace(' ', '').replace('TL', '').strip()
    if not s or s in ('.', '-', '−', '0.0'):
        return Decimal('0')
    try:
        return Decimal(s)
    except InvalidOperation:
        return Decimal('0')


def _normalize_fatura_no(s) -> str:
    if not s:
        return ''
    return re.sub(r'[\s\-]+', '', str(s)).upper().strip()


def _tr_upper(s: str) -> str:
    """Türkçe karakterleri büyütür ve ASCII'ye çevirir."""
    s = str(s or '').upper()
    for old, new in [('İ','I'),('Ğ','G'),('Ş','S'),('Ü','U'),('Ö','O'),('Ç','C')]:
        s = s.replace(old, new)
    return s


def _norm_header(s: str) -> str:
    """Header değerini boşluk kaldırarak normalize et."""
    return re.sub(r'\s+', '', _tr_upper(s))


def _parse_tarih(val):
    """
    Returns (tarih_str 'DD.MM.YYYY', ay int, yil int) veya ('', None, None).
    """
    if val is None:
        return '', None, None
    if isinstance(val, datetime):
        return val.strftime('%d.%m.%Y'), val.month, val.year
    s = str(val).strip()
    m = re.search(r'(\d{2})[./](\d{2})[./](\d{4})', s)
    if m:
        return m.group(0).replace('/', '.'), int(m.group(2)), int(m.group(3))
    m = re.search(r'(\d{4})-(\d{2})-(\d{2})', s)
    if m:
        return f'{m.group(3)}.{m.group(2)}.{m.group(1)}', int(m.group(2)), int(m.group(1))
    return '', None, None


def _find_einvoice(text) -> str:
    """Metinde ilk e-fatura numarasını bul."""
    m = _EFATURA_RE.search(_tr_upper(str(text or '')))
    return m.group(1) if m else ''


def _is_skip_row(row_values) -> bool:
    """Toplam/devir/açılış satırı mı?"""
    text = _tr_upper(' '.join(str(v or '') for v in row_values if v is not None))
    for kw in _SKIP_TEXT:
        if kw in text:
            return True
    return False


def _kdv_section_rate(row_values):
    """
    KDV bölüm başlığı satırıysa rate string döner ('1','10','20'), yoksa None.
    Örn: '191 00 01 - İNDİRİLECEK KDV % 1'  →  '1'
         'Alt Hesap Kodu : 191.10 - % 10'     →  '10'
    """
    text = _tr_upper(' '.join(str(v or '') for v in row_values if v is not None))
    m = re.search(r'%\s*(\d+)', text)
    if m and ('KDV' in text or 'INDIRILECEK' in text or '191' in text):
        return m.group(1)
    if re.search(r'\b191[\s.][0-9]', text) and 'KDV' in text:
        return None   # section header ama rate bilinmiyor
    return False      # False = section header değil


# ══════════════════════════════════════════════════════════════════════════════
# .XLS DESTEĞI (xlrd wrapper)
# ══════════════════════════════════════════════════════════════════════════════

class _XlsSheet:
    """xlrd sheet'ini openpyxl iter_rows arayüzüyle sarmalar."""
    def __init__(self, xls_sheet, name):
        self._s    = xls_sheet
        self.title = name

    def iter_rows(self, values_only=True, min_row=1, max_row=None):
        end = (max_row or self._s.nrows)
        for r in range(min_row - 1, min(end, self._s.nrows)):
            row = []
            for c in range(self._s.ncols):
                cell = self._s.cell(r, c)
                ctype = cell.ctype
                if ctype == 0 or ctype == 6:    # empty/blank
                    row.append(None)
                elif ctype == 1:                 # text
                    row.append(cell.value if cell.value.strip() else None)
                elif ctype == 2:                 # number
                    v = cell.value
                    row.append(int(v) if v == int(v) else v)
                elif ctype == 3:                 # date
                    import xlrd
                    try:
                        t = xlrd.xldate_as_tuple(cell.value, self._s.book.datemode)
                        row.append(datetime(*t))
                    except Exception:
                        row.append(cell.value)
                else:
                    row.append(cell.value)
            yield tuple(row)


class _XlsWb:
    def __init__(self, path):
        import xlrd
        self._wb      = xlrd.open_workbook(path)
        self.sheetnames = self._wb.sheet_names()

    def __getitem__(self, name):
        return _XlsSheet(self._wb.sheet_by_name(name), name)

    @property
    def worksheets(self):
        return [_XlsSheet(self._wb.sheet_by_name(n), n) for n in self.sheetnames]


def _open_wb(path):
    """Hem .xlsx hem .xls açar (sadece okuma)."""
    if str(path).lower().endswith('.xls'):
        return _XlsWb(path)
    import openpyxl
    return openpyxl.load_workbook(path, data_only=True)


def _xls_colour_to_hex(colour_map, idx):
    """xlrd renk indeksini 'RRGGBB' hex'e çevirir. Bulunamazsa None döner."""
    if idx is None or idx > 63:
        return None
    rgb = colour_map.get(idx)
    if rgb:
        return '{:02X}{:02X}{:02X}'.format(*rgb)
    # Standart Excel renk paleti (yedek)
    _PALETTE = {
        0:'000000',1:'FFFFFF',2:'FF0000',3:'00FF00',4:'0000FF',
        5:'FFFF00',6:'FF00FF',7:'00FFFF',8:'800000',9:'008000',
        10:'000080',11:'808000',12:'800080',13:'008080',14:'C0C0C0',
        15:'808080',16:'9999FF',17:'993366',18:'FFFFCC',19:'CCFFFF',
        20:'660066',21:'FF8080',22:'0066CC',23:'CCCCFF',24:'000080',
        25:'FF00FF',26:'FFFF00',27:'00FFFF',28:'800080',29:'800000',
        30:'008080',31:'0000FF',32:'00CCFF',33:'CCFFFF',34:'CCFFCC',
        35:'FFFF99',36:'99CCFF',37:'FF99CC',38:'CC99FF',39:'FFCC99',
        40:'3366FF',41:'33CCCC',42:'99CC00',43:'FFCC00',44:'FF9900',
        45:'FF6600',46:'666699',47:'969696',48:'003366',49:'339966',
        50:'003300',51:'333300',52:'993300',53:'993366',54:'333399',55:'333333',
        63:'FFFFFF',
    }
    return _PALETTE.get(idx)


_XLS_HALIGN = {1:'left', 2:'center', 3:'right', 4:'fill',
               5:'justify', 6:'centerContinuous', 7:'distributed'}
_XLS_VALIGN = {0:'top', 1:'center', 2:'bottom', 3:'justify', 4:'distributed'}


def _xls_to_openpyxl_wb(xls_path):
    """
    .xls dosyasını xlrd ile okur, openpyxl Workbook olarak döner.
    Font, renk, arka plan, hizalama, satır yüksekliği, sütun genişliği kopyalanır.
    """
    import xlrd, openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin   = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    xls_wb     = xlrd.open_workbook(xls_path, formatting_info=True)
    colour_map = xls_wb.colour_map
    xlsx_wb    = openpyxl.Workbook()
    xlsx_wb.remove(xlsx_wb.active)

    for sh_name in xls_wb.sheet_names():
        xls_sh  = xls_wb.sheet_by_name(sh_name)
        xlsx_ws = xlsx_wb.create_sheet(title=sh_name)

        for r in range(xls_sh.nrows):
            for c in range(xls_sh.ncols):
                cell  = xls_sh.cell(r, c)
                ctype = cell.ctype

                # ── Değer ─────────────────────────────────────────────────
                if ctype == 0 or ctype == 6:
                    val = None
                elif ctype == 1:
                    val = cell.value if cell.value.strip() else None
                elif ctype == 2:
                    v   = cell.value
                    val = int(v) if v == int(v) else v
                elif ctype == 3:
                    try:
                        t   = xlrd.xldate_as_tuple(cell.value, xls_wb.datemode)
                        val = datetime(*t).strftime('%d.%m.%Y')
                    except Exception:
                        val = cell.value
                else:
                    val = cell.value

                oc = xlsx_ws.cell(row=r + 1, column=c + 1, value=val)

                # ── Biçimlendirme ─────────────────────────────────────────
                try:
                    xf_idx = xls_sh.cell_xf_index(r, c)
                    xf     = xls_wb.xf_list[xf_idx]
                    xfont  = xls_wb.font_list[xf.font_index]

                    # Font
                    font_colour = _xls_colour_to_hex(colour_map, xfont.colour_index)
                    oc.font = Font(
                        name   = xfont.name or 'Calibri',
                        size   = xfont.height / 20,
                        bold   = bool(xfont.bold),
                        italic = bool(xfont.italic),
                        color  = font_colour if font_colour else '000000',
                    )

                    # Arka plan dolgusu
                    bg = xf.background
                    if bg.fill_pattern == 1:   # solid fill
                        fg_hex = _xls_colour_to_hex(colour_map, bg.pattern_colour_index)
                        if fg_hex and fg_hex not in ('FFFFFF', 'FFFF00FF'):
                            oc.fill = PatternFill('solid', fgColor=fg_hex)

                    # Hizalama
                    aln = xf.alignment
                    oc.alignment = Alignment(
                        horizontal = _XLS_HALIGN.get(aln.hor_align, 'general'),
                        vertical   = _XLS_VALIGN.get(aln.vert_align, 'bottom'),
                        wrap_text  = bool(aln.text_wrapped),
                    )
                except Exception:
                    pass  # biçim okunamazsa varsayılan bırak

                if val is not None:
                    oc.border = border

        # Satır yükseklikleri
        default_h = (xls_sh.default_row_height / 20) if xls_sh.default_row_height else 15
        for r in range(xls_sh.nrows):
            if r in xls_sh.rowinfo_map:
                h = xls_sh.rowinfo_map[r].height / 20
                xlsx_ws.row_dimensions[r + 1].height = h if h > 0 else default_h
            else:
                xlsx_ws.row_dimensions[r + 1].height = default_h

        # Sütun genişlikleri
        for c in range(xls_sh.ncols):
            if c in xls_sh.colinfo_map:
                w = xls_sh.colinfo_map[c].width / 256
                xlsx_ws.column_dimensions[get_column_letter(c + 1)].width = max(w, 4) if w > 0 else 10

    return xlsx_wb


# ══════════════════════════════════════════════════════════════════════════════
# MUAVIN SÜTUN TESPİTİ
# ══════════════════════════════════════════════════════════════════════════════

# Sütun adı anahtar kelimeleri (boşluksuz, TR normalleştirilmiş)
_HDR_BORC       = {'BORC','BORCTL','ANABOYCBORÇ','ANADOVIZBORC','BORCTUTARI','TUTAR',
                   'ANADOVIZBORC','BORCBALANCE','ANADÖVIZBORÇ'}
_HDR_TARIH      = {'TARIH','FISTTARIHI','TARIHI'}
_HDR_EVRAK_TAR  = {'EVR.TAR.','EVRAKTARIHI','EVRAKTAR.','BELGETARIHI','BELGRETARIHI',
                   'FATURAARIHI','EVRAKSAATI','EVRAKTARIH'}
_HDR_EVRAK      = {'EVRAKNO','SERINO','FISNO','BELGENO','FATURAONO','BELGE'}
_HDR_ACK        = {'ACIKLAMA','AÇIKLAMA','ACIKLAMAEVRAKNO','DESCRIPTION'}
_HDR_ALACAK     = {'ALACAK','ALACAKTL','ANADOVIZALACAK','ALACAKBALANCE'}
_HDR_VKN        = {'VKN','VERGINO','VERGINOVELTATC','TCKN','TCNO','VERGIKIMLIK'}
_HDR_FIRMA      = {'UNVAN','CARIUNVAN','FIRMAADI','CARIAD','SATICIADI'}


def _detect_header(sheet) -> tuple[int, dict]:
    """
    İlk 25 satırı tarar, header satırını ve sütun haritasını bulur.
    Returns: (header_row_1indexed, col_map)
    col_map = {'borc': ci, 'tarih': ci, 'evrak_tarihi': ci, 'evrak_no': ci, ...}
    """
    GROUPS = [
        (_HDR_BORC,      'borc'),
        (_HDR_TARIH,     'tarih'),
        (_HDR_EVRAK_TAR, 'evrak_tarihi'),
        (_HDR_EVRAK,     'evrak_no'),
        (_HDR_ACK,       'aciklama'),
        (_HDR_ALACAK,    'alacak'),
        (_HDR_VKN,       'vkn'),
        (_HDR_FIRMA,     'firma'),
    ]

    best_row   = 4
    best_score = 0
    best_map   = {}

    for ri, row in enumerate(sheet.iter_rows(values_only=True, min_row=1, max_row=25), 1):
        col_map = {}
        score   = 0
        for ci, val in enumerate(row, 1):
            if val is None:
                continue
            v = _norm_header(val)
            for hdrs, field in GROUPS:
                for hdr in hdrs:
                    if v == hdr or (len(hdr) >= 4 and hdr in v):
                        if field not in col_map:
                            col_map[field] = ci
                            score += 1
                        break
        if score > best_score:
            best_score = score
            best_row   = ri
            best_map   = col_map

    return best_row, best_map


# ══════════════════════════════════════════════════════════════════════════════
# MUAVIN OKUYUCU
# ══════════════════════════════════════════════════════════════════════════════

def _firma_from_aciklama(aciklama: str, fatura_no: str, tarih: str) -> str:
    """Açıklama alanından firma adını ayıkla."""
    text = aciklama.strip()

    # Format: 'Al.fat. : CODE/DATE/NUM/FIRMA[/...]'
    m = re.match(
        r'^(?:Al\.fat\.|Sat\.fat\.|[A-Za-z.]+\s*:)\s*[A-Z0-9]+/[^/]+/[^/]*/(.+?)/?$',
        text, re.IGNORECASE
    )
    if m:
        return m.group(1).strip(' /')

    # Format: 'DATE-CODE-FIRMA'
    m = re.match(r'^\d{2}[./]\d{2}[./]\d{4}\s*[-–]\s*[A-Z0-9]+\s*[-–]\s*(.+)$',
                 text, re.IGNORECASE)
    if m:
        return m.group(1).strip()

    # Format: 'DATE SPACE_CODE FIRMA'
    m = re.match(r'^\d{2}[./]\d{2}[./]\d{4}\s+(.+)$', text)
    if m:
        rest = m.group(1)
        # Remove internal account code like 'A.82'
        rest = re.sub(r'^[A-Z]\.\d+\s*', '', rest)
        return rest.strip()

    # Fallback: remove known parts
    remaining = text
    if fatura_no:
        remaining = re.sub(re.escape(fatura_no), '', remaining, flags=re.IGNORECASE)
    if tarih:
        remaining = remaining.replace(tarih, '')
    remaining = re.sub(r'^[\s\-/.,;:]+|[\s\-/.,;:]+$', '', remaining.strip())
    return remaining[:80] if remaining else text[:60]


def _read_muavin_sheet(sheet, cb_log=None) -> list[dict]:
    """Tek bir muavin sheet'inden ham satırları okur."""
    raw       = []
    hdr_row, col_map = _detect_header(sheet)

    if cb_log:
        cb_log(f'  "{sheet.title}" — hdr:{hdr_row} cols:{col_map}', 'info')

    # BORÇ kolonu zorunlu
    if 'borc' not in col_map:
        if cb_log:
            cb_log(f'  "{sheet.title}" — BORÇ kolonu bulunamadı, atlanıyor.', 'warn')
        return raw

    kdv_rate   = ''
    data_start = hdr_row + 1

    for ri, row in enumerate(sheet.iter_rows(values_only=True), 1):
        row_list = list(row)

        # Başlık öncesi: sadece KDV bölüm başlıklarını yakala
        if ri <= hdr_row:
            rate = _kdv_section_rate(row_list)
            if rate:
                kdv_rate = rate
            continue

        if not any(c for c in row_list if c is not None):
            continue

        # KDV bölüm başlığı mı?
        rate = _kdv_section_rate(row_list)
        if rate is not False:
            if rate:
                kdv_rate = rate
            continue

        # Atlanan satır mı?
        if _is_skip_row(row_list):
            continue

        # BORÇ al
        borc_val = row_list[col_map['borc'] - 1] if col_map['borc'] <= len(row_list) else None
        borc     = _to_decimal(borc_val)
        if borc <= 0:
            continue

        # ── E-fatura numarasını TÜM sütunlarda ara ──────────────────────────
        fatura_no = ''
        for val in row_list:
            if val is None:
                continue
            fn = _find_einvoice(val)
            if fn:
                fatura_no = fn
                break

        # Evrak/fiş no kolonundan fallback (e-fatura değilse bile kullan)
        evrak_val = ''
        if not fatura_no and 'evrak_no' in col_map and col_map['evrak_no'] <= len(row_list):
            ev = row_list[col_map['evrak_no'] - 1]
            if ev:
                evrak_str = str(ev).strip()
                # Sadece anlamlı referansları kullan (harf içeriyorsa veya uzunsa)
                if re.search(r'[A-ZĞÜŞİÖÇ]', evrak_str.upper()) or len(evrak_str) >= 8:
                    evrak_val = _normalize_fatura_no(evrak_str)

        if not fatura_no and evrak_val:
            fatura_no = evrak_val

        # ── Açıklama alanını önce oku (tarih için birincil kaynak) ──────────
        aciklama = ''
        if 'aciklama' in col_map and col_map['aciklama'] <= len(row_list):
            aciklama = str(row_list[col_map['aciklama'] - 1] or '')

        # ── Tarih: 1) açıklama  2) evrak_tarihi  3) tarih sütunu ────────────
        tarih_str, ay, yil = '', None, None

        # Önce açıklamadaki tarih
        if aciklama:
            dm = re.search(r'\d{2}[./]\d{2}[./]\d{4}', aciklama)
            if dm:
                tarih_str, ay, yil = _parse_tarih(dm.group())

        # Evrak tarihi sütunu
        if not tarih_str and 'evrak_tarihi' in col_map and col_map['evrak_tarihi'] <= len(row_list):
            tarih_str, ay, yil = _parse_tarih(row_list[col_map['evrak_tarihi'] - 1])

        # Genel tarih sütunu (son çare)
        if not tarih_str and 'tarih' in col_map and col_map['tarih'] <= len(row_list):
            tarih_str, ay, yil = _parse_tarih(row_list[col_map['tarih'] - 1])

        # ── Firma adı ─────────────────────────────────────────────────────────
        firma = ''
        if 'firma' in col_map and col_map['firma'] <= len(row_list):
            firma = str(row_list[col_map['firma'] - 1] or '').strip()
        if not firma and aciklama:
            firma = _firma_from_aciklama(aciklama, fatura_no, tarih_str)

        # ── VKN ───────────────────────────────────────────────────────────────
        vkn = ''
        if 'vkn' in col_map and col_map['vkn'] <= len(row_list):
            vkn = str(row_list[col_map['vkn'] - 1] or '').strip()

        # En az tarih veya fatura_no olması lazım
        if not tarih_str and not fatura_no:
            continue

        raw.append({
            'fatura_no': fatura_no,
            'tarih':     tarih_str,
            'firma':     firma,
            'vkn':       vkn,
            'aciklama':  aciklama,
            'borc':      borc,
            'kdv_orani': kdv_rate,
            'ay':        ay,
            'yil':       yil,
            'satir_no':  ri,
        })

    return raw


def _aggregate_by_fatura(raw_rows: list[dict]) -> list[dict]:
    """
    Aynı fatura_no'ya sahip satırları birleştirir, KDV toplamlarını toplar.
    Fatura no'su olmayan satırlar için sentetik anahtar kullanılır.
    """
    grouped: dict[str, dict] = {}
    no_fno_idx = 0

    for row in raw_rows:
        key = row['fatura_no'] or ''
        if not key:
            # E-fatura numarası yoksa: tarih + firma (benzersiz tutmak için)
            key = f'_NOFNO_{row["tarih"]}_{row["firma"][:20]}_{no_fno_idx}'
            no_fno_idx += 1

        if key not in grouped:
            grouped[key] = {
                'fatura_no':    row['fatura_no'],
                'tarih':        row['tarih'],
                'firma':        row['firma'],
                'vkn':          row['vkn'],
                'toplam_kdv':   Decimal('0'),
                'kdv_oranlari': set(),
                'ay':           row['ay'],
                'yil':          row['yil'],
                'ham_satirlar': [],
            }
        g = grouped[key]
        g['toplam_kdv']   += row['borc']
        g['ham_satirlar'].append(row)
        if row['kdv_orani']:
            g['kdv_oranlari'].add(row['kdv_orani'])
        # İlk değer yoksa güncelle
        if not g['firma'] and row['firma']:
            g['firma'] = row['firma']
        if not g['vkn'] and row['vkn']:
            g['vkn'] = row['vkn']

    return list(grouped.values())


def read_muavin(path: str, cb_log=None) -> list[dict]:
    """
    191 Muavin Excel dosyasını okur.
    Her formata uyumlu evrensel okuyucu.
    Aynı fatura_no'nun birden fazla KDV satırını toplar.
    Returns: aggregated list[dict]
    """
    wb = _open_wb(path)
    all_raw = []

    for sheet in wb.worksheets:
        if cb_log:
            cb_log(f'Muavin sheet okunuyor: "{sheet.title}"', 'info')
        raw = _read_muavin_sheet(sheet, cb_log=cb_log)
        all_raw.extend(raw)

    records = _aggregate_by_fatura(all_raw)

    if cb_log:
        real = sum(1 for r in records if not r['fatura_no'].startswith('_NOFNO_'))
        cb_log(f'Muavin: {len(all_raw)} ham satir, {real} benzersiz fatura (e-fatura), '
               f'{len(records) - real} eFatura numarasiz.', 'ok')

    return records


# ══════════════════════════════════════════════════════════════════════════════
# IKDVL OKUYUCU
# ══════════════════════════════════════════════════════════════════════════════

def _sheet_is_monthly(name: str) -> bool:
    n = _norm_header(name)
    for ay_name in _AYLAR:
        if _norm_header(ay_name) in n:
            return True
    if re.search(r'(?<!\d)(0[1-9]|1[0-2])(?!\d)', n):
        return True
    if re.match(r'^20\d{2}(0[1-9]|1[0-2])$', n) or re.match(r'^(0[1-9]|1[0-2])20\d{2}$', n):
        return True
    return False


def _detect_ikdvl_header(sheet) -> tuple[int, dict]:
    """
    IKDVL sheet'inde header satırını ve sütun indekslerini (0-tabanlı) tespit et.
    Returns: (data_start_row_1indexed, col_idx_dict)
    """
    # Varsayılan (xlsx formatı): col indeksleri 0-tabanlı
    best = {
        'tarih': 2, 'seri': 3, 'fatura_no': 4, 'tip': 5,
        'satici': 6, 'vkn': 7, 'matrah': 10,
        'hesaplanan_kdv': 11, 'toplam_kdv': 14,
        'kdv_donemi': 16,
    }
    data_start = 5
    HDR_MAP = {
        'ALISF': 'tarih',          # 'Alış Faturasının Tarihi'
        'SATICI': 'satici',
        'VERGIKI': 'vkn',
        "KDV'SI": 'hesaplanan_kdv',
        'HESAPLANAN': 'hesaplanan_kdv',
        'TOPLAMIND': 'toplam_kdv',
        'TOPLAMIND': 'toplam_kdv',
        'KDVD': 'kdv_donemi',
        'BELGENI': 'kdv_donemi',
    }

    for ri, row in enumerate(sheet.iter_rows(values_only=True, min_row=1, max_row=15), 1):
        non_null = sum(1 for c in row if c is not None)
        if non_null < 5:
            continue
        # Bu satır header mı? "sıra", "fatura", "satıcı" gibi kelimeler aranır
        row_text = _tr_upper(' '.join(str(c or '') for c in row))
        if 'FATURA' not in row_text and 'SATICI' not in row_text:
            data_start = ri + 1
            continue
        # Header satırı — sütunları tespit et
        data_start = ri + 1
        for ci, val in enumerate(row):
            if val is None:
                continue
            v = _norm_header(val)
            # Fatura sıra no
            if 'FATURASIRANO' in v or 'SIRANOSU' in v or ('FATURA' in v and 'SIRA' in v):
                best['fatura_no'] = ci
            # Fatura tarihi
            elif 'ALISF' in v and 'TARIH' in v:
                best['tarih'] = ci
            elif ci == best['tarih']:
                pass
            # Satıcı
            elif 'SATICI' in v and 'ADI' in v:
                best['satici'] = ci
            # VKN
            elif 'VERGIKIM' in v or 'VERGIKIMLIK' in v:
                best['vkn'] = ci
            # Hesaplanan KDV / KDV'si
            elif ("KDV" in v and "SI" in v and "TOPLAM" not in v) or \
                 ("HESAPLANAN" in v and "KDV" in v):
                best['hesaplanan_kdv'] = ci
            # Tevkifatlı (indirilecek)
            elif 'TEVKIFATLI' in v:
                best['tevkifatli'] = ci
            # 2 No'lu
            elif '2NOLU' in v or 'NOLUBEY' in v:
                best['nolu2'] = ci
            # Toplam İndirilen KDV
            elif 'TOPLAMIND' in v and 'KDV' in v:
                best['toplam_kdv'] = ci
            # Dönem
            elif 'KDVD' in v or ('BELGEN' in v and 'INDIRIM' in v) or 'DONEM' in v:
                best['kdv_donemi'] = ci
        break

    return data_start, best


def _read_ikdvl_sheet(sheet, cb_log=None) -> list[dict]:
    results = []
    data_start, ci = _detect_ikdvl_header(sheet)

    def _get(row, key, default=None):
        idx = ci.get(key, -1)
        if idx < 0 or idx >= len(row):
            return default
        return row[idx]

    for ri, row in enumerate(sheet.iter_rows(values_only=True), 1):
        if ri < data_start:
            continue
        if not any(c for c in row if c is not None):
            continue

        fatura_no = _normalize_fatura_no(_get(row, 'fatura_no'))
        if not fatura_no:
            continue

        # Toplam satırlarını atla
        row_text = _tr_upper(' '.join(str(c or '') for c in row))
        if 'TOPLAM' in row_text and len(fatura_no) < 5:
            continue

        tarih_str, ay, yil = _parse_tarih(_get(row, 'tarih'))
        toplam_kdv = _to_decimal(_get(row, 'toplam_kdv'))
        if toplam_kdv == 0:
            continue

        results.append({
            'sira':          _get(row, 'sira', ''),
            'tarih':         tarih_str,
            'seri':          str(_get(row, 'seri') or ''),
            'fatura_no':     fatura_no,
            'tip':           str(_get(row, 'tip') or ''),
            'satici':        str(_get(row, 'satici') or ''),
            'vkn':           str(_get(row, 'vkn') or ''),
            'matrah':        _to_decimal(_get(row, 'matrah')),
            'toplam_kdv':    toplam_kdv,
            'kdv_donemi':    str(_get(row, 'kdv_donemi') or ''),
            'satir_no':      ri,
            'sheet_name':    sheet.title,
            'ay':            ay,
            'yil':           yil,
        })

    return results


def read_ikdvl(path: str, cb_log=None) -> tuple[list[dict], bool]:
    """
    IKDVL Excel dosyasını okur.
    Returns: (rows, has_monthly_sheets)
    """
    wb      = _open_wb(path)
    monthly = [n for n in wb.sheetnames if _sheet_is_monthly(n)]

    if monthly:
        if cb_log:
            cb_log(f'Aylik sayfalar: {", ".join(monthly)}', 'info')
        results = []
        for name in monthly:
            rows = _read_ikdvl_sheet(wb[name], cb_log)
            if cb_log and rows:
                cb_log(f'  {name}: {len(rows)} kayit', 'info')
            results.extend(rows)
        if cb_log:
            cb_log(f'IKDVL toplam: {len(results)} kayit.', 'ok')
        return results, True

    sheet = None
    for name in wb.sheetnames:
        if 'indirilecek' in name.lower() or 'kdv' in name.lower():
            sheet = wb[name]
            break
    if sheet is None:
        sheet = wb.worksheets[0]

    if cb_log:
        cb_log(f'IKDVL sheet: "{sheet.title}"', 'info')
    results = _read_ikdvl_sheet(sheet, cb_log)
    if cb_log:
        cb_log(f'IKDVL: {len(results)} kayit.', 'ok')
    return results, False


# ══════════════════════════════════════════════════════════════════════════════
# KARŞILAŞTIRMA
# ══════════════════════════════════════════════════════════════════════════════

def _compare_group(ikdvl_rows: list, muavin_records: list, cb_log=None):
    """Bir grup (ay ya da tüm) IKDVL satırını muavin kayıtlarıyla karşılaştırır."""
    def log(msg, tag='info'):
        if cb_log: cb_log(msg, tag)

    # Muavin sözlüğü: fatura_no → aggregated record
    muavin_dict: dict[str, dict] = {}
    for mr in muavin_records:
        fn = mr['fatura_no']
        if fn and not fn.startswith('_NOFNO_'):
            muavin_dict[fn] = mr

    matched_keys = set()
    eslesen      = []
    ikdvl_fazla  = []
    tutar_farki  = []

    for ir in ikdvl_rows:
        fno = ir['fatura_no']
        mr  = muavin_dict.get(fno)

        # Kısmi eşleşme (fatura no kısa versiyonu veya uzun versiyonu)
        if not mr and len(fno) > 8:
            for key in muavin_dict:
                if fno in key or key in fno:
                    mr = muavin_dict[key]
                    break

        if not mr:
            ikdvl_fazla.append(ir)
            log(f'IKDVL Fazla: {fno}  ({ir["satici"]})  {ir["toplam_kdv"]} TL', 'err')
        else:
            matched_keys.add(mr['fatura_no'])
            muavin_kdv = mr['toplam_kdv'].quantize(Decimal('0.01'))
            ikdvl_kdv  = ir['toplam_kdv'].quantize(Decimal('0.01'))
            fark       = abs(muavin_kdv - ikdvl_kdv)

            if fark > Decimal('0.05'):
                tutar_farki.append({
                    'ikdvl':        ir,
                    'muavin':       mr,
                    'muavin_tutar': muavin_kdv,
                    'fark':         fark,
                })
                log(f'Tutar Farki: {fno}  IKDVL={ikdvl_kdv} Muavin={muavin_kdv}  '
                    f'Fark={fark}  ({len(mr["ham_satirlar"])} muavin satiri)', 'warn')
            else:
                eslesen.append({'ikdvl': ir, 'muavin': mr, 'muavin_tutar': muavin_kdv})
                log(f'Eslesti: {fno}  {ikdvl_kdv} TL  '
                    f'({len(mr["ham_satirlar"])} muavin satiri)', 'ok')

    # Muavinde olup IKDVL'de olmayanlar (e-fatura olanlar)
    muavin_fazla = [
        mr for mr in muavin_records
        if mr['fatura_no']
        and not mr['fatura_no'].startswith('_NOFNO_')
        and mr['fatura_no'] not in matched_keys
        and mr['toplam_kdv'] > 0
    ]

    for mr in muavin_fazla:
        log(f'Muavin Fazla: {mr["fatura_no"]}  {mr["firma"][:40]}  {mr["toplam_kdv"]} TL', 'warn')

    return eslesen, ikdvl_fazla, tutar_farki, muavin_fazla


def compare(
    ikdvl_path: str,
    muavin_path: str,
    stop_flag=None,
    cb_log=None,
    cb_progress=None,
    cb_stats=None,
    cb_done=None,
    max_muavin_rows=None,
):
    def log(msg, tag='info'):
        if cb_log: cb_log(msg, tag)

    def progress(pct):
        if cb_progress: cb_progress(pct)

    log('=' * 50, 'head')
    log('  KARSILASTIRMA BASLIYOR', 'head')
    log('=' * 50, 'head')

    log('IKDVL dosyasi okunuyor...', 'info')
    try:
        ikdvl_rows, has_monthly = read_ikdvl(ikdvl_path, cb_log=cb_log)
    except Exception as e:
        log(f'IKDVL okunamadi: {e}', 'err')
        if cb_done: cb_done(0, 0, 0, 0)
        return None

    progress(20)
    if stop_flag and stop_flag.is_set():
        return None

    log('Muavin dosyasi okunuyor...', 'info')
    try:
        muavin_records = read_muavin(muavin_path, cb_log=cb_log)
    except Exception as e:
        log(f'Muavin okunamadi: {e}', 'err')
        if cb_done: cb_done(0, 0, 0, 0)
        return None

    progress(40)
    if stop_flag and stop_flag.is_set():
        return None

    # Trial satır limiti — muavin okunduktan hemen sonra, karşılaştırmadan önce
    _trial_truncated = False
    if max_muavin_rows is not None:
        total_muavin_rows = sum(len(rec.get('ham_satirlar', [])) for rec in muavin_records)
        if total_muavin_rows > max_muavin_rows:
            # Kota aşıldı — fatura sınırında kes, kalan kota kadar işle
            truncated = []
            row_count = 0
            for rec in muavin_records:
                n = len(rec.get('ham_satirlar', []))
                if row_count + n > max_muavin_rows:
                    break
                truncated.append(rec)
                row_count += n
            log(f'⚠️  Trial kotası: {row_count}/{total_muavin_rows} muavin satırı işlenecek '
                f'(kota: {max_muavin_rows})', 'warn')
            muavin_records = truncated
            _trial_truncated = True

    log('Karsilastirma yapiliyor...', 'info')

    if has_monthly:
        # Aylik karsilastirma: her IKDVL ayini sadece o ayin muavin satirlariyla karsilastir
        log('Aylik karsilastirma modu aktif.', 'info')

        # IKDVL'yi aya gore grupla
        ikdvl_by_month: dict[tuple, list] = defaultdict(list)
        for ir in ikdvl_rows:
            ikdvl_by_month[(ir.get('yil'), ir.get('ay'))].append(ir)

        # Muavini aya gore grupla
        muavin_by_month: dict[tuple, list] = defaultdict(list)
        for mr in muavin_records:
            muavin_by_month[(mr.get('yil'), mr.get('ay'))].append(mr)

        all_eslesen     = []
        all_ikdvl_fazla = []
        all_tutar_farki = []
        all_muavin_fazla = []

        months = sorted(ikdvl_by_month.keys(), key=lambda k: (k[0] or 0, k[1] or 0))
        for i, month_key in enumerate(months):
            yil, ay = month_key
            log(f'-- {ay}/{yil} --', 'head')
            m_ikdvl  = ikdvl_by_month[month_key]
            m_muavin = muavin_by_month.get(month_key, [])

            if not m_muavin:
                log(f'  Muavinde {ay}/{yil} donemi bulunamadi!', 'warn')

            es, if_, tf, mf = _compare_group(m_ikdvl, m_muavin, cb_log)
            all_eslesen.extend(es)
            all_ikdvl_fazla.extend(if_)
            all_tutar_farki.extend(tf)
            all_muavin_fazla.extend(mf)

            progress(40 + int(50 * (i + 1) / len(months)))

        # IKDVL'de olmayan aylardaki muavin kayıtları da muavin_fazla'ya ekle
        for month_key, m_muavin in muavin_by_month.items():
            if month_key in ikdvl_by_month:
                continue  # zaten karşılaştırıldı
            for mr in m_muavin:
                if (mr['fatura_no']
                        and not mr['fatura_no'].startswith('_NOFNO_')
                        and mr['toplam_kdv'] > 0):
                    all_muavin_fazla.append(mr)
                    log(f'Muavin Fazla (eslesmeyen ay): {mr["fatura_no"]}  '
                        f'{mr["firma"][:30]}  {mr["toplam_kdv"]} TL', 'warn')
    else:
        # Genel karsilastirma
        all_eslesen, all_ikdvl_fazla, all_tutar_farki, all_muavin_fazla = \
            _compare_group(ikdvl_rows, muavin_records, cb_log)
        progress(90)

    progress(95)
    log('-' * 50, 'head')
    log(f'  IKDVL Toplam     : {len(ikdvl_rows)}',       'info')
    log(f'  Muavin Benzersiz : {len(muavin_records)}',   'info')
    log(f'  Eslesen          : {len(all_eslesen)}',      'ok')
    log(f'  IKDVL Fazlasi    : {len(all_ikdvl_fazla)}', 'err')
    log(f'  Muavin Fazlasi   : {len(all_muavin_fazla)}','warn')
    log(f'  Tutar Farki      : {len(all_tutar_farki)}', 'warn')
    log('=' * 50, 'head')
    progress(100)

    if cb_stats:
        cb_stats(len(ikdvl_rows), len(muavin_records),
                 len(all_eslesen),
                 len(all_ikdvl_fazla) + len(all_muavin_fazla) + len(all_tutar_farki))

    if cb_done:
        cb_done(len(all_eslesen), len(all_ikdvl_fazla),
                len(all_muavin_fazla), len(all_tutar_farki))

    if _trial_truncated:
        log('─' * 50, 'head')
        log('⛔  Trial kotanız doldu. Lisans alarak tüm muavini işleyebilirsiniz.', 'err')

    return {
        'ikdvl_path':    ikdvl_path,
        'muavin_path':   muavin_path,
        'ikdvl_rows':    ikdvl_rows,
        'muavin_records': muavin_records,
        # Geriye dönük uyumluluk için
        'muavin_rows':   [r for rec in muavin_records for r in rec['ham_satirlar']],
        'eslesen':       all_eslesen,
        'ikdvl_fazla':   all_ikdvl_fazla,
        'muavin_fazla':  all_muavin_fazla,
        'tutar_farki':   all_tutar_farki,
    }


# ══════════════════════════════════════════════════════════════════════════════
# DÜZELTME MOTORU
# ══════════════════════════════════════════════════════════════════════════════

_GREEN_HEX       = 'C8F7C5'
_YELLOW_HEX      = 'FFFDE7'
_PURPLE_HEX      = 'E1BEE7'
_PURPLE_HEAD_HEX = '6A1B9A'

def _make_thin_border():
    from openpyxl.styles import Border, Side
    s = Side(style='thin')
    return Border(left=s, right=s, top=s, bottom=s)

_THIN_BORDER = _make_thin_border()


def _find_ikdvl_cols(ws) -> dict:
    """IKDVL sheet'inde sütun pozisyonlarını dinamik tespit et."""
    found = {
        'tarih': 3, 'fatura_no': 5, 'satici': 7, 'vkn': 8,
        'matrah': 11, 'hesaplanan_kdv': 12, 'indirilecek': 13,
        'tevkifat': 14, 'toplam_kdv': 15, 'kdv_donemi': 17, 'last_col': 18,
        'format': 'our',  # 'our' (hesaplanan=L=12) veya 'other' (hesaplanan=K=11)
    }
    real_last = 1
    for row in ws.iter_rows(min_row=1, max_row=8):
        # Başlık satırı değilse (tek hücreli satır = sayfa başlığı) atla
        non_null = sum(1 for cell in row if cell.value is not None)
        if non_null < 3:
            continue
        for cell in row:
            if cell.value is None:
                continue
            v  = _norm_header(cell.value)
            vc = v.replace("'", '')   # apostrof temizlenmiş versiyon
            c  = cell.column
            if ('TOPLAMINDIRILENKDV' in v) or ('TOPLAMKDV' in v) or \
               ('TOPLAMINDIRILEN' in v) or (v == 'TOPLAMINDIRILENK'):
                found['toplam_kdv'] = c
            elif ('HESAPLANAN' in v and 'KDV' in v) or \
                 vc in ('HESAPLANANKDV', 'KDVSI', 'KDVTUTARI'):
                found['hesaplanan_kdv'] = c
                found['format'] = 'other' if c == 11 else 'our'
            elif ('2NOLUBEY' in v) or ('NOLUBEYANNAME' in v) or \
                 ('ODENEN' in v and 'KDV' in v and 'TEVKIFAT' not in v):
                found['tevkifat'] = c
            elif 'INDIRILECEK' in v and 'KDV' in v and 'TOPLAM' not in v \
                 and 'LISTE' not in v:
                found['indirilecek'] = c
            elif v in ('TARIH', 'FATURAARIHI', 'EVRAKTARIHI'):
                found['tarih'] = c
            elif 'FATURAON' in v:
                found['fatura_no'] = c
            elif v in ('SATICI', 'UNVAN', 'FIRMAADI'):
                found['satici'] = c
            if cell.value is not None and c > real_last:
                real_last = c
    if real_last >= 5:
        found['last_col'] = real_last

    # Format'a göre bağımlı sütunları düzelt
    hk = found['hesaplanan_kdv']
    if found['format'] == 'our':
        # L=hesaplanan, M=indirilecek(L-N), N=tevkifat(sabit), O=toplam(M+N)
        if found['indirilecek'] == 13: found['indirilecek'] = hk + 1
        if found['tevkifat']    == 14: found['tevkifat']    = hk + 2
        if found['toplam_kdv']  == 15: found['toplam_kdv']  = hk + 3
    else:
        # K=hesaplanan, L=indirilecek(K-M), M=tevkifat(sabit), N=toplam(L+M)
        found['indirilecek'] = hk + 1
        found['tevkifat']    = hk + 2
        found['toplam_kdv']  = hk + 3

    return found


def _cell_has_formatting(cell) -> bool:
    """Hücrenin orijinal özel biçimi (font rengi, arka plan, kalın vb.) var mı?"""
    f = cell.font
    if f:
        if f.bold:                                          return True
        if f.color and f.color.type == 'rgb' and f.color.rgb not in ('FF000000', '00000000'): return True
    fl = cell.fill
    if fl and fl.fill_type == 'solid':
        fg = fl.fgColor
        if fg and fg.type == 'rgb' and fg.rgb not in ('FFFFFFFF', '00000000', 'FF000000'): return True
    return False


def _detect_base_font(ws, data_start, last_col):
    """Veri satırlarından kullanılan temel font adını ve boyutunu tespit et."""
    from collections import Counter
    counter = Counter()
    for r in range(data_start, min(data_start + 20, ws.max_row + 1)):
        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            if cell.value is not None and cell.font:
                fn = cell.font.name or 'Calibri'
                fs = cell.font.size or 10
                counter[(fn, fs)] += 1
    if counter:
        return counter.most_common(1)[0][0]  # (font_name, font_size)
    return ('Calibri', 10)


def _apply_yeni_formatting(ws, cols):
    """
    'Yeni Düzenlenmiş Liste Al' modunda çıktıya biçimlendirme uygular:
    - Tüm veri satırlarına ince kenarlık (sütun 1 → last_col)
    - Biçimi olan hücreler aynen korunur (font rengi, arka plan vs.)
    - Biçimi olmayan hücrelerde dosyanın kendi fontu/boyu kullanılır
    - Sayısal sütunlara: Virgül stili (#,##0.00)
    """
    from openpyxl.styles import Font, Alignment

    CENTER_WRAP = Alignment(horizontal='center', vertical='center', wrap_text=True)
    CENTER_MID  = Alignment(horizontal='center', vertical='center')
    LEFT_MID    = Alignment(horizontal='left',   vertical='center')
    COMMA_FMT   = '#,##0.00'

    last_col = cols.get('last_col', 20)
    numeric_cols = {cols.get(k) for k in
                    ('hesaplanan_kdv', 'indirilecek', 'tevkifat', 'toplam_kdv', 'matrah')
                    if cols.get(k)}
    center_cols = {cols.get(k) for k in ('tarih', 'kdv_donemi') if cols.get(k)}
    center_cols.add(1)  # sıra no

    # Son dolu satırı bul
    last_row = 1
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None:
                last_row = cell.row

    # Başlık bitiş satırını bul
    fno_col    = cols.get('fatura_no', 4)
    header_end = 4
    for r in range(1, min(20, last_row + 1)):
        val = ws.cell(row=r, column=fno_col).value
        if val and _EFATURA_RE.search(str(val)):
            header_end = r - 1
            break

    # Dosyanın temel fontunu tespit et (veri satırlarından)
    base_font_name, base_font_size = _detect_base_font(ws, header_end + 1, last_col)
    HDR_FONT = Font(name=base_font_name, size=base_font_size, bold=True)
    DAT_FONT = Font(name=base_font_name, size=base_font_size)

    for r in range(1, last_row + 1):
        has_data = any(
            ws.cell(row=r, column=c).value is not None
            for c in range(1, last_col + 1)
        )
        if not has_data:
            continue

        is_header = (r <= header_end)

        for c in range(1, last_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _THIN_BORDER

            has_fmt = _cell_has_formatting(cell)

            if is_header:
                if not has_fmt:
                    cell.font = HDR_FONT
                # Hizalama: biçimi olan hücrede wrap_text aç, yoksa tamamen uygula
                if not cell.alignment or cell.alignment.horizontal in (None, 'general'):
                    cell.alignment = CENTER_WRAP
                else:
                    cell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical or 'center',
                        wrap_text=True,
                    )
            else:
                if not has_fmt:
                    cell.font = DAT_FONT
                if c in center_cols:
                    cell.alignment = CENTER_MID
                else:
                    cell.alignment = LEFT_MID
                if c in numeric_cols and isinstance(cell.value, (int, float)):
                    cell.number_format = COMMA_FMT


def _find_toplam_rows(ws, from_row=5) -> list[int]:
    rows = []
    for row in ws.iter_rows(min_row=from_row):
        for cell in row:
            v = _tr_upper(str(cell.value or '')).strip()
            if v in ('TOPLAM', 'GENEL TOPLAM', 'ARA TOPLAM', 'GRAND TOTAL', 'TOP'):
                rows.append(cell.row)
                break
    return rows


def _duzelt_ws(ws, tf_list, if_list, mf_list, fills, fonts, CENTER, log, log_entries,
               mode='mevcut'):
    """Tek bir IKDVL sheet'ini düzelt.
    mode='mevcut' → renkli, IKDVL fazlası alta mor bölüm
    mode='yeni'   → renksiz, IKDVL fazlası sadece silindi (alt bölüm yok)
    """
    GREEN_FILL, YELLOW_FILL, PURPLE_FILL, PHEAD_FILL = fills
    PHEAD_FONT = fonts[0]

    cols     = _find_ikdvl_cols(ws)
    last_col = cols['last_col']
    kdv_col  = cols['toplam_kdv']
    log(f'  Sheet "{ws.title}": tf={len(tf_list)} if={len(if_list)} mf={len(mf_list)} '
        f'KDVkol={kdv_col} sonkol={last_col}', 'info')

    # Normal veri satırlarından tipik satır yüksekliği al (yeni satırlar için)
    typical_h = None
    for row in ws.iter_rows(min_row=5, max_row=20):
        rh = ws.row_dimensions[row[0].row].height
        if rh and rh > 5:
            typical_h = rh
            break

    hk_col  = cols['hesaplanan_kdv']   # L veya K
    ind_col = cols['indirilecek']      # M veya L (hesaplanan - tevkifat)
    tev_col = cols['tevkifat']         # N veya M (sabit kesinti)
    top_col = cols['toplam_kdv']       # O veya N (indirilecek + tevkifat)

    # 1. Tutar farklarını güncelle (yeşil)
    for pair in tf_list:
        ir    = pair['ikdvl']
        eski  = float(ir['toplam_kdv'])
        yeni  = round(float(pair['muavin_tutar']), 2)
        satir = ir['satir_no']

        # Hesaplanan KDV sütununu güncelle
        ws.cell(row=satir, column=hk_col).value = yeni

        # Bağımlı sütunları yeniden hesapla
        tev = _to_decimal(ws.cell(row=satir, column=tev_col).value)
        ind = Decimal(str(yeni)) - tev          # M = L - N  veya  L = K - M
        top = ind + tev                          # O = M + N  veya  N = L + M
        ws.cell(row=satir, column=ind_col).value = float(round(ind, 2))
        ws.cell(row=satir, column=top_col).value = float(round(top, 2))

        # Tarih hücresini DD.MM.YYYY yap
        tc = ws.cell(row=satir, column=cols['tarih'])
        ts, _, _ = _parse_tarih(tc.value)
        if ts: tc.value = ts

        for col in range(1, last_col + 1):
            c = ws.cell(row=satir, column=col)
            c.fill = GREEN_FILL
            c.border = _THIN_BORDER
        log(f'  Guncellendi: {ir["fatura_no"]}  {eski:.2f} -> {yeni:.2f} TL', 'ok')
        log_entries.append({'islem': 'TUTAR_GUNCELLENDI', 'fatura': ir['fatura_no'],
                            'tarih': ir['tarih'], 'eski': f'{eski:.2f}',
                            'yeni': f'{yeni:.2f}', 'satir_no': satir,
                            'sheet': ws.title})

    # 2. IKDVL fazlası — önce veri topla, sonra sil (büyük satır no'dan küçüğe)
    if_data = []
    for ir in if_list:
        satir    = ir['satir_no']
        row_vals = [ws.cell(row=satir, column=c).value for c in range(1, last_col + 1)]
        if_data.append((ir, row_vals))

    for ir, _ in sorted(if_data, key=lambda x: x[0]['satir_no'], reverse=True):
        ws.delete_rows(ir['satir_no'])
        log_entries.append({'islem': 'SATIR_TASINAK', 'fatura': ir['fatura_no'],
                            'tarih': ir['tarih'], 'eski': f'{float(ir["toplam_kdv"]):.2f}',
                            'yeni': 'Alta tasindi', 'satir_no': ir['satir_no'],
                            'sheet': ws.title})

    # 3. Toplam satırını bul ve sil; ekleme noktasını belirle
    toplam_rows = _find_toplam_rows(ws, from_row=5)
    if toplam_rows:
        insert_start = min(toplam_rows)
        for tr in sorted(toplam_rows, reverse=True):
            ws.delete_rows(tr)
    else:
        last_row = 4
        for row in ws.iter_rows(min_row=5):
            for cell in row:
                if cell.value is not None:
                    last_row = cell.row
                    break
        insert_start = last_row + 2

    # Merge temizliği: insert_start ve sonrasındaki tüm merge aralıklarını kaldır.
    # delete_rows() openpyxl'de merge range'leri otomatik shift etmez;
    # stale merge'ler veri satırlarına otomatik uygulanır ve veri kaymasına yol açar.
    stale_merges = [
        mr for mr in list(ws.merged_cells.ranges)
        if mr.min_row >= insert_start
    ]
    for mr in stale_merges:
        ws.merged_cells.remove(mr)

    # 4. Muavin fazlası ekle (sarı / renksiz)
    insert_row = insert_start
    for mr in mf_list:
        # KDV dönemi: MMYYYY formatı (ör. "012025")
        ay_str  = f'{mr.get("ay") or 0:02d}'
        yil_str = str(mr.get("yil") or '')
        donem   = (ay_str + yil_str) if (mr.get("ay") and mr.get("yil")) else ''

        tarih_raw = mr.get('tarih', '')
        ts_tarih, _, _ = _parse_tarih(tarih_raw)
        tarih_fmt = ts_tarih or tarih_raw

        kdv_val = round(float(mr['toplam_kdv']), 2)
        tev_val = 0.0   # muavinden gelen satır için tevkifat bilinmiyor
        ind_val = round(kdv_val - tev_val, 2)
        top_val = round(ind_val + tev_val, 2)

        row_data = [None] * last_col
        def _set(key, val):
            ci = cols.get(key, 0)
            if 1 <= ci <= last_col:
                row_data[ci - 1] = val

        _set('tarih',          tarih_fmt)
        _set('fatura_no',      mr.get('fatura_no') or '')
        _set('satici',         mr.get('firma') or '')
        _set('vkn',            mr.get('vkn') or '')
        _set('hesaplanan_kdv', kdv_val)
        _set('indirilecek',    ind_val)
        _set('tevkifat',       tev_val if tev_val else None)
        _set('toplam_kdv',     top_val)
        _set('kdv_donemi',     donem)

        for ci, val in enumerate(row_data, start=1):
            cell = ws.cell(row=insert_row, column=ci)
            cell.value = val
            cell.fill   = YELLOW_FILL
            cell.border = _THIN_BORDER
        if typical_h:
            ws.row_dimensions[insert_row].height = typical_h

        log(f'  Eklendi: {mr.get("fatura_no","")}  {mr.get("firma","")[:30]}  '
            f'{kdv_val:.2f} TL  donem={donem}', 'ok')
        log_entries.append({'islem': 'SATIR_EKLENDI', 'fatura': mr.get('fatura_no', ''),
                            'tarih': tarih_fmt, 'eski': '-',
                            'yeni': f'{kdv_val:.2f}',
                            'satir_no': insert_row, 'sheet': ws.title})
        insert_row += 1

    if mf_list:
        insert_row += 1  # boşluk

    # 5. IKDVL fazlasını alta yapıştır — her iki modda da
    if if_data:
        hc = ws.cell(row=insert_row, column=1)
        hc.value     = 'INDIRILECEK KDV LISTESINDE OLUP, 191 MUAVINDE OLMAYAN FATURALAR'
        hc.fill      = PHEAD_FILL
        hc.font      = PHEAD_FONT
        hc.alignment = CENTER
        ws.merge_cells(start_row=insert_row, start_column=1,
                       end_row=insert_row, end_column=last_col)
        ws.row_dimensions[insert_row].height = 24
        insert_row += 1
        tarih_col = cols['tarih']
        for ir, row_vals in if_data:
            for ci, val in enumerate(row_vals, start=1):
                cell = ws.cell(row=insert_row, column=ci)
                if ci == tarih_col and val is not None:
                    ts, _, _ = _parse_tarih(val)
                    cell.value = ts if ts else str(val)
                else:
                    cell.value = val
                cell.fill = PURPLE_FILL
            if typical_h:
                ws.row_dimensions[insert_row].height = typical_h
            insert_row += 1

    # Yeni mod: tam biçimlendirme uygula
    if mode == 'yeni':
        _apply_yeni_formatting(ws, cols)


def duzelt(
    ikdvl_path: str,
    result: dict,
    output_path: str,
    cb_log=None,
    cb_progress=None,
    mode: str = 'mevcut',   # 'mevcut' veya 'yeni'
):
    """IKDVL dosyasını karşılaştırma sonucuna göre düzeltir ve kaydeder.
    mode='mevcut': renkli, IKDVL fazlası alta mor bölüm olarak taşınır
    mode='yeni'  : renksiz, sadece veriler düzeltilir/eklenir, fazlalar silinir
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment

    def log(msg, tag='info'):
        if cb_log: cb_log(msg, tag)

    def progress(pct):
        if cb_progress: cb_progress(pct)

    GREEN_FILL  = PatternFill('solid', fgColor=_GREEN_HEX)
    YELLOW_FILL = PatternFill('solid', fgColor=_YELLOW_HEX)
    PURPLE_FILL = PatternFill('solid', fgColor=_PURPLE_HEX)
    PHEAD_FILL  = PatternFill('solid', fgColor=_PURPLE_HEAD_HEX)
    PHEAD_FONT  = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    CENTER      = Alignment(horizontal='center', vertical='center', wrap_text=True)

    fills = (GREEN_FILL, YELLOW_FILL, PURPLE_FILL, PHEAD_FILL)
    fonts = (PHEAD_FONT,)

    log('=' * 50, 'head')
    log('  IKDVL DUZELTILIYOR', 'head')
    log('=' * 50, 'head')

    tutar_farki  = result.get('tutar_farki',  [])
    ikdvl_fazla  = result.get('ikdvl_fazla',  [])
    muavin_fazla = result.get('muavin_fazla', [])

    log(f'Tutar guncellecek: {len(tutar_farki)}', 'info')
    log(f'IKDVL tasinacak  : {len(ikdvl_fazla)}', 'info')
    log(f'Muavinden eklenck: {len(muavin_fazla)}', 'info')

    if not tutar_farki and not ikdvl_fazla and not muavin_fazla:
        log('Hic duzeltme gerekmedi.', 'ok')
        if cb_progress: cb_progress(100)
        return

    if str(ikdvl_path).lower().endswith('.xls'):
        wb = _xls_to_openpyxl_wb(ikdvl_path)
        log('XLS dosyasi XLSX formatina donusturuldu.', 'info')
    else:
        wb = openpyxl.load_workbook(ikdvl_path)
    log_entries = []

    # ── Aylık sheet yapısı var mı? ────────────────────────────────────────────
    monthly_names = [n for n in wb.sheetnames if _sheet_is_monthly(n)]

    if monthly_names:
        log(f'Aylik mod: {len(monthly_names)} sayfa', 'info')

        # (yil, ay) → sheet_name haritası (IKDVL kayıtlarından)
        month_to_sheet: dict[tuple, str] = {}
        for ir in result.get('ikdvl_rows', []):
            k = (ir.get('yil'), ir.get('ay'))
            if k not in month_to_sheet and ir.get('sheet_name'):
                month_to_sheet[k] = ir['sheet_name']

        # Tutar farkı sayfaya göre grupla
        tf_by_sheet: dict[str, list] = defaultdict(list)
        for pair in tutar_farki:
            tf_by_sheet[pair['ikdvl']['sheet_name']].append(pair)

        # IKDVL fazlası sayfaya göre grupla
        if_by_sheet: dict[str, list] = defaultdict(list)
        for ir in ikdvl_fazla:
            if_by_sheet[ir['sheet_name']].append(ir)

        # Muavin fazlası sayfaya göre grupla (ay/yıl → sheet)
        mf_by_sheet: dict[str, list] = defaultdict(list)
        unmapped_mf = []
        for mr in muavin_fazla:
            k  = (mr.get('yil'), mr.get('ay'))
            sn = month_to_sheet.get(k)
            if sn:
                mf_by_sheet[sn].append(mr)
            else:
                unmapped_mf.append(mr)

        # Eşleşmeyen muavin fazlalarını son sayfaya ekle
        if unmapped_mf:
            last_sheet = monthly_names[-1]
            mf_by_sheet[last_sheet].extend(unmapped_mf)
            log(f'{len(unmapped_mf)} muavin fazlasi ay eslesemedi, '
                f'"{last_sheet}" sayfasina eklendi.', 'warn')

        sheets_to_process = sorted(
            set(list(tf_by_sheet) + list(if_by_sheet) + list(mf_by_sheet)),
            key=lambda n: monthly_names.index(n) if n in monthly_names else 999
        )

        total = max(len(sheets_to_process), 1)
        for i, sheet_name in enumerate(sheets_to_process):
            if sheet_name not in wb.sheetnames:
                log(f'Sayfa bulunamadi: {sheet_name}', 'warn')
                continue
            ws = wb[sheet_name]
            _duzelt_ws(ws,
                       tf_by_sheet.get(sheet_name, []),
                       if_by_sheet.get(sheet_name, []),
                       mf_by_sheet.get(sheet_name, []),
                       fills, fonts, CENTER, log, log_entries, mode=mode)
            progress(20 + int(65 * (i + 1) / total))

    else:
        # Tek sayfa modu
        ws = None
        for name in wb.sheetnames:
            if 'indirilecek' in name.lower() or 'kdv' in name.lower():
                ws = wb[name]; break
        if ws is None:
            ws = wb.active
        _duzelt_ws(ws, tutar_farki, ikdvl_fazla, muavin_fazla,
                   fills, fonts, CENTER, log, log_entries, mode=mode)
        progress(85)

    # ── DEĞİŞİKLİK_LOG ────────────────────────────────────────────────────────
    if 'DEGISIKLIK_LOG' in wb.sheetnames:
        del wb['DEGISIKLIK_LOG']

    from openpyxl.utils import get_column_letter
    ws_log = wb.create_sheet('DEGISIKLIK_LOG')
    ws_log.sheet_view.showGridLines = False
    LOG_HFILL = PatternFill('solid', fgColor='0B1F3A')
    LOG_HFONT = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    log_hdrs  = ['Tarih/Saat', 'Sayfa', 'Fatura No', 'Fatura Tarihi', 'Islem', 'Eski', 'Yeni', 'Satir']
    log_ws    = [20, 14, 24, 14, 22, 14, 14, 8]
    for ci, (h, w) in enumerate(zip(log_hdrs, log_ws), 1):
        c = ws_log.cell(row=1, column=ci, value=h)
        c.fill = LOG_HFILL; c.font = LOG_HFONT
        c.alignment = CENTER
        ws_log.column_dimensions[get_column_letter(ci)].width = w

    now_str = datetime.now().strftime('%d.%m.%Y %H:%M:%S')
    for ri, ent in enumerate(log_entries, start=2):
        for ci, v in enumerate([now_str, ent.get('sheet', ''), ent['fatura'], ent['tarih'],
                                 ent['islem'], ent['eski'], ent['yeni'], ent['satir_no']], 1):
            ws_log.cell(row=ri, column=ci, value=v)
    ws_log.freeze_panes = 'A2'

    progress(95)
    wb.save(output_path)
    log(f'Kaydedildi: {output_path}', 'ok')
    log(f'Toplam {len(log_entries)} degisiklik ({len(monthly_names) if monthly_names else 1} sayfa).', 'info')
    log('=' * 50, 'head')
    progress(100)


# ══════════════════════════════════════════════════════════════════════════════
# RAPOR
# ══════════════════════════════════════════════════════════════════════════════

def create_report(result: dict, output_path: str, cb_log=None):
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    def log(msg, tag='info'):
        if cb_log: cb_log(msg, tag)

    GREEN_F  = PatternFill('solid', fgColor='C6EFCE')
    RED_F    = PatternFill('solid', fgColor='FFC7CE')
    ORANGE_F = PatternFill('solid', fgColor='FFE0B2')
    YELLOW_F = PatternFill('solid', fgColor='FFEB9C')
    HEAD_F   = PatternFill('solid', fgColor='0B1F3A')
    H_FONT   = Font(name='Segoe UI', bold=True, color='FFFFFF', size=10)
    N_FONT   = Font(name='Segoe UI', size=9)
    B_FONT   = Font(name='Segoe UI', bold=True, size=9)
    C_ALN    = Alignment(horizontal='center', vertical='center', wrap_text=True)
    L_ALN    = Alignment(horizontal='left',   vertical='center', wrap_text=True)
    thin     = Side(style='thin', color='D3D3D3')
    BRD      = Border(left=thin, right=thin, top=thin, bottom=thin)

    wb = openpyxl.Workbook()

    def mksheet(title):
        ws = wb.create_sheet(title)
        ws.sheet_view.showGridLines = False
        return ws

    def hdr_row(ws, headers, fills=None):
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.fill = HEAD_F; c.font = H_FONT; c.alignment = C_ALN; c.border = BRD
        ws.freeze_panes = 'A2'

    def data_row(ws, ri, vals, fill):
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = fill; c.font = N_FONT; c.alignment = L_ALN; c.border = BRD

    def auto_width(ws, n):
        for ci in range(1, n + 1):
            ml = 10
            for row in ws.iter_rows(min_col=ci, max_col=ci):
                for cell in row:
                    if cell.value:
                        ml = max(ml, len(str(cell.value)))
            ws.column_dimensions[get_column_letter(ci)].width = min(ml + 3, 50)

    # Özet
    ws0 = wb.active; ws0.title = 'Ozet'
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions['A'].width = 28; ws0.column_dimensions['B'].width = 20
    c = ws0.cell(row=1, column=1, value='191 MUAVIN KARSILASTIRMA RAPORU')
    c.fill = HEAD_F; c.font = H_FONT; c.alignment = C_ALN
    ws0.merge_cells('A1:B1'); ws0.row_dimensions[1].height = 26
    ozet_data = [
        ('Rapor Tarihi', datetime.now().strftime('%d.%m.%Y %H:%M')),
        ('IKDVL Kayit',  len(result['ikdvl_rows'])),
        ('Muavin Fatura', len(result.get('muavin_records', []))),
        ('Eslesen',       len(result['eslesen'])),
        ('IKDVL Fazlasi', len(result['ikdvl_fazla'])),
        ('Muavin Fazlasi',len(result['muavin_fazla'])),
        ('Tutar Farki',   len(result['tutar_farki'])),
    ]
    fills_o = [None, None, None, GREEN_F, RED_F, ORANGE_F, YELLOW_F]
    for i, (lbl, val) in enumerate(ozet_data, start=2):
        cl = ws0.cell(row=i, column=1, value=lbl)
        cv = ws0.cell(row=i, column=2, value=val)
        cl.font = B_FONT; cl.alignment = L_ALN; cl.border = BRD
        cv.font = N_FONT; cv.alignment = C_ALN; cv.border = BRD
        if fills_o[i-2]: cv.fill = fills_o[i-2]

    # Eşleşenler
    ws1 = mksheet('Eslenenler')
    hs1 = ['Fatura No', 'Tarih', 'Satici', 'VKN', 'IKDVL KDV', 'Muavin KDV', 'Fark']
    hdr_row(ws1, hs1)
    for ri, p in enumerate(result['eslesen'], 2):
        ir = p['ikdvl']; mr = p['muavin']
        mt = float(p.get('muavin_tutar', mr.get('toplam_kdv', 0)))
        data_row(ws1, ri, [ir['fatura_no'], ir['tarih'], ir['satici'], ir['vkn'],
                           float(ir['toplam_kdv']), mt, 0.0], GREEN_F)
    auto_width(ws1, len(hs1))

    # IKDVL Fazlası
    ws2 = mksheet('IKDVL Fazlasi')
    hs2 = ['Fatura No', 'Tarih', 'Satici', 'VKN', 'Matrah', 'Toplam KDV', 'Donem']
    hdr_row(ws2, hs2)
    for ri, ir in enumerate(result['ikdvl_fazla'], 2):
        data_row(ws2, ri, [ir['fatura_no'], ir['tarih'], ir['satici'], ir['vkn'],
                           float(ir['matrah']), float(ir['toplam_kdv']), ir['kdv_donemi']], RED_F)
    auto_width(ws2, len(hs2))

    # Muavin Fazlası
    ws3 = mksheet('Muavin Fazlasi')
    hs3 = ['Fatura No', 'Tarih', 'Firma', 'VKN', 'Toplam KDV', 'KDV Oranlari']
    hdr_row(ws3, hs3)
    for ri, mr in enumerate(result['muavin_fazla'], 2):
        oranlar = ', '.join(f'%{r}' for r in sorted(mr.get('kdv_oranlari', set())))
        data_row(ws3, ri, [mr['fatura_no'], mr['tarih'], mr['firma'], mr['vkn'],
                           float(mr['toplam_kdv']), oranlar], ORANGE_F)
    auto_width(ws3, len(hs3))

    # Tutar Farkları
    ws4 = mksheet('Tutar Farklari')
    hs4 = ['Fatura No', 'Tarih', 'Satici', 'IKDVL KDV', 'Muavin KDV', 'Fark']
    hdr_row(ws4, hs4)
    for ri, p in enumerate(result['tutar_farki'], 2):
        ir = p['ikdvl']; mr = p['muavin']
        mt = float(p.get('muavin_tutar', mr.get('toplam_kdv', 0)))
        data_row(ws4, ri, [ir['fatura_no'], ir['tarih'], ir['satici'],
                           float(ir['toplam_kdv']), mt, float(p['fark'])], YELLOW_F)
    auto_width(ws4, len(hs4))

    wb.save(output_path)
    log(f'Rapor kaydedildi: {output_path}', 'ok')
